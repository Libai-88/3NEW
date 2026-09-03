# -*- coding: utf-8 -*-
"""查看 7.26 文件 18-32 行的完整布局，理解重复行组织结构"""
import openpyxl

U = '/workspace/.uploads/'
path = U + 'cb7988e8-dc53-4617-9e33-71aa45000fcf_7.26配料测试汇总.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb['Sheet1']
print('max_row=', ws.max_row, 'max_col=', ws.max_column)
for r in range(14, 36):
    vals = [ws.cell(r, c).value for c in range(1, 21)]
    vals = ['--' if v is None else (round(v, 2) if isinstance(v, float) else v) for v in vals]
    print(f'row{r}:', vals)