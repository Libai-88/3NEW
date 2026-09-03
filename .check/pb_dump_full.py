# -*- coding: utf-8 -*-
"""配比方案：全sheet网格转储（含全部列，行<=80），供人工核对。用openpyxl data_only读缓存值。"""
import openpyxl
F = "/workspace/.uploads/b36f4809-d40f-491a-b9d7-2a9c4e359b1b_fbfc94091fa4955969e5d0fff7df0fd6_789507228604997242_m_3NX240913-6C--AI研发26.7.22配比方案.xlsx"
wb = openpyxl.load_workbook(F, data_only=True)
out = open('/tmp/pb_dump_full.txt', 'w')
for ws in wb.worksheets:
    out.write('=' * 120 + '\n')
    out.write('SHEET: %s  rows=%d cols=%d\n' % (ws.title, ws.max_row, ws.max_column))
    for r in range(1, min(ws.max_row, 80) + 1):
        vals = []
        for c in range(1, min(ws.max_column, 20) + 1):
            v = ws.cell(r, c).value
            if v is None:
                s = ''
            elif isinstance(v, float):
                s = ('%.4f' % v).rstrip('0').rstrip('.')
            else:
                s = str(v).replace('\n', '\\n')
            vals.append(s)
        line = 'r%-2d| %s' % (r, ' | '.join(vals))
        if line.strip(' |'):
            out.write(line + '\n')
out.close()
print('done')