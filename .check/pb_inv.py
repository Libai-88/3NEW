# -*- coding: utf-8 -*-
"""配比方案：全sheet 结构盘点 —— 确认每个sheet的样本列、是否含衍生列(百分比/1000KG)、每列合计"""
import openpyxl, unicodedata, re

FA = '/workspace/.uploads/b36f4809-d40f-491a-b9d7-2a9c4e359b1b_fbfc94091fa4955969e5d0fff7df0fd6_789507228604997242_m_3NX240913-6C--AI研发26.7.22配比方案.xlsx'


def norm(s):
    return unicodedata.normalize('NFKC', str(s)).replace('\n', '').replace(' ', '')


def isnum(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


wb = openpyxl.load_workbook(FA, data_only=True)
for ws in wb.worksheets:
    if ws.title == 'Sheet1':
        continue
    ncols = min(ws.max_column, 16)
    grid = [[ws.cell(r, c).value for c in range(1, ncols + 1)] for r in range(1, (ws.max_row or 1) + 1)]

    def is_sname(n):
        return ('#' in n) or (n in ('百分比', '1000KG'))
    hdr = None
    for gi, row in enumerate(grid):
        names = []
        for j in range(1, ncols):
            v = row[j]
            if isinstance(v, str) and v.strip() and is_sname(norm(v)) and not isnum(v):
                names.append((j, norm(v)))
        if len(names) >= 2:
            hdr = names
            hdr_gi = gi
            break
    if hdr is None:
        # 单列样本sheet：名称可能在 row[1..]
        for gi, row in enumerate(grid):
            names = []
            for j in range(1, ncols):
                v = row[j]
                if isinstance(v, str) and v.strip() and is_sname(norm(v)) and not isnum(v):
                    names.append((j, norm(v)))
            if len(names) == 1 and gi > 0:
                hdr = names
                hdr_gi = gi
                break
    if hdr is None:
        print('%-22s <无表头>' % ws.title)
        continue
    end_gi = len(grid)
    for gi in range(hdr_gi + 1, len(grid)):
        nm = grid[gi][0] if grid[gi] else None
        nm2 = grid[gi][1] if len(grid[gi]) > 1 else None
        if (isinstance(nm, str) and norm(nm) == '合计') or (isinstance(nm2, str) and norm(nm2) == '合计'):
            end_gi = gi
            break
    print('%-22s 表头行r%-2d 样本列=%s' % (ws.title, hdr_gi + 1, [n for _, n in hdr]))
    for j, n in hdr:
        s = 0.0
        for gi in range(hdr_gi + 1, end_gi):
            row = grid[gi]
            if j < len(row) and isnum(row[j]) and row[j] > 0:
                s += float(row[j])
        print('    col%-3d %-10s Σ=%10.4f' % (j, n, s))