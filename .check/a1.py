# -*- coding: utf-8 -*-
"""附件A 配比方案：逐 sheet 概要（配方列数、T弯G/水煮字段、原料行数）"""
import re, collections
import openpyxl

F = '/workspace/.uploads/b36f4809-d40f-491a-b9d7-2a9c4e359b1b_fbfc94091fa4955969e5d0fff7df0fd6_789507228604997242_m_3NX240913-6C--AI研发26.7.22配比方案.xlsx'
SKIP = re.compile(r'^(合计|产品|应用|背景|工艺|基材|理化|备注|外观|附着力|50KG|T弯|MEK|电腐蚀|3%|2%|三合一|BOX|单涂|双涂|双烘|序号|粘度|固含|说明|次数|时间|验收)$')


def isnum(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


wb = openpyxl.load_workbook(F, data_only=True)
for ws in wb.worksheets:
    ncols = min(ws.max_column, 30)
    grid = [[ws.cell(r, c).value for c in range(1, ncols + 1)] for r in range(1, (ws.max_row or 1) + 1)]
    ing_rows = []
    for gi, row in enumerate(grid):
        nm = None
        base = 0
        for j in (0, 1):
            v = row[j]
            if isinstance(v, str) and v.strip() and not isnum(v):
                nm = str(v).strip().replace('\n', '')
                base = j + 1
                break
        if not nm or SKIP.match(nm) or len(nm) > 20:
            continue
        vals = [j for j in range(base, min(ncols, 14)) if isnum(row[j]) and row[j] > 0]
        if vals:
            ing_rows.append((gi + 1, nm, vals))
    nvar = max((len(v) for _, _, v in ing_rows), default=0)
    label_rows = []
    for gi, row in enumerate(grid):
        t = str(row[0]).strip().replace('\n', '') if row[0] is not None else ''
        t2 = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
        if 'T弯G' in t or 'T弯G' in t2 or '水煮' in t or '杀菌水煮' in t:
            label_rows.append((gi + 1, (t or t2), [c for c in range(0, ncols) if row[c] is not None]))
    print('%-14s rows=%3d cols=%2d ingredients=%2d n_var=%d 标签行=%s' % (
        ws.title, ws.max_row, ws.max_column, len(ing_rows), nvar,
        label_rows[:3]))