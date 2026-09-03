# -*- coding: utf-8 -*-
"""环氧酚醛：标签冲突核对——merged 有值但与源文件不一致"""
import pickle, re, collections
import openpyxl
import numpy as np

U = '/workspace/.uploads/'
FILES = {
    '7.26': (U + 'cb7988e8-dc53-4617-9e33-71aa45000fcf_7.26配料测试汇总.xlsx', 'Sheet1'),
    '8.6': (U + 'a7f5f7b0-69c6-4226-bd58-0d02c928336b_8.6配料测试汇总.xlsx', '8.6配料测试汇总'),
    '8.14': (U + '05ddeff0-8071-4cc7-8e93-5df104d14162_8.14配料测试汇总.xlsx', '8.14配料测试汇总'),
}
COL_T, COL_M, COL_W = 18, 19, 20


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    m = re.match(r'^(\d+(?:\.\d+)?)\+$', s)
    return float(m.group(1)) if m else None


def parse(path, sheet):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    recs = {}
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(r, 2).value
        if pid is None:
            continue
        pid = str(pid).strip()
        d = {'T': num(ws.cell(r, COL_T).value), 'M': num(ws.cell(r, COL_M).value), 'W': num(ws.cell(r, COL_W).value)}
        if pid not in recs or any(recs[pid].get(k) is None for k in ('T', 'M', 'W')):
            sub = recs.setdefault(pid, {})
            for k in ('T', 'M', 'W'):
                if sub.get(k) is None and d[k] is not None:
                    sub[k] = d[k]
    return recs


src = {k: parse(*v) for k, v in FILES.items()}
D = pickle.load(open('/workspace/generalization/data/merged_data.pkl', 'rb'))
ms = {s['样本ID']: s for s in D['all_samples'] if s['体系'] == '环氧酚醛'}

print('### 标签冲突：merged 有值 ≠ 源文件值（同一体系内多源时以最新8.14优先，仍冲突才报）')
conf = collections.defaultdict(list)
for sid, s in ms.items():
    for tname, key in (('T弯', 'T'), ('MEK', 'M'), ('水煮', 'W')):
        mv = s.get(tname)
        if mv is None or (isinstance(mv, float) and np.isnan(mv)):
            continue
        mv = float(mv)
        srcvals = {}
        for k, recs in src.items():
            if sid in recs and recs[sid].get(key) is not None:
                srcvals[k] = recs[sid][key]
        if not srcvals:
            continue
        # 如果有任一源值与merged不一致（不同时所有源一致才叫唯一冲突，两源冲突取较新）
        diff = {k: v for k, v in srcvals.items() if abs(v - mv) > 1e-6}
        if diff:
            conf[tname].append((sid, diff, mv))
for t, lst in sorted(conf.items()):
    print(f'-- {t}: {len(lst)} 条')
    for sid, diff, mv in lst:
        print(f'      {sid:8s} merged={mv}  源差异={diff}')