# -*- coding: utf-8 -*-
"""配比方案 权威解析：逐sheet重建规范组分（重复原料行求和），自校验合计，与merged比对"""
import pickle, re, collections, unicodedata
import openpyxl

FA = '/workspace/.uploads/b36f4809-d40f-491a-b9d7-2a9c4e359b1b_fbfc94091fa4955969e5d0fff7df0fd6_789507228604997242_m_3NX240913-6C--AI研发26.7.22配比方案.xlsx'
D = pickle.load(open('/workspace/generalization/data/merged_data.pkl', 'rb'))


def norm(s):
    return unicodedata.normalize('NFKC', str(s)).replace('\n', '').replace(' ', '')


def isnum(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def txt(v):
    if v is None:
        return ''
    if isinstance(v, float) and v != v:
        return ''
    return str(v).strip()


NON_ING = re.compile(
    r'^(合计|含量|总和|总价|产品|应用|背景|工艺|基材|理化|备注|外观|附着力|50KG|T弯G?|MEK|电腐蚀|三合一|BOX|单涂|双涂|双烘|双烘方法|序号|粘度|固含|说明|次数|时间|验收|杀菌|水煮|马口铁|镀铬|膜厚|烘烤|T板|评级|等级|划格|冲击|盐雾|折弯|焊点|卷封|水浴|蒸汽|检验|检测|结论|结果|标准|要求|目标|玻璃化|密度|酸值|羟值|胺值|环氧当量|官能度|分子量|沸点|闪点|挥发|灰分|细度|滑度|流平|光泽|粗糙|铅笔|附着力级|百格|剥离|耐溶剂|耐酸碱|花架印|S$|3H|2H|3%氯化钠|三圈盖硫酸铜|121\*60|121\*6|0913)\D*$')


def classify(t):
    t = norm(t)
    if 'T弯G' in t:
        return 'T弯G'
    if '水煮' in t or '杀菌' in t:
        return '水煮'
    return None


wb = openpyxl.load_workbook(FA, data_only=True)
total_m = total_s = 0
for ws in wb.worksheets:
    if ws.title == 'Sheet1':
        continue
    ncols = min(ws.max_column, 16)
    grid = [[ws.cell(r, c).value for c in range(1, ncols + 1)] for r in range(1, (ws.max_row or 1) + 1)]

    # 表头行：含>=2个样本名列（名含# 或是 百分比/1000KG）
    def is_sname(n):
        return ('#' in n) or (n in ('百分比', '1000KG'))
    hdr_gi = None
    hdr_names = []
    for gi, row in enumerate(grid):
        names = []
        for j in range(1, ncols):
            v = row[j]
            if isinstance(v, str) and v.strip() and is_sname(norm(v)) and not isnum(v):
                names.append((j, norm(v)))
        if len(names) >= 2:
            hdr_gi = gi
            hdr_names = names
            break
    if hdr_gi is None:
        continue
    col_order = [j for j, _ in hdr_names]
    # 数据结束行（合计）
    end_gi = len(grid)
    for gi in range(hdr_gi + 1, len(grid)):
        nm = grid[gi][0] if grid[gi] else None
        nm2 = grid[gi][1] if len(grid[gi]) > 1 else None
        if (isinstance(nm, str) and norm(nm) == '合计') or (isinstance(nm2, str) and norm(nm2) == '合计'):
            end_gi = gi
            break
    # 每列组分（重复原料行求和）
    colcomp = {j: collections.defaultdict(float) for j in col_order}
    for gi in range(hdr_gi + 1, end_gi):
        row = grid[gi]
        nm = None
        for j in (0, 1, 2):
            v = row[j] if j < len(row) else None
            if isinstance(v, str) and norm(v) and not isnum(v):
                nm = re.sub(r'\s', '', norm(v))
                break
        if not nm or len(nm) > 30 or NON_ING.match(nm) or classify(nm):
            continue
        for j in col_order:
            if j < len(row) and isnum(row[j]) and row[j] > 0:
                colcomp[j][nm] += float(row[j])
    # 合计行原始值（自校验：完整和 vs 合计；差异若等于末行溶剂则视为源表quirk）
    sumrow = None
    if end_gi < len(grid):
        sumrow = [grid[end_gi][j] for j in col_order]
    # 每列“末行值”（最后一条原料记录）
    lastrow = {j: None for j in col_order}
    for gi in range(hdr_gi + 1, end_gi):
        row = grid[gi]
        nm = None
        for j0 in (0, 1, 2):
            v = row[j0] if j0 < len(row) else None
            if isinstance(v, str) and norm(v) and not isnum(v):
                nm = re.sub(r'\s', '', norm(v))
                break
        if not nm or len(nm) > 30 or NON_ING.match(nm) or classify(nm):
            continue
        for j in col_order:
            if j < len(row) and isnum(row[j]) and row[j] > 0:
                lastrow[j] = float(row[j])
    print('=' * 110)
    print('SHEET %-22s 表头=%s 原料行%d~%d' % (ws.title, [n for _, n in hdr_names], hdr_gi + 2, end_gi))
    for j, _ in hdr_names:
        s = sum(colcomp[j].values())
        if sumrow is None or sumrow[col_order.index(j)] is None:
            flag = '(无合计行)'
        else:
            src = float(sumrow[col_order.index(j)])
            if abs(src - s) < 0.05:
                flag = ''
            elif lastrow[j] and abs(src - (s - lastrow[j])) < 0.05:
                flag = '  合计=%.4f(不含末行)' % src
            else:
                flag = '  <-- 源合计%.4f 不符' % src
        print('  col%d %-10s Σ=%.4f %s' % (j, hdr_names[col_order.index(j)][1], s, flag))
    # 与merged比对（按顺序）
    tail = [s['样本ID'].split('环氧-配比方案-', 1)[1] for s in D['all_samples']
            if s['体系'] == '环氧-配比方案' and s['样本ID'].split('环氧-配比方案-', 1)[1].startswith(ws.title + '-')]
    ms = [s for s in D['all_samples'] if s['体系'] == '环氧-配比方案' and
          s['样本ID'].split('环氧-配比方案-', 1)[1].startswith(ws.title + '-')]
    # 按序号排序以与col_order对应
    ms_sorted = sorted(ms, key=lambda s: int(s['样本ID'].rsplit('-', 1)[1]))
    for k, m in enumerate(ms_sorted):
        mc = collections.defaultdict(float)
        for a, b in m['组分'].items():
            mc[re.sub(r'\s', '', norm(a))] += b
        if k < len(col_order):
            j = col_order[k]
            sc = dict(colcomp[j])
            keys = set(mc) | set(sc)
            dv = []
            for a in keys:
                mv = mc.get(a, 0.0)
                sv = sc.get(a, 0.0)
                if abs(mv - sv) > 0.05:
                    dv.append('%s:%s→%s' % (a, round(mv, 3), round(sv, 3)))
            if dv:
                print('  [DIFF] %-44s %s' % (m['样本ID'], '; '.join(dv)))
            total_m += 1
            if not dv:
                total_s += 1
print()
print('### 配比方案样本 %d，完全一致 %d' % (total_m, total_s))