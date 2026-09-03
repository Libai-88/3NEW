# -*- coding: utf-8 -*-
"""判定每个标签取自哪个文件，并量化 8.14 独有的缺失数据"""
import pickle, re, collections, statistics
import openpyxl

U = '/workspace/.uploads/'
FILES = {
    '7.26': (U + 'cb7988e8-dc53-4617-9e33-71aa45000fcf_7.26配料测试汇总.xlsx', 'Sheet1'),
    '8.6': (U + 'a7f5f7b0-69c6-4226-bd58-0d02c928336b_8.6配料测试汇总.xlsx', '8.6配料测试汇总'),
    '8.14': (U + '05ddeff0-8071-4cc7-8e93-5df104d14162_8.14配料测试汇总.xlsx', '8.14配料测试汇总'),
}
T, M, W = 18, 19, 20


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ('', '——', '-'):
        return None
    m = re.match(r'^(\d+(?:\.\d+)?)\+$', s)
    return float(m.group(1)) if m else None


def parse(path, sheet):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    heads = [ws.cell(1, c).value for c in range(1, 21)]
    recs = []
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(r, 2).value
        if pid is None:
            continue
        recs.append({'row': r, 'id': str(pid).strip(), 'bar': str(ws.cell(r, 3).value or '').strip(),
                     'T': num(ws.cell(r, T).value), 'M': num(ws.cell(r, M).value), 'W': num(ws.cell(r, W).value)})
    return recs


src = {k: parse(*v) for k, v in FILES.items()}
D = pickle.load(open('/workspace/generalization/data/merged_data.pkl', 'rb'))
ms = {s['样本ID']: s for s in D['all_samples'] if s['来源'] == '配料测试数据汇总V1'}

by = {k: collections.defaultdict(list) for k in src}
for k, rs in src.items():
    for r in rs:
        by[k][r['id']].append(r)

allids = sorted(set(ms))
verdict = collections.Counter()
detail = collections.defaultdict(list)
for sid in allids:
    for tname, key in (('T弯', 'T'), ('MEK', 'M'), ('水煮', 'W')):
        mv = ms[sid][tname]
        mv = None if mv is None else round(float(mv), 4)
        per = {}
        for k in src:
            vals = [r[key] for r in by[k].get(sid, []) if r[key] is not None]
            per[k] = round(statistics.mean(vals), 4) if vals else None
        match = [k for k in src if per[k] is not None and per[k] == mv]
        if mv is None and all(v is None for v in per.values()):
            verdict[(tname, 'all-empty')] += 1
        elif match:
            verdict[(tname, 'matches:' + '+'.join(sorted(match)))] += 1
        elif mv is None:
            verdict[(tname, 'MISSING-in-merged')] += 1
            detail[(tname, 'missing')].append((sid, per))
        else:
            verdict[(tname, 'no-match')] += 1
            detail[(tname, 'nomatch')].append((sid, mv, per))

print('### 每个标签与哪个源文件一致（按 样本×目标 计数）')
for k in sorted(verdict, key=lambda x: (x[0], -verdict[x])):
    print(f'  {k[0]:4s} {k[1]:28s} {verdict[k]}')

print()
print('### 合并版缺失、但源文件有值的样本（前 60 条）')
for t in ('T弯', 'MEK', '水煮'):
    lst = detail[(t, 'missing')]
    print(f'-- {t}: {len(lst)} 条')
    for x in lst[:60]:
        print('    ', x[0], {k: v for k, v in x[1].items() if v is not None})
print()
print('### 合并版有值但与任何单文件均值都不同的样本')
for t in ('T弯', 'MEK', '水煮'):
    lst = detail[(t, 'nomatch')]
    print(f'-- {t}: {len(lst)} 条')
    for x in lst[:40]:
        print('    ', x[0], 'merged=', x[1], 'per-file=', x[2])
