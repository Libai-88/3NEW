# -*- coding: utf-8 -*-
"""附件A 配比方案：逐 sheet 提取列位配方 + T弯G/水煮档位，与 merged 样本比对"""
import pickle, re, sys, collections, unicodedata
import openpyxl

sys.path.insert(0, '/workspace/generalization/workbench')
from DataPrepWorkbench import clean_code

FA = '/workspace/.uploads/b36f4809-d40f-491a-b9d7-2a9c4e359b1b_fbfc94091fa4955969e5d0fff7df0fd6_789507228604997242_m_3NX240913-6C--AI研发26.7.22配比方案.xlsx'
D = pickle.load(open('/workspace/generalization/data/merged_data.pkl', 'rb'))
merged = [s for s in D['all_samples'] if s['体系'] == '环氧-配比方案']
by_sheet = collections.defaultdict(list)
for s in merged:
    tail = s['样本ID'].split('环氧-配比方案-', 1)[1]
    by_sheet[tail[:tail.rfind('-')]].append(s)

NON_ING = re.compile(r'^(合计|产品|应用|背景|工艺|基材|理化|备注|外观|附着力|50KG|T弯|MEK|电腐蚀|3%|2%|三合一|BOX|单涂|双涂|双烘|序号|粘度|固含|说明|次数|时间|验收|粘度/|按205|0913|1KG|MIBK$|DMP$|DBE$|PM$)$')


def isnum(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def norm(s):
    return unicodedata.normalize('NFKC', str(s)).replace('\n', '').replace(' ', '')


def classify(t):
    t = norm(t)
    if 'T弯G' in t:
        return 'T弯G'
    if '水煮' in t and '杀菌' in t:
        return '水煮'
    return None


wb = openpyxl.load_workbook(FA, data_only=True)
total_ok = total_pct = total_unmatched = 0
for ws in wb.worksheets:
    if ws.title == 'Sheet1':
        continue
    ncols = min(ws.max_column, 16)
    grid = [[ws.cell(r, c).value for c in range(1, ncols + 1)] for r in range(1, (ws.max_row or 1) + 1)]
    # 原料行 → 各列 amount（同类累加）
    colcomp = collections.defaultdict(collections.Counter)
    for gi, row in enumerate(grid):
        nm = None
        base = 0
        for j in (0, 1, 2):
            v = row[j]
            if isinstance(v, str) and norm(v) and not isnum(v):
                nm = norm(v)
                base = j + 1
                break
        if not nm or NON_ING.match(nm) or len(nm) > 25:
            continue
        if classify(nm) is not None:
            continue
        try:
            code = clean_code(nm)
        except Exception:
            code = nm
        for j in range(base, ncols):
            v = row[j]
            if isnum(v) and v > 0 and abs(v) < 1e6:
                colcomp[j][code] += float(v)
    cols = {j: dict(c) for j, c in colcomp.items() if c}
    # 性能行
    perf = {}
    for gi, row in enumerate(grid):
        key = None
        for j in range(min(4, len(row))):
            t = row[j]
            if isinstance(t, str) and classify(t):
                key = classify(t)
                break
        if key:
            for j in list(cols):
                v = row[j] if j < len(row) else None
                if v is not None and str(v).strip() not in ('', None):
                    perf.setdefault(key, {})[j] = v
    # 汇总该 sheet 列数/总数
    weight = {}
    for j, c in cols.items():
        tot = sum(c.values())
        if tot > 0:
            weight[j] = tot
    print('=' * 100)
    print('SHEET %-14s  列数=%d 合并版样本=%d' % (ws.title, len(cols), len(by_sheet.get(ws.title, []))))
    if perf:
        for k, v in perf.items():
            print('   T弯G/水煮 行: %s' % {j: str(x) for j, x in v.items()})
    for j, c in sorted(cols.items()):
        tot = sum(c.values())
        sample = ' '.join('%s=%.2f' % (k, v) for k, v in sorted(c.items())[:8])
        print('   列%-2d n=%2d Σ=%.2f  %s' % (j, len(c), tot, sample))
    # merged 样本匹配
    for m in by_sheet.get(ws.title, []):
        mc = {}
        for a, b in m['组分'].items():
            cc = clean_code(a)
            mc[cc] = mc.get(cc, 0.0) + b
        mset = set(mc)
        best = None
        for j, c in cols.items():
            cset = set(c)
            miss = len(mset ^ cset)
            dv = [a for a in mset & cset if abs(mc[a] - c[a]) > max(0.05, 0.01 * c[a])]
            score = miss + len(dv)
            if best is None or score < best[0]:
                best = (score, j, cset, mset ^ cset, dv)
        sc, j, cset, diffk, diffv = best
        tag = 'OK' if sc == 0 else 'DIFF(%d)' % sc
        print('   [%s] %-38s (Σmerged=%.2f) 列%d: 差组=%s 差量=%s' % (
            tag, m['样本ID'], sum(mc.values()), j, diffk if diffk else '', {a: (mc[a], cset) for a in diffv} or ''))
    total_ok += sum(1 for m in by_sheet.get(ws.title, []) if True)