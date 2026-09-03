# -*- coding: utf-8 -*-
"""配比方案：问题sheet全格转储（所有列，供人工核对对齐）"""
import openpyxl
F = "/workspace/.uploads/b36f4809-d40f-491a-b9d7-2a9c4e359b1b_fbfc94091fa4955969e5d0fff7df0fd6_789507228604997242_m_3NX240913-6C--AI研发26.7.22配比方案.xlsx"
wb = openpyxl.load_workbook(F, data_only=True)

def fmt(v):
    if v is None:
        return ''
    if isinstance(v, float):
        return ('%.6f' % v).rstrip('0').rstrip('.')
    return str(v).replace('\n', '\\n')

SHEETS = ['25.1.13', '25.1.16', '25.1.16-2', '25.1.17', '25.3.12-180KG配方确认',
          '25.9.30', '25.11.18', '25.11.24-2', '25.11.26', '25.12.31']
out = open('/tmp/pb_grid3.txt', 'w')
for name in SHEETS:
    if name not in wb.sheetnames:
        out.write('MISSING SHEET: %s\n' % name)
        continue
    ws = wb[name]
    out.write('=' * 120 + '\n')
    out.write('SHEET: %s  rows=%d cols=%d\n' % (ws.title, ws.max_row, ws.max_column))
    for r in range(1, ws.max_row + 1):
        vals = []
        for c in range(1, ws.max_column + 1):
            v = fmt(ws.cell(r, c).value)
            if v:
                vals.append('c%d=%s' % (c, v))
        if vals:
            out.write('r%-2d: %s\n' % (r, '  '.join(vals)))
out.close()
print('done')