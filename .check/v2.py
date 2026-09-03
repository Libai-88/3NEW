# -*- coding: utf-8 -*-
"""逐样本比对：环氧酚醛 组成用量 + 三项性能标签"""
import pickle, re, collections, statistics
import openpyxl

U = '/workspace/.uploads/'
FILES = {
    '7.26': (U + 'cb7988e8-dc53-4617-9e33-71aa45000fcf_7.26配料测试汇总.xlsx', 'Sheet1'),
    '8.6': (U + 'a7f5f7b0-69c6-4226-bd58-0d02c928336b_8.6配料测试汇总.xlsx', '8.6配料测试汇总'),
    '8.14': (U + '05ddeff0-8071-4cc7-8e93-5df104d14162_8.14配料测试汇总.xlsx', '8.14配料测试汇总'),
}
MATCOLS = list(range(4, 18))
T, M, W = 18, 19, 20


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ('', '——', '-'):
        return None
    m = re.match(r'^(\d+(?:\.\d+)?)\+$', s)
    if m:
        return float(m.group(1))
    return None


def parse(path, sheet):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    heads = [ws.cell(1, c).value for c in range(1, 21)]
    recs = []
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(r, 2).value
        if pid is None:
            continue
        bar = str(ws.cell(r, 3).value or '').strip()
        comp = {}
        for c in MATCOLS:
            nm = heads[c - 1]
            raw = ws.cell(r, c).value
            v = num(raw)
            if raw is not None and str(raw).strip() != '' and v is None:
                comp['<非数值:%s=%r>' % (nm, raw)] = None
            elif v is not None:
                comp[nm] = v
        recs.append({'row': r, 'id': str(pid).strip(), 'bar': bar, 'comp': comp,
                     'T': num(ws.cell(r, T).value), 'M': num(ws.cell(r, M).value),
                     'W': num(ws.cell(r, W).value),
                     'Traw': ws.cell(r, T).value, 'Mraw': ws.cell(r, M).value, 'Wraw': ws.cell(r, W).value})
    return recs


src = {k: parse(*v) for k, v in FILES.items()}
D = pickle.load(open('/workspace/generalization/data/merged_data.pkl', 'rb'))
ms = {s['样本ID']: s for s in D['all_samples'] if s['来源'] == '配料测试数据汇总V1'}


def norm_comp(comp):
    return {k: round(v, 6) for k, v in comp.items() if v is not None and v != 0}


def fnum(v):
    return '' if v is None else round(v, 6)


print('### A. 组成用量比对（合并版 vs 各文件同 ID 行）')
comp_issues = []
for sid, s in sorted(ms.items()):
    mc = norm_comp(s['组分'])
    rows = [(k, r) for k in src for r in src[k] if r['id'] == sid]
    if not rows:
        continue
    for k, r in rows:
        sc = norm_comp(r['comp'])
        if sc != mc:
            only_m = {a: b for a, b in mc.items() if sc.get(a) != b}
            only_s = {a: b for a, b in sc.items() if mc.get(a) != b}
            comp_issues.append((sid, k, r['row'], r['bar'], only_m, only_s,
                                [x for x in r['comp'] if x.startswith('<非数值')]))
print(f'不一致 (样本,文件,行) 组合数: {len(comp_issues)}')
cnt = collections.Counter()
for sid, k, row, bar, om, os_, nonnum in comp_issues:
    cnt[(k, 'only_in_merged' if om and not os_ else ('only_in_src' if os_ and not om else 'diff'))] += 1
print('分类计数:', dict(cnt))
for x in comp_issues[:45]:
    print('  ', x[0], x[1], 'row%d' % x[2], x[3], '| 合并版独有/不同:', x[4], '| 源文件独有/不同:', x[5], x[6])

print()
print('### B. 性能标签比对')
rows_by = collections.defaultdict(list)
for k, rs in src.items():
    for r in rs:
        rows_by[r['id']].append((k, r))

tgt_issues = []
agree = collections.Counter()
for sid, s in sorted(ms.items()):
    rs = rows_by.get(sid, [])
    if not rs:
        continue
    for tname, key in (('T弯', 'T'), ('MEK', 'M'), ('水煮', 'W')):
        vals = [r[key] for _, r in rs if r[key] is not None]
        srcmean = statistics.mean(vals) if vals else None
        mv = s[tname]
        if mv is None and srcmean is None:
            agree[(tname, 'both-empty')] += 1
            continue
        a = None if mv is None else round(float(mv), 4)
        b = None if srcmean is None else round(srcmean, 4)
        if a == b:
            agree[(tname, 'mean-match')] += 1
        elif mv is not None and vals and any(abs(mv - v) < 1e-6 for v in vals):
            agree[(tname, 'single-value-match')] += 1
        else:
            tgt_issues.append((sid, tname, a, b, [(k, r['bar'], r[key], r[key + 'raw']) for k, r in rs]))
print('一致分类:', dict(agree))
print(f'不一致: {len(tgt_issues)}')
for x in tgt_issues[:40]:
    print('  ', x[0], x[1], '合并版=', x[2], '各源行均值=', x[3], '源行=', x[4])
