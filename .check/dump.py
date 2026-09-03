import sys
import openpyxl

path = sys.argv[1]
sheets = sys.argv[2].split(",") if len(sys.argv) > 2 and sys.argv[2] != "*" else None
maxcol = int(sys.argv[3]) if len(sys.argv) > 3 else 40
wb = openpyxl.load_workbook(path, data_only=True)
for ws in wb.worksheets:
    if sheets and ws.title not in sheets:
        continue
    print("=" * 100)
    print("SHEET:", ws.title, "dims:", ws.dimensions, "max_row:", ws.max_row, "max_col:", ws.max_column)
    print("MERGED:", [str(r) for r in ws.merged_cells.ranges][:80])
    print("=" * 100)
    for r in range(1, (ws.max_row or 1) + 1):
        vals = []
        for c in range(1, min(ws.max_column or 1, maxcol) + 1):
            v = ws.cell(r, c).value
            vals.append("" if v is None else str(v).replace("\n", "\\n"))
        if any(x != "" for x in vals):
            print(r, "|", " | ".join(vals))
