# -*- coding: utf-8 -*-
"""转储 配比方案 全部 sheet 网格到文本"""
import openpyxl
F = "/workspace/.uploads/b36f4809-d40f-491a-b9d7-2a9c4e359b1b_fbfc94091fa4955969e5d0fff7df0fd6_789507228604997242_m_3NX240913-6C--AI研发26.7.22配比方案.xlsx"
wb = openpyxl.load_workbook(F, data_only=True)
print('sheets:', wb.sheetnames)
out = open('/tmp/pb_sheets.txt', 'w')
for ws in wb.worksheets:
    out.write('=' * 100 + '\n')
    out.write('SHEET: %s  rows=%d cols=%d\n' % (ws.title, ws.max_row, ws.max_column))
    for r in range(1, min(ws.max_row, 200) + 1):
        vals = []
        for c in range(1, min(ws.max_column, 16) + 1):
            v = ws.cell(r, c).value
            vals.append('' if v is None else str(v).replace('\n', '\\n'))
        line = ' | '.join(vals)
        if line.strip(' |'):
            out.write('r%d: %s\n' % (r, line))
out.close()
print('written /tmp/pb_sheets.txt')