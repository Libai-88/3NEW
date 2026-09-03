# -*- coding: utf-8 -*-
"""列出三个配料测试汇总文件的所有 sheet + 各 sheet 内 样本ID 前缀分布"""
import openpyxl, collections

U = '/workspace/.uploads/'
FILES = {
    '7.26': U + 'cb7988e8-dc53-4617-9e33-71aa45000fcf_7.26配料测试汇总.xlsx',
    '8.6': U + 'a7f5f7b0-69c6-4226-bd58-0d02c928336b_8.6配料测试汇总.xlsx',
    '8.14': U + '05ddeff0-8071-4cc7-8e93-5df104d14162_8.14配料测试汇总.xlsx',
}
for k, f in FILES.items():
    wb = openpyxl.load_workbook(f, data_only=True)
    print('=' * 70)
    print(k, wb.sheetnames)
    for ws in wb.worksheets:
        pids = []
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, 2).value
            if v is not None and not isinstance(v, (int, float)):
                pids.append(str(v).strip())
            elif isinstance(v, (int, float)) and not isinstance(v, bool):
                pids.append(str(v))
        pref = collections.Counter()
        for p in pids:
            m = p
            pref[m.split('-')[0]] += 1
        # 也扫描第1列
        pids1 = []
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v is not None:
                pids1.append(str(v).strip())
        pref1 = collections.Counter(p.split('-')[0] for p in pids1)
        print('   sheet=%-16s 行数=%-4d 前缀(col2)=%s' % (ws.title, ws.max_row, dict(pref)))
        if pref1 and any(x != 'None' for x in pids1):
            print('     前缀(col1)=%s' % dict(pref1))