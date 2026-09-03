# -*- coding: utf-8 -*-
"""配比方案最终核对：按列序号一一对应比对（合并样本 → 同序源列），
源列仅统计公式区（表头行之后到性能行之前），排除性能行/非原料行。
输出：每个 sheet 每列的组成、合并样本一一对应后的差异明细。
"""
import pickle, re, sys, collections, unicodedata
import openpyxl

sys.path.insert(0, '/workspace/generalization/workbench')
from DataPrepWorkbench import clean_code

FA = '/workspace/.uploads/b36f4809-d40f-491a-b9d7-2a9c4e359b1b_fbfc94091fa4955969e5d0fff7df0fd6_789507228604997242_m_3NX240913-6C--AI研发26.7.22配比方案.xlsx'
D = pickle.load(open('/workspace/generalization/data/merged_data.pkl', 'rb'))
merged_all = [s for s in D['all_samples'] if s['体系'] == '环氧-配比方案']
by_sheet = collections.defaultdict(list)
for s in merged_all:
    tail = s['样本ID'].split('环氧-配比方案-', 1)[1]
    by_sheet[tail[:tail.rfind('-')]].append(s)

# 非原料行：只排除明确的性能/说明/表头词。绝对不按 `%` 前缀排除（'3%气硅'等是原料）
NON_ING = re.compile(r'^(合计|含量|总和|总价|产品|应用|背景|工艺|基材|理化|备注|外观|附着力|50KG|T弯G?|MEK|电腐蚀|三合一|BOX|单涂|双涂|双烘|双烘方法|序号|粘度|固含|说明|次数|时间|验收|杀菌|水煮|马口铁|镀铬|膜厚|烘烤|T板|评级|等级|划格|冲击|盐雾|折弯|焊点|卷封|水浴|蒸汽|检验|检测|结论|结果|标准|要求|目标|玻璃化|密度|酸值|羟值|胺值|环氧当量|官能度|分子量|沸点|闪点|含量|挥发|灰分|细度|滑度|流平|光泽|粗糙|铅笔|附着力级|百格|剥离|耐溶剂|耐酸碱)\D*$')


def fw(t):
    return t.replace('（', '(').replace('）', ')').replace('，', ',').replace('：', ':')


def isnum(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def norm(s):
    return unicodedata.normalize('NFKC', str(s)).replace('\n', '').replace(' ', '')


def classify(t):
    t = norm(t)
    if 'T弯G' in t:
        return 'T弯G'
    if '水煮' in t or '杀菌' in t:
        return '水煮'
    return None


wb = openpyxl.load_workbook(FA, data_only=True)
n_total = n_ok = 0
miss_rows = []
for ws in wb.worksheets:
    if ws.title == 'Sheet1':
        continue
    ncols = min(ws.max_column, 16)
    grid = [[ws.cell(r, c).value for c in range(1, ncols + 1)] for r in range(1, (ws.max_row or 1) + 1)]
    # 找公式区的行范围：从第一行含“序号”+ 样本表头行开始到“合计”行结束
    # 样本表头行：该行含 >=2 个短文本样本名（'6-1# 等）
    hdr_gi = None
    for gi, row in enumerate(grid):
        cnt = 0
        for j in range(1, ncols):
            v = row[j]
            if isinstance(v, str) and v.strip() and len(norm(v)) <= 8 and not isnum(v):
                cnt += 1
        if cnt >= 2:
            hdr_gi = gi
            break
    # 列顺序：表头行的短文本所在列（保持原文顺序）
    col_order = []
    if hdr_gi is not None:
        for j in range(1, ncols):
            v = grid[hdr_gi][j]
            if isinstance(v, str) and v.strip() and len(norm(v)) <= 8 and not classify(v):
                col_order.append(j)
    else:
        # 无表头：用第1个数据列开始的所有有数值列
        col_order = [j for j in range(2, ncols) if any(isnum(grid[gi][j]) for gi in range(1, len(grid)))][:8]
    end_gi = None
    for gi in range(hdr_gi + 1 if hdr_gi is not None else 1, len(grid)):
        nm = grid[gi][0] if grid[gi] else None
        nm2 = grid[gi][1] if len(grid[gi]) > 1 else None
        if isinstance(nm, str) and norm(nm) == '合计':
            end_gi = gi
            break
        if isinstance(nm2, str) and norm(nm2) == '合计':
            end_gi = gi
            break
    # 每列组成
    colcomp = {j: collections.Counter() for j in col_order}
    for gi in range((hdr_gi + 1) if hdr_gi is not None else 1, (end_gi if end_gi is not None else len(grid))):
        row = grid[gi]
        nm = None
        base = 0
        for j in (0, 1, 2):
            v = row[j] if j < len(row) else None
            if isinstance(v, str) and norm(v) and not isnum(v):
                nm = norm(v)
                base = j + 1
                break
        if not nm or len(nm) > 30 or NON_ING.match(nm) or classify(nm):
            continue
        try:
            code = clean_code(fw(nm))
        except Exception:
            code = fw(nm)
        for j in col_order:
            if j < len(row) and isnum(row[j]) and row[j] > 0 and abs(row[j]) < 1e6:
                colcomp[j][code] += float(row[j])
    # 性能行（用于标签核对，单独打印）
    perf = collections.defaultdict(dict)
    for gi, row in enumerate(grid):
        key = None
        for j in range(min(4, len(row))):
            t = row[j]
            if isinstance(t, str) and classify(t):
                key = classify(t)
                break
        if key:
            for j in col_order:
                if j < len(row) and row[j] is not None and str(row[j]).strip() not in ('', 'None'):
                    perf[key][j] = row[j]
    print('=' * 100)
    print('SHEET %-16s 源列=%s 合并样本=%d' % (ws.title, col_order, len(by_sheet.get(ws.title, []))))
    if perf:
        for k, v in perf.items():
            print('   性能 %s: %s' % (k, {col_order.index(j) if j in col_order else j: str(x) for j, x in sorted(v.items()) if j in col_order}))
    # 一一对应
    ms = by_sheet.get(ws.title, [])
    for k, m in enumerate(ms):
        mc = {}
        for a, b in m['组分'].items():
            cc = clean_code(fw(a))
            mc[cc] = mc.get(cc, 0.0) + b
        if k < len(col_order):
            j = col_order[k]
            c = colcomp[j]
            cset = set(c)
            mset = set(mc)
            dv = [a for a in mset & cset if abs(mc[a] - c[a]) > max(0.05, 0.01 * c[a])]
            only_m = sorted(mset - cset)
            only_c = sorted(cset - mset)
            ok = (not only_m and not only_c and not dv)
            tag = 'OK' if ok else 'DIFF'
            if ok:
                n_ok += 1
            extra = ''
            if not ok:
                extra = ' 仅合并版有=%s 仅源列有=%s 值差=%s' % (only_m, only_c,
                       {a: (round(mc[a], 3), round(c[a], 3)) for a in dv} if dv else '')
            print('   [%s] %-44s Σ源=%8.3f Σ并=%8.3f%s' % (tag, m['样本ID'], sum(c.values()), sum(mc.values()), extra))
        else:
            print('   [??] %-44s 无对应源列（合并样本数多于源列数）' % m['样本ID'])
        n_total += 1
print()
print('### 汇总：一一对应样本 %d 个，完全一致 %d 个，存在差异 %d 个' % (n_total, n_ok, n_total - n_ok))