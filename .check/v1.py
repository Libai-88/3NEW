# -*- coding: utf-8 -*-
"""独立解析 3 个配料测试汇总文件，与 merged_data.pkl 逐样本比对（覆盖 + 数值）"""
import pickle, re, collections, math
import openpyxl

U = '/workspace/.uploads/'
FILES = {
    '7.26': (U + 'cb7988e8-dc53-4617-9e33-71aa45000fcf_7.26配料测试汇总.xlsx', 'Sheet1'),
    '8.6': (U + 'a7f5f7b0-69c6-4226-bd58-0d02c928336b_8.6配料测试汇总.xlsx', '8.6配料测试汇总'),
    '8.14': (U + '05ddeff0-8071-4cc7-8e93-5df104d14162_8.14配料测试汇总.xlsx', '8.14配料测试汇总'),
}
MATCOLS = list(range(4, 18))   # D..R = 14 个原料列
T, M, W = 18, 19, 20


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ('', '——', '-', '#DIV/0!'):
        return None
    m = re.match(r'^(\d+(?:\.\d+)?)\s*([+])?$', s)
    if m:
        return float(m.group(1))          # '300+' / '4+' -> 数值部分
    return None


def flag(v):
    s = str(v).strip() if v is not None else ''
    return '+' in s or '<' in s or '>' in s


def parse(path, sheet):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    heads = [ws.cell(1, c).value for c in range(1, 21)]
    recs = []
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(r, 2).value
        bar = ws.cell(r, 3).value
        if pid is None:
            continue
        comp = {}
        for c in MATCOLS:
            name = heads[c - 1]
            v = ws.cell(r, c).value
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
            comp[name] = num(v)
            if num(v) is None:
                comp['!!RAW' + str(name)] = v
        tv, mv, wv = ws.cell(r, T).value, ws.cell(r, M).value, ws.cell(r, W).value
        recs.append({
            'row': r, 'id': str(pid).strip(), 'bar': str(bar).strip() if bar else '',
            'comp': comp,
            'T': num(tv), 'M': num(mv), 'W': num(wv),
            'Traw': tv, 'Mraw': mv, 'Wraw': wv,
        })
    return heads, recs


D = pickle.load(open('/workspace/generalization/data/merged_data.pkl', 'rb'))
ms = {s['样本ID']: s for s in D['all_samples']}
msrc = [s for s in D['all_samples'] if s['来源'] == '配料测试数据汇总V1']
mids = set(s['样本ID'] for s in msrc)

parsed = {}
for k, (p, sh) in FILES.items():
    heads, recs = parse(p, sh)
    parsed[k] = recs
    print(f'[{k}] 行数={len(recs)}  唯一配方ID={len(set(r["id"] for r in recs))}')

# 每个文件的 ID -> 合并版是否存在
for k, recs in parsed.items():
    ids = sorted(set(r['id'] for r in recs))
    miss = [i for i in ids if i not in mids]
    print(f'[{k}] 未在合并版中的配方ID({len(miss)}): {miss}')

# 合并版中 R/D/C 系列 ID 不出现在任何文件中的
allfileids = collections.Counter()
for k, recs in parsed.items():
    for i in set(r['id'] for r in recs):
        allfileids[i] += 1
orph = [i for i in sorted(mids) if i not in allfileids]
print(f'\n合并版环氧酚醛样本 {len(mids)} 条；其中不出现在 3 个附件中的 {len(orph)} 条:')
print('  ', orph)
print('  按系列:', collections.Counter(i.rsplit("-", 1)[0] for i in orph))
