# -*- coding: utf-8 -*-
"""转储指定 配比方案 sheet 的完整网格（含全部列），用于人工核验"""
import openpyxl
F = "/workspace/.uploads/b36f4809-d40f-491a-b9d7-2a9c4e359b1b_fbfc94091fa4955969e5d0fff7df0fd6_789507228604997242_m_3NX240913-6C--AI研发26.7.22配比方案.xlsx"
SHEETS = ['25.1.13', '25.1.16', '25.1.16-2', '25.1.17', '25.3.12-180KG配方确认',
          '25.9.30', '25.11.18', '25.11.24-2', '25.11.26', '25.12.31']
wb = openpyxl.load_workbook(F, data_only=True)


def fmt(v):
    if v is None:
        return ''
    if isinstance(v, float):
        return ('%.6f' % v).rstrip('0').rstrip('.')
    return str(v).replace('\n', '\\n')


for name in SHEETS:
    if name not in wb.sheetnames:
        print('MISSING SHEET:', name)
        continue
    ws = wb[name]
    print('=' * 120)
    print('SHEET: %s  rows=%d cols=%d' % (ws.title, ws.max_row, ws.max_column))
    # 先打印前几行找表头结构
    for r in range(1, ws.max_row + 1):
        vals = []
        for c in range(1, ws.max_column + 1):
            v = fmt(ws.cell(r, c).value)
            if v:
                vals.append('c%d=%s' % (c, v))
        if vals:
            print('r%d: %s' % (r, '  '.join(vals)))