# -*- coding: utf-8 -*-
"""配比方案：严谨逐 sheet 比对。每 sheet 的每个配方列（真实表头）视为一个源配方；
汇总同代码行（重复原料行累加），与 merged_data.pkl 中同 sheet 样本按组成严格比对。
性能行（T弯G/水煮/杀菌）单独提取核对。
"""
import pickle, re, sys, collections, unicodedata
import openpyxl

sys.path.insert(0, '/workspace/generalization/workbench')
from DataPrepWorkbench import clean_code

FA = '/workspace/.uploads/b36f4809-d40f-491a-b9d7-2a9c4e359b1b_fbfc94091fa4955969e5d0fff7df0fd6_789507228604997242_m_3NX240913-6C--AI研发26.7.22配比方案.xlsx'
D = pickle.load(open('/workspace/generalization/data/merged_data.pkl', 'rb'))
merged = [s for s in D['all_samples'] if s['体系'] == '环氧-配比方案']
by_sheet = collections.defaultdict(list)
for s in merged:
    tail = s['样本ID'].split('环氧-配比方案-', 1)[1]
    by_sheet[tail[:tail.rfind('-')]].append(s)

# 非原料行判定：表头/说明/性能项。注意 MIBK/DMP/DBE/PM 是真实溶剂不能排除；
# 3%气硅/10%磷酸/10%135 等以“%”开头的也常是原料名，不能按前缀排除。只排除明确匹配的非原料词或带性能数值短语的行。
NON_ING = re.compile(r'^(合计|含量|总和|总价|产品|应用|背景|工艺|基材|理化|备注|外观|附着力|50KG|T弯G|T弯|MEK|电腐蚀|三合一|BOX|单涂|双涂|双烘|双烘方法|序号|粘度|固含|说明|次数|时间|验收|杀菌|水煮|马口铁|镀铬|膜厚|烘烤|T板|评级|等级|划格|冲击|盐雾|折弯|焊点|卷封|水浴|蒸汽|检验|检测|结论|结果|标准|要求|目标|Tg|玻璃化|反应|熔化|软化|凝|密度|pH|酸值|羟值|胺值|环氧当量|官能度|分子量|沸点|闪点|含量|挥发|灰分|细度|粘度计)\D*$')


def fw(t):
    """全角括号/数字→半角，便于名称归一"""
    t = t.replace('（', '(').replace('）', ')').replace('，', ',').replace('：', ':')
    return t


def isnum(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def norm(s):
    return unicodedata.normalize('NFKC', str(s)).replace('\n', '').replace(' ', '')


def classify(t):
    t = norm(t)
    if 'T弯G' in t:
        return 'T弯G'
    if '水煮' in t or '杀菌' in t:
        return '水煮'
    return None


def pick_header(grid):
    """找表头行：该行含 >=2 个非原料短文本（样本名），且下方跟着数值块"""
    ncols = len(grid[0])
    best = None
    for gi, row in enumerate(grid):
        txts = []
        for j in range(ncols):
            v = row[j]
            if isinstance(v, str) and v.strip():
                t = norm(v)
                if t and len(t) <= 12 and not isnum(v):
                    txts.append((j, t))
        if len(txts) >= 2:
            nz = 0
            for j in range(ncols):
                if gi + 1 < len(grid) and isnum(grid[gi + 1][j]):
                    nz += 1
            if nz >= 2:
                best = (gi, txts)
                break
    return best


wb = openpyxl.load_workbook(FA, data_only=True)
report = []
for ws in wb.worksheets:
    if ws.title == 'Sheet1':
        continue
    ncols = min(ws.max_column, 16)
    grid = [[ws.cell(r, c).value for c in range(1, ncols + 1)] for r in range(1, (ws.max_row or 1) + 1)]
    hdr = pick_header(grid)
    # 配方列：表头行中以短文本命名的列
    headermap = {}
    if hdr is not None:
        hgi, txts = hdr
        for j, t in txts:
            if classify(t):
                continue
            headermap[j] = t
    else:
        # 无表头：直接用所有有数值但不在第0/1列的列
        pass
    # 收集列组成（累加同代码），原料名在 0/1 列
    colcomp = collections.defaultdict(collections.Counter)
    for gi, row in enumerate(grid):
        if hdr is not None and gi <= hdr[0]:
            continue
        nm = None
        base = 0
        for j in (0, 1, 2):
            v = row[j]
            if isinstance(v, str) and norm(v) and not isnum(v):
                nm = norm(v)
                base = j + 1
                break
        if not nm or len(nm) > 25:
            continue
        if NON_ING.match(nm):
            continue
        if classify(nm):
            continue
        try:
            code = clean_code(fw(nm))
        except Exception:
            code = fw(nm)
        for j in range(base, ncols):
            v = row[j]
            if isnum(v) and v > 0 and abs(v) < 1e6:
                colcomp[j][code] += float(v)
    cols = {j: dict(c) for j, c in colcomp.items() if c}
    # 性能行（T弯G / 水煮）
    perf = collections.defaultdict(dict)
    for gi, row in enumerate(grid):
        key = None
        for j in range(min(4, len(row))):
            t = row[j]
            if isinstance(t, str) and classify(t):
                key = classify(t)
                break
        if key:
            for j in list(cols):
                if j < len(row) and row[j] is not None and str(row[j]).strip() not in ('', 'None'):
                    perf[key][j] = row[j]
    print('=' * 100)
    print('SHEET %-16s 配方列=%s 样本(合并版)=%d' % (
        ws.title, [headermap.get(j, j) for j in sorted(cols)], len(by_sheet.get(ws.title, []))))
    if perf:
        for k, v in perf.items():
            print('   性能行 %s: %s' % (k, {headermap.get(j, j): str(x) for j, x in sorted(v.items())}))
    for j, c in sorted(cols.items()):
        print('   列%2d %-12s n=%2d Σ=%9.3f  %s' % (j, headermap.get(j, j), len(c), sum(c.values()),
              ' '.join('%s=%.3f' % (k, v) for k, v in sorted(c.items())[:6])))
    # merged 样本 → 最优列
    for m in by_sheet.get(ws.title, []):
        mc = {}
        for a, b in m['组分'].items():
            cc = clean_code(fw(a))
            mc[cc] = mc.get(cc, 0.0) + b
        mset = set(mc)
        best = None
        for j, c in cols.items():
            cset = set(c)
            mind = mset - cset
            minc = cset - mset
            dv = [a for a in mset & cset if abs(mc[a] - c[a]) > max(0.05, 0.01 * c[a])]
            score = len(mind) + len(minc) + len(dv)
            if best is None or score < best[0]:
                best = (score, j, mind, minc, dv, c)
        sc, j, mind, minc, dv, c = best
        tag = 'OK' if sc == 0 else 'DIFF'
        print('   [%s] %-42s Σ=%8.3f 列%d: 仅合并版有=%s 仅源列有=%s 值差=%s' % (
            tag, m['样本ID'], sum(mc.values()), j,
            sorted(mind), sorted(minc),
            {a: (round(mc[a], 3), round(c[a], 3)) for a in dv} if dv else ''))