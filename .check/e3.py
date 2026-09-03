# -*- coding: utf-8 -*-
"""检查源文件中重复样本行到底记载了什么（T/M/W 与组分）"""
import openpyxl, re

U = '/workspace/.uploads/'
FILES = {
    '7.26': (U + 'cb7988e8-dc53-4617-9e33-71aa45000fcf_7.26配料测试汇总.xlsx', 'Sheet1'),
    '8.6': (U + 'a7f5f7b0-69c6-4226-bd58-0d02c928336b_8.6配料测试汇总.xlsx', '8.6配料测试汇总'),
    '8.14': (U + '05ddeff0-8071-4cc7-8e93-5df104d14162_8.14配料测试汇总.xlsx', '8.14配料测试汇总'),
}
CHECK = {
    '7.26': ['R01-23', 'R02-16', 'R03-16', 'R4-16', 'R5-06', 'R7-16'],
    '8.6': ['D1-24', 'D2-24', 'D3-24', 'C4-8', 'C5-16', 'D7-35'],
    '8.14': ['D1-24', 'D2-24', 'D3-24', 'D6-6', 'C4-8', 'C5-16', 'C6-5', 'D6-3'],
}
for k, (path, sheet) in FILES.items():
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    print('=' * 110)
    print(f'FILE {k}  sheet={sheet}  (列1-20 前4行表头)')
    print('  ', [ws.cell(1, c).value for c in range(1, 21)])
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(r, 2).value
        if pid is None:
            continue
        pid = str(pid).strip()
        if pid in CHECK.get(k, []):
            # 行全部单元格
            rowvals = [ws.cell(r, c).value for c in range(1, 21)]
            print(f'  row{r}: ID={pid} | 组分={rowvals[3:18]} | T={rowvals[17]} M={rowvals[18]} W={rowvals[19]}')