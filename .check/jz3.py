# -*- coding: utf-8 -*-
"""聚酯金黄：列位对齐 + 重复行求和 + 代码映射，与 merged pkl 严格比对"""
import pickle, re, sys, collections
import openpyxl

sys.path.insert(0, '/workspace/generalization/workbench')
from DataPrepWorkbench import clean_code

F = '/workspace/.uploads/09e51279-275e-4020-b4e4-03288225e478_聚酯金黄-AI(1).xlsx'
D = pickle.load(open('/workspace/generalization/data/merged_data.pkl', 'rb'))
merged = collections.defaultdict(list)
for s in D['all_samples']:
    if s['体系'] != '聚酯金黄':
        continue
    tail = s['样本ID'].split('聚酯金黄-', 1)[1]
    merged[tail[:tail.rfind('-')]].append(s)

NON_ING = re.compile(r'^(合计|粘度|固含|外观|粉末|佳仪滑度|50KG|T弯|MEK|电腐蚀|附着力|硬度|刮伤|BOX|百分比|500克|序号|备注|121|126|度系数|PVC|产品|工艺|基材|单涂|双涂|低价)')


def isnum(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def colname(row, j):
    v = row[j]
    return str(v).strip().replace('\n', '') if isinstance(v, str) else ''


wb = openpyxl.load_workbook(F, data_only=True)
report = []
for ws in wb.worksheets:
    ncols = min(ws.max_column, 12)
    grid = [[ws.cell(r, c).value for c in range(1, ncols + 1)] for r in range(1, ws.max_row + 1)]
    # 样本名列：出现 '百分比'/'500克' 之外的表头行
    hdr_row = None
    for gi, row in enumerate(grid):
        txts = [colname(row, j) for j in range(ncols)]
        if any(t.startswith(('聚酯金黄-', '808-', '3144-')) for t in txts):
            hdr_row = gi
            break
    cols = []
    if hdr_row is not None:
        for j in range(ncols):
            t = colname(grid[hdr_row], j)
            if t and not NON_ING.match(t):
                cols.append((j, t))
    # 采集原料行（列位保留）
    per = {j: collections.defaultdict(float) for j, _ in cols}
    for gi, row in enumerate(grid):
        if gi <= (hdr_row if hdr_row is not None else 1):
            continue
        nm = None
        base = 0
        for j in (0, 1):
            v = row[j]
            if isinstance(v, str) and v.strip() and not isnum(v):
                nm = re.sub(r'\s', '', v)
                base = j + 1
                break
        if nm is None or NON_ING.match(nm) or nm.startswith(('121', '126', '3%醋', '3%盐', '2%柠')):
            continue
        got = False
        for j, _ in cols:
            if j < len(row) and isnum(row[j]) and row[j] > 0:
                try:
                    code = clean_code(nm)
                except Exception:
                    code = nm
                per[j][code] += float(row[j])
                got = True
        if not got:
            pass
    print('=' * 96)
    print('SHEET %-10s 样本列: %s' % (ws.title, [c for _, c in cols]))
    for j, cn in cols:
        tot = sum(per[j].values())
        print('   %-12s 组分数=%2d 合计=%10.4f   %s' % (cn, len(per[j]), tot,
              ' '.join('%s=%g' % (k, v) for k, v in sorted(per[j].items()))))
    ms = merged.get(ws.title, [])
    print('   合并版 %d 条:' % len(ms))
    for k, m in enumerate(ms):
        mc = {a: round(b, 4) for a, b in m['组分'].items()}
        best = None
        for j, cn in cols:
            sc = {a: round(b, 4) for a, b in per[j].items()}
            diff_keys = set(sc) ^ set(mc)
            diff_val = [a for a in set(sc) & set(mc) if abs(sc[a] - mc.get(a, 0)) > 1e-4]
            score = len(diff_keys) + len(diff_val)
            if best is None or score < best[0]:
                best = (score, cn, diff_keys, diff_val, sc)
        s, cn, dk, dv, sc = best
        flag = 'OK' if s == 0 else 'DIFF'
        extra = [a for a in mc if a not in sc]
        missing = [a for a in sc if a not in mc]
        print('      [%s] %-22s 最近列=%-10s 合并版缺:%s 合并版多:%s 值不同:%s' % (
            flag, m['样本ID'], cn, missing, extra,
            {a: (sc[a], mc[a]) for a in dv} if dv else {}))
    report.append((ws.title, [c for _, c in cols], len(ms)))
print()
print('总源配方列数=%d  合并版样本数=%d' % (sum(len(b) for a, b, c in report), sum(c for a, b, c in report)))
