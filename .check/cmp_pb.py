# -*- coding: utf-8 -*-
"""逐sheet逐格比较两个配比方案文件（仅打印差异），定位merged数值来源"""
import openpyxl

F1 = '/workspace/.uploads/b36f4809-d40f-491a-b9d7-2a9c4e359b1b_fbfc94091fa4955969e5d0fff7df0fd6_789507228604997242_m_3NX240913-6C--AI研发26.7.22配比方案.xlsx'
F2 = '/workspace/.uploads/423e234b-3ab5-43a8-add2-4e324135808e_fbfc94091fa4955969e5d0fff7df0fd6_789507228604997242_m_3NX240913-6C--AI研发26.7.22配比方案.xlsx'
w1 = openpyxl.load_workbook(F1, data_only=True)
w2 = openpyxl.load_workbook(F2, data_only=True)
print('w1 sheets:', w1.sheetnames)
print('w2 sheets:', w2.sheetnames)
diffs = 0
for name in w1.sheetnames:
    if name not in w2.sheetnames:
        print('W2 MISSING SHEET:', name)
        continue
    s1, s2 = w1[name], w2[name]
    mr = max(s1.max_row, s2.max_row)
    mc = max(s1.max_column, s2.max_column)
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            a, b = s1.cell(r, c).value, s2.cell(r, c).value
            if isinstance(a, float) and isinstance(b, float) and abs(a - b) < 1e-9:
                continue
            if a == b:
                continue
            diffs += 1
            if diffs <= 60:
                print('%s r%d c%d: %.9r vs %.9r' % (name, r, c, a, b))
print('total diffs:', diffs)