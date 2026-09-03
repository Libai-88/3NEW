# -*- coding: utf-8 -*-
"""聚酯金黄：按列位对齐的独立解析 + 与 merged_data.pkl 比对"""
import pickle, re, collections
import openpyxl

F = '/workspace/.uploads/09e51279-275e-4020-b4e4-03288225e478_聚酯金黄-AI(1).xlsx'
wb = openpyxl.load_workbook(F, data_only=True)
D = pickle.load(open('/workspace/generalization/data/merged_data.pkl', 'rb'))
merged = [s for s in D['all_samples'] if s['体系'] == '聚酯金黄']
by_sheet = collections.defaultdict(list)
for s in merged:
    tail = s['样本ID'].split('聚酯金黄-', 1)[1]
    i = tail.rfind('-')
    by_sheet[tail[:i]].append(s)

HDR_SKIP = {'合计', '粘度', '固含', '外观', '粉末', '佳仪滑度', '50KG', 'T弯', 'MEK', '电腐蚀',
            '附着力', '硬度', '刮伤', 'BOX', '百分比', '500克', '序号', ''}


def isnum(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def classify(t):
    t = str(t or '').replace('\n', '').replace(' ', '')
    if 'T弯' in t:
        return 'T弯'
    if t in ('MEK', 'MEK擦拭'):
        return 'MEK'
    if '水煮' in t and not any(b in t for b in ('BOX', '蒸汽', '盐', '柠檬', '酸', '三合一', 'H', 'S', '铜', '后')):
        return '水煮'
    return None


for ws in wb.worksheets:
    nrows, ncols = ws.max_row, min(ws.max_column, 12)
    grid = [[ws.cell(r, c).value for c in range(1, ncols + 1)] for r in range(1, nrows + 1)]
    # 原料行：col A 或 B 为名称，右侧有数值
    ing = []
    for gi, row in enumerate(grid):
        nm = None
        base = 0
        for j in (0, 1):
            v = row[j]
            if isinstance(v, str) and v.strip() and not isnum(v):
                nm = v.strip().replace('\n', '')
                base = j + 1
                break
        if nm is None or nm in HDR_SKIP or nm.startswith('121'):
            continue
        vals = {j: row[j] for j in range(base, ncols) if isnum(row[j]) and row[j] not in (None,)}
        if not vals:
            continue
        ing.append((gi + 1, nm, vals))
    colset = sorted(set().union(*[set(v[2]) for v in ing]) if ing else set())
    print('=' * 90)
    print('SHEET', ws.title, ' 原料行数=%d  数值列(0-based)=%s' % (len(ing), colset))
    per = collections.defaultdict(dict)
    for r, nm, vals in ing:
        hits = [c for c in colset if c in vals]
        tag = '' if len(hits) == len(colset) else '  <== 仅出现在列 %s（源表该行有空列）' % hits
        for c in hits:
            per[c][nm] = vals[c]
        if len(hits) != len(colset):
            print('   稀疏原料行 r%-3d %-14s -> 列%s = %s%s' % (r, nm, hits, [vals[c] for c in hits], ''))
    for c in colset:
        tot = sum(per[c].values())
        print('   列%d 合计=%.4f 组分数=%d' % (c, tot, len(per[c])))
    # 该 sheet 的性能行
    for gi, row in enumerate(grid):
        key = None
        for j in range(min(4, len(row))):
            v = row[j]
            if isinstance(v, str) and classify(v):
                key = classify(v)
                break
        if key:
            print('   性能行 r%-3d label=%r -> %s : %s' % (gi + 1, row[1] if len(row) > 1 else row[0], key,
                                                            {c: row[c] for c in colset if c < len(row) and row[c] is not None}))
    ms = by_sheet.get(ws.title, [])
    print('   合并版样本 %d 条: %s' % (len(ms), [m['样本ID'].rsplit('-', 1)[1] for m in ms]))
    for k, m in enumerate(ms):
        mc = {re.sub(r'\s', '', a): round(b, 4) for a, b in m['组分'].items()}
        for c in colset:
            sc = {re.sub(r'\s', '', a): round(b, 4) for a, b in per[c].items()}
            if sc == mc:
                print('      %-24s 与列%d 完全一致' % (m['样本ID'], c))
                break
        else:
            print('      %-24s 无匹配列! 组分数=%d 合计=%.4f' % (m['样本ID'], len(mc), sum(mc.values())))
