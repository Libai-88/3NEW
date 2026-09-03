# -*- coding: utf-8 -*-
"""环氧酚醛：全源核对——8.14/8.6/7.26 中的配方+标签 vs merged"""
import pickle, re, collections, statistics
import openpyxl
import numpy as np

U = '/workspace/.uploads/'
FILES = {
    '7.26': (U + 'cb7988e8-dc53-4617-9e33-71aa45000fcf_7.26配料测试汇总.xlsx', 'Sheet1'),
    '8.6': (U + 'a7f5f7b0-69c6-4226-bd58-0d02c928336b_8.6配料测试汇总.xlsx', '8.6配料测试汇总'),
    '8.14': (U + '05ddeff0-8071-4cc7-8e93-5df104d14162_8.14配料测试汇总.xlsx', '8.14配料测试汇总'),
}
NCOL_AMT = (4, 18)  # 用量列: col4..col17
COL_T, COL_M, COL_W = 18, 19, 20


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    m = re.match(r'^(\d+(?:\.\d+)?)\+$', s)
    return float(m.group(1)) if m else None


def parse(path, sheet):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    hdr = {ws.cell(1, c).value: c for c in range(1, 21)}
    amt_cols = [c for c in range(4, 18) if ws.cell(1, c).value]
    recs = {}
    dup = collections.Counter()
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(r, 2).value
        if pid is None:
            continue
        pid = str(pid).strip()
        comp = {}
        for c in amt_cols:
            v = ws.cell(r, c).value
            name = str(ws.cell(1, c).value or '').strip()
            if isinstance(v, (int, float)) and v != 0 and not (isinstance(v, float) and np.isnan(v)):
                comp[name] = round(float(v), 6)
        d = {'T': num(ws.cell(r, COL_T).value), 'M': num(ws.cell(r, COL_M).value), 'W': num(ws.cell(r, COL_W).value),
             'comp': comp, 'row': r}
        if pid in recs:
            dup[pid] += 1
        sub = recs.setdefault(pid, {})
        for k in ('T', 'M', 'W'):
            if sub.get(k) is None and d[k] is not None:
                sub[k] = d[k]
        if not sub.get('comp'):
            sub['comp'] = comp
    return recs, dup


src = {k: parse(*v)[0] for k, v in FILES.items()}
dups = {k: parse(*v)[1] for k, v in FILES.items()}
D = pickle.load(open('/workspace/generalization/data/merged_data.pkl', 'rb'))
ms = {s['样本ID']: s for s in D['all_samples'] if s['体系'] == '环氧酚醛'}
print('merged 环氧酚醛样本数:', len(ms))

# 1) 各源文件中的样本ID集合 vs merged
for k, recs in src.items():
    ids = set(recs)
    print(f'-- {k}: 文件中样本 {len(ids)} 个, 不在 merged 中: {sorted(ids - set(ms))[:20]}')
    print(f'   重复行样本(去重后仍保留首次): {dict(dups[k]) if dups[k] else "无"}')
    print(f'   在merged但文件无这组? merged缺: {len(set(ms)-ids)} 个(merged多)')

# 2) 配方一致性：merged 组分 vs 各文件组分（对每个样本，找与其组分完全一致的源文件）
print()
print('### 配方(组分)一致性核对（merged vs 各文件）')
mismatch = collections.defaultdict(list)
for sid, s in ms.items():
    mc = {k: round(float(v), 2) for k, v in s['组分'].items()}
    sim = []
    for k, recs in src.items():
        if sid not in recs:
            continue
        sc = {kk: round(float(vv), 2) for kk, vv in recs[sid]['comp'].items()}
        same = mc == sc
        sim.append((k, same, sorted(set(sc) ^ set(mc)), {a: (sc.get(a), mc.get(a)) for a in set(sc) & set(mc) if abs(sc[a] - mc[a]) > 0.05}))
    if not sim:
        mismatch['not-in-any-file'].append(sid)
    else:
        for k, same, only, dv in sim:
            if not same:
                mismatch[(k, 'comp-diff')].append((sid, only[:6], dict(list(dv.items())[:4])))
for k, v in mismatch.items():
    if k == 'not-in-any-file':
        print(f'  在任何文件都无配方的样本: {len(v)} 个: {v[:15]}')
    else:
        print(f'  文件{k[0]} 配方与merged不同的样本: {len(v)} 个')
        for sid, only, dv in v[:10]:
            print('      ', sid, '仅文件有=', only, '值差=', dv)

# 3) 缺失标签明细（合并版为空但源文件有值）——重点输出
print()
print('### 合并版缺失标签（源文件有值）——待补全清单')
need = collections.defaultdict(list)
for sid, s in ms.items():
    for tname, key in (('T弯', 'T'), ('MEK', 'M'), ('水煮', 'W')):
        mv = s.get(tname)
        mv = None if mv is None else round(float(mv), 4)
        if mv is not None:
            continue
        vals = {}
        for k, recs in src.items():
            if sid in recs and recs[sid].get(key) is not None:
                vals[k] = recs[sid][key]
        if vals:
            need[(tname, tuple(sorted(vals)))].append((sid, vals))
for k, lst in sorted(need.items()):
    print(f'-- {k[0]}  来源{list(k[1])}: {len(lst)} 条')
    for sid, vals in lst:
        print(f'      {sid:8s} {vals}')