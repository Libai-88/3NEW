# -*- coding: utf-8 -*-
"""
数据整理与特征转换工作台 (Data Prep Workbench) v1.0
===================================================
配套「终极版数据集模板 v3」使用的前置自动化工具，聚焦固定且机械的流程自动化，
减少人工误差。当前阶段包含两大功能（不含模型训练）：

  1. 数据整理：导入多种数据源（模板Excel / 配料测试汇总 / 配比方案 / 聚酯金黄 / 原料数据）
     → 自动识别格式、清洗原料代码（别名映射+大小写统一）、重复测量取均值
     → 导出为终极版模板结构 Excel（原料主数据/配方明细/性能结果/工艺条件）
  2. 特征一键转换：从模板结构数据自动计算特征矩阵
     （组分用量 + 增强描述符 + 显式比例 + SMILES 分子描述符加权聚合）
     → 导出「配方级描述符」与「建模输入」工作表

打包：python -m PyInstaller --onefile --windowed --name 数据整理与特征转换工作台 DataPrepWorkbench.py
"""
import os
import sys
import re
import traceback
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

import numpy as np
import pandas as pd

# 复用工作台核心特征函数（smi_desc 内嵌，避免运行时依赖 RDKit）
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from CoatingModelWorkbench import (
    load_dataset, build_sample_features, build_feature_matrix,
    enhanced_descriptors, explicit_ratios, smi_aggregate,
    canon, ENH_FEATURES, SMI_AGG_KEYS, CONT_DESC, ROLES, RTYPES,
)
from materials import MAT, ALIAS

# ---------- 数据整理：格式识别与解析 ----------
NOISE = {'合计', '固含', '硬度', '刮伤', '度系数最终值', '佳仪滑度'}
NOISE_PREFIX = ('121℃', '121°C', '121℃/60min')


def clean_code(code):
    """清洗原料代码：去换行、别名映射、大小写统一"""
    code = str(code).strip().replace('\n', '')
    if code in ALIAS:
        return ALIAS[code]
    for k in MAT.keys():
        if code.upper() == k.upper():
            return k
    return code


def is_num(v):
    return isinstance(v, (int, float)) and not (isinstance(v, float) and np.isnan(v))


def est_material(code):
    """按代码模式估算新原料描述符（未登记原料自动登记，保证模板完整性）"""
    c = code.upper()
    base = dict(role='助剂', rtype='其他', NV=50, density=1.0, Mw=1000, EEW=0, AV=0, OHV=0,
                amine=0, func=0, Tg=20, bp=200, fp=80, dD=17.5, dP=6.0, dH=8.0, pol=3.0,
                evap=0.1, C=65, H=9, O=25, N=0, S=0, Cl=0,
                fg_epoxy=0, fg_oh=0.02, fg_cooh=0.005, fg_ester=0.2, fg_amine=0,
                fg_amide=0.05, fg_arom=0.2, fg_ether=0.1, wax=0, pig=0)
    if c.startswith('IR') or c.startswith('IA') or 'R170M' in c or '50177' in c or 'FT960' in c or '818' in c:
        base.update(role='树脂', rtype='环氧', NV=55, Mw=1300, EEW=1800, OHV=32, func=2, Tg=68,
                    fg_epoxy=0.056, fg_oh=0.019, fg_arom=0.4, fg_ether=0.15)
    elif c.startswith('RF') or c.startswith('RH'):
        base.update(role='固化剂', rtype='酚醛', NV=60, Mw=900, OHV=280, func=3, Tg=70,
                    fg_oh=0.165, fg_arom=0.6, fg_ether=0.05)
    elif c.startswith('RJ'):
        base.update(role='树脂', rtype='聚酯', NV=60, Mw=3000, OHV=60, func=2, Tg=40,
                    fg_oh=0.035, fg_ester=0.35, fg_arom=0.3)
    elif c.startswith('TF'):
        base.update(role='树脂', rtype='乙烯基', NV=25, Mw=20000, OHV=20, func=1, Tg=55, Cl=30,
                    fg_oh=0.012, fg_arom=0.1)
    elif c.startswith('AL') or c.startswith('AS') or c.startswith('RX'):
        base.update(role='树脂', rtype='丙烯酸', NV=50, Mw=15000, OHV=50, func=1, Tg=45,
                    fg_oh=0.030, fg_ester=0.30, fg_arom=0.25)
    elif (c.startswith('TM') or c.startswith('TZ') or c.startswith('TT')):
        base.update(role='溶剂', rtype='其他', NV=0, Mw=120, Tg=-70, bp=150, fp=40,
                    fg_oh=0.5, fg_ether=0.5, evap=0.3)
    elif c.startswith('RY') or '白浆' in code or '炭黑' in code or 'PVC' in code:
        base.update(role='颜料', rtype='其他', NV=100, density=3.0, Mw=200, pig=100, fg_arom=0.3)
    elif c.startswith('AZ') or c.startswith('BYK') or c.startswith('FL') or c.startswith('RA') or '气硅' in code or 'CAB' in code:
        base.update(role='助剂', rtype='其他', NV=50, Mw=3000, OHV=30, fg_oh=0.02, fg_ester=0.2)
    elif 'DBE' in c or 'DMP' in c or 'DPM' in c or 'MEK' in c or 'MIBK' in c:
        base.update(role='溶剂', rtype='其他', NV=0, Mw=110, Tg=-70, bp=160, fp=40,
                    fg_ester=0.5, fg_ether=0.3, evap=0.5)
    return base


def detect_format(path):
    """识别数据源格式：template / labeled / formulation / materials / unknown"""
    try:
        xl = pd.ExcelFile(path)
        names = [str(s) for s in xl.sheet_names]
        if any('原料主数据' in s and '配方明细' in s for s in names):
            return 'template'
        if any('配方与结果' in s or '配料' in s for s in names):
            return 'labeled'
        if any('原料' in s and '送检' in s for s in names) or any('原材料' in s for s in names):
            return 'materials'
        # 配方方案类：扫描前几行，识别 代码|用量1|用量2... 或 序号|代码|用量1...
        df = xl.parse(xl.sheet_names[0], header=None)
        for i in range(1, min(6, len(df))):
            row = df.iloc[i]
            c0 = row[0]
            c1 = row[1] if df.shape[1] > 1 else None
            c0_txt = str(c0).strip() if pd.notna(c0) else ''
            c1_txt = str(c1).strip() if pd.notna(c1) else ''
            if (not is_num(c0) and 0 < len(c0_txt) <= 20) or \
               (is_num(c0) and c1_txt and not is_num(c1) and 0 < len(c1_txt) <= 20):
                return 'formulation'
        return 'unknown'
    except Exception:
        return 'unknown'


def _read_excel(path):
    """读取 Excel，openpyxl 失败时尝试 LibreOffice 转换"""
    try:
        return pd.ExcelFile(path)
    except Exception:
        import subprocess, tempfile
        tmp = tempfile.mkdtemp()
        try:
            subprocess.run(['soffice', '--headless', '--convert-to', 'xlsx', '--outdir', tmp, path],
                           capture_output=True, timeout=180)
            out = os.path.join(tmp, os.path.basename(path))
            if os.path.exists(out):
                return pd.ExcelFile(out)
        except Exception:
            pass
        raise


def parse_formulation_sheet(df):
    """解析单个配方sheet：结构A(序号|代码|用量1|用量2...) / 结构B(代码|用量1|用量2...)"""
    ing_rows = []
    for i in range(1, len(df)):
        row = df.iloc[i]
        c0 = row[0]
        c1 = row[1] if len(row) > 1 else None
        c0_txt = str(c0).strip() if pd.notna(c0) else ''
        c1_txt = str(c1).strip() if pd.notna(c1) else ''
        c0_is_num = is_num(c0)
        c1_is_num = is_num(c1)
        if not c0_txt or len(c0_txt) > 20:
            continue
        if c0_is_num and (not c1_txt or len(c1_txt) > 20):
            continue
        if not c0_is_num and len(c0_txt) > 20:
            continue
        if c0_is_num and c1_txt and not c1_is_num:
            code = c1_txt
            nums = [v for v in row[2:] if is_num(v) and v > 0]
        elif not c0_is_num and c0_txt:
            code = c0_txt
            nums = [v for v in row[1:] if is_num(v) and v > 0]
        else:
            continue
        if not nums:
            continue
        ing_rows.append((code, nums))
    if not ing_rows:
        return []
    n_var = max(len(n) for _, n in ing_rows)
    formulas = []
    for v in range(n_var):
        comp = {}
        for code, nums in ing_rows:
            if v < len(nums):
                comp[code] = comp.get(code, 0.0) + float(nums[v])
        if comp:
            formulas.append(comp)
    return formulas


def parse_formulation_file(path, system_hint=''):
    """解析配方方案文件（多sheet），返回样本列表"""
    xl = _read_excel(path)
    samples = []
    for s in xl.sheet_names:
        df = xl.parse(s, header=None)
        for comp in parse_formulation_sheet(df):
            comp2 = {}
            for code, amt in comp.items():
                cc = clean_code(code)
                if cc in NOISE or cc.startswith(NOISE_PREFIX) or cc.isdigit():
                    continue
                comp2[cc] = comp2.get(cc, 0.0) + float(amt)
            if not comp2:
                continue
            samples.append({
                '样本ID': f'{system_hint}-{s}-{len(samples)+1}', '体系': system_hint,
                '系列': system_hint, '组分': comp2,
                '烘烤温度': None, '烘烤时间': None,
                'T弯': None, 'MEK': None, '水煮': None,
                '标签状态': '无标签', '来源': os.path.basename(path),
            })
    return samples


def parse_bake(s):
    if pd.isna(s):
        return None, None
    s = str(s)
    temps = re.findall(r'(\d{3})\s*[℃°]', s)
    times = re.findall(r'(\d+)\s*min', s)
    return (int(temps[0]) if temps else None), (int(times[0]) if times else None)


def parse_labeled_file(path):
    """解析配料测试数据汇总V1（有标签）"""
    xl = _read_excel(path)
    sheet = None
    for s in xl.sheet_names:
        if '配方与结果' in s or '配料' in s:
            sheet = s
            break
    if sheet is None:
        sheet = xl.sheet_names[0]
    df = xl.parse(sheet)
    excl = ['批次', '追溯编号', '配方ID', '配方系列', '配方类型', '线棒号', '烘烤条件',
            'T弯(mm)_原始', 'MEK擦拭(次)_原始', '水煮（等级）_原始', '检测指标数量', '检测完整率',
            '检测完整性', '复核状态', '来源文件']
    comp_cols = [c for c in df.columns if c not in excl]
    bake = df['烘烤条件'].apply(parse_bake) if '烘烤条件' in df.columns else None
    samples = []
    for _, row in df.iterrows():
        comp = {}
        for c in comp_cols:
            v = row[c]
            if pd.notna(v) and isinstance(v, (int, float)) and v > 0:
                comp[c] = float(v)
        if not comp:
            continue
        bt = bake.iloc[_] if bake is not None else (None, None)
        def _num(col):
            if col not in df.columns:
                return None
            v = row[col]
            try:
                return float(v)
            except (ValueError, TypeError):
                return None
        t_w = _num('T弯(mm)_原始')
        m_w = _num('MEK擦拭(次)_原始')
        z_w = _num('水煮（等级）_原始')
        samples.append({
            '样本ID': str(row.get('配方ID', f'L{len(samples)+1}')), '体系': '环氧酚醛',
            '系列': str(row.get('配方系列', '')).strip(), '组分': comp,
            '烘烤温度': bt[0], '烘烤时间': bt[1],
            'T弯': t_w, 'MEK': m_w, '水煮': z_w,
            '标签状态': '实测', '来源': os.path.basename(path),
        })
    return samples


def parse_materials_file(path):
    """解析原料数据文件 → mat_lib 补充（支持 原料代码/存货编码 列）"""
    xl = _read_excel(path)
    sheet = None
    for s in xl.sheet_names:
        if '原料主数据' in s:
            sheet = s
            break
    if sheet is None:
        sheet = xl.sheet_names[0]
    df = xl.parse(sheet)
    mat_lib = {}
    code_col = None
    for cand in ['原料代码', '存货编码', '存货代码', '代码']:
        if cand in df.columns:
            code_col = cand
            break
    if code_col is None:
        return mat_lib
    for _, row in df.iterrows():
        code = str(row.get(code_col, '')).strip()
        if not code or code == 'nan':
            continue
        role = str(row.get('角色', '')).strip()
        rtype = str(row.get('树脂类型', '')).strip()
        if role not in ROLES:
            role = est_material(code)['role']
        if rtype not in RTYPES:
            rtype = est_material(code)['rtype']
        m = {'role': role, 'rtype': rtype}
        for d in CONT_DESC:
            v = row.get(d)
            m[d] = float(v) if pd.notna(v) else 0.0
        mat_lib[code] = m
    return mat_lib


def organize_files(file_paths):
    """整理多个数据源 → (mat_lib, samples, perf, proc, report)"""
    mat_lib = dict(MAT)
    samples = {}
    perf = {}
    proc = {}
    report = []
    for path in file_paths:
        kind = detect_format(path)
        base = os.path.basename(path)
        try:
            if kind == 'template':
                ml, sm, pf, pr = load_dataset(path)
                mat_lib.update(ml)
                for sid, s in sm.items():
                    samples.setdefault(sid, s)
                for sid, p in pf.items():
                    perf.setdefault(sid, {}).update(p)
                for sid, p in pr.items():
                    proc.setdefault(sid, p)
                report.append(f'[模板] {base}: 原料{len(ml)} 样本{len(sm)}')
            elif kind == 'labeled':
                for s in parse_labeled_file(path):
                    sid = s['样本ID']
                    samples[sid] = s
                    for tgt, key in [('T弯', 'T弯'), ('MEK擦拭', 'MEK'), ('水煮等级', '水煮')]:
                        v = s.get(key)
                        if v is not None and not (isinstance(v, float) and np.isnan(v)):
                            perf.setdefault(sid, {})[tgt] = float(v)
                    if s.get('烘烤温度') is not None or s.get('烘烤时间') is not None:
                        proc.setdefault(sid, {'烘烤温度': s['烘烤温度'], '烘烤时间': s['烘烤时间']})
                report.append(f'[有标签] {base}: 样本{sum(1 for s in samples.values() if s.get("来源")==base)}')
            elif kind == 'formulation':
                sys_hint = '环氧-配比方案' if '配比' in base else ('聚酯金黄' if '聚酯' in base else '配方')
                for s in parse_formulation_file(path, sys_hint):
                    samples[s['样本ID']] = s
                report.append(f'[配方方案] {base}: 样本{sum(1 for s in samples.values() if s.get("来源")==base)}')
            elif kind == 'materials':
                ml = parse_materials_file(path)
                added = [c for c in ml if c not in mat_lib]
                for c in added:
                    mat_lib[c] = ml[c]
                report.append(f'[原料数据] {base}: 新增原料{len(added)}（已有{len(ml)-len(added)}种保留原描述符）')
            else:
                report.append(f'[未识别] {base}: 无法自动识别格式，跳过')
        except Exception as e:
            report.append(f'[失败] {base}: {e}')
    # 未登记原料自动估算登记（保证模板完整性）
    all_codes = set(c for s in samples.values() for c in s['组分'])
    new_codes = sorted(c for c in all_codes if c not in mat_lib)
    for c in new_codes:
        mat_lib[c] = est_material(c)
    if new_codes:
        report.append(f'自动估算登记 {len(new_codes)} 种新原料：{new_codes[:10]}{"..." if len(new_codes) > 10 else ""}')
    # 按样本ID去重（重复测量取均值，诚实评估协议）
    def _mean(vals):
        vals = [v for v in vals if v is not None and not (isinstance(v, float) and np.isnan(v))]
        return float(np.mean(vals)) if vals else None
    agg = {}
    for sid, s in samples.items():
        if sid not in agg:
            agg[sid] = dict(s)
            agg[sid]['_T弯'] = [s.get('T弯')]; agg[sid]['_MEK'] = [s.get('MEK')]; agg[sid]['_水煮'] = [s.get('水煮')]
        else:
            agg[sid]['_T弯'].append(s.get('T弯')); agg[sid]['_MEK'].append(s.get('MEK')); agg[sid]['_水煮'].append(s.get('水煮'))
    dedup = []
    for sid, s in agg.items():
        s = dict(s)
        s['T弯'] = _mean(s.pop('_T弯')); s['MEK'] = _mean(s.pop('_MEK')); s['水煮'] = _mean(s.pop('_水煮'))
        dedup.append(s)
    n_dup = len(samples) - len(dedup)
    if n_dup > 0:
        report.append(f'按样本ID去重: {len(samples)} → {len(dedup)}（重复测量取均值）')
    samples = {s['样本ID']: s for s in dedup}
    return mat_lib, samples, perf, proc, report


# ---------- 导出：终极版模板结构 ----------
def export_template(mat_lib, samples, perf, proc, out_path):
    """导出终极版模板结构 Excel（原料主数据/配方明细/性能结果/工艺条件）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    HDR_FILL = PatternFill('solid', fgColor='1F2937')
    HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    ZEBRA_1 = PatternFill('solid', fgColor='FFFFFF')
    ZEBRA_2 = PatternFill('solid', fgColor='F7F9FC')
    THIN = Side(style='thin', color='D9DEE7')
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    BODY_FONT = Font(name='Arial', size=10)
    TITLE_FONT = Font(name='Arial', size=14, bold=True, color='1F2937')
    BOLD_FONT = Font(name='Arial', size=10, bold=True)

    def style_table(ws, header_row, n_cols, n_rows, kpi_cols=None):
        for c in range(1, n_cols + 1):
            cell = ws.cell(header_row, c)
            cell.fill = HDR_FILL; cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = BORDER
        for r in range(header_row + 1, header_row + n_rows + 1):
            fill = ZEBRA_1 if (r - header_row) % 2 == 0 else ZEBRA_2
            for c in range(1, n_cols + 1):
                cell = ws.cell(r, c)
                cell.fill = fill; cell.font = BODY_FONT; cell.border = BORDER
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        if kpi_cols:
            for c in kpi_cols:
                for r in range(header_row + 1, header_row + n_rows + 1):
                    ws.cell(r, c).fill = PatternFill('solid', fgColor='EAF2FF')
                    ws.cell(r, c).font = BOLD_FONT

    wb = Workbook()
    ws = wb.active; ws.title = '使用说明'
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 120
    lines = [
        ('整理后涂料配方-性能数据集（终极版模板 v3 结构）', ''),
        ('', ''),
        ('一、数据构成', '本文件由「数据整理与特征转换工作台」自动生成：'),
        ('', f'  · 原料主数据 {len(mat_lib)} 种'),
        ('', f'  · 配方样本 {len(samples)} 个'),
        ('', f'  · 性能记录 {sum(len(v) for v in perf.values())} 条'),
        ('', ''),
        ('二、工作表说明', '1. 原料主数据：原料描述符（建模特征基础）。'),
        ('', '2. 配方明细：长表，每行=一个样本中的一个组分（含系列）。'),
        ('', '3. 性能结果：实测性能（T弯/MEK擦拭/水煮等级）。'),
        ('', '4. 工艺条件：烘烤温度/时间等工艺参数。'),
        ('', '5. 配方级描述符：样本级特征矩阵（由特征一键转换生成）。'),
        ('', '6. 建模输入：宽表，一行=一个样本，特征+目标。'),
        ('', ''),
        ('三、使用建议', '1. 新增原料请在「原料主数据」补充真实描述符（SDS/TDS）。'),
        ('', '2. 用「特征一键转换」生成配方级描述符与建模输入。'),
    ]
    r = 1
    for text, _ in lines:
        cell = ws.cell(r, 2, text)
        if r == 1:
            cell.font = TITLE_FONT
        elif text.startswith(('一、', '二、', '三、')):
            cell.font = BOLD_FONT
        else:
            cell.font = BODY_FONT
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        r += 1

    # 原料主数据
    ws = wb.create_sheet('原料主数据')
    mat_headers = ['原料代码', '原料名称', '所属体系', '角色', '树脂类型', 'SMILES', '描述符状态'] + \
                  ['固含NV(%)', '密度(g/cm³)', '分子量(g/mol)', '环氧当量EEW(g/eq)', '酸值AV(mgKOH/g)',
                   '羟值OHV(mgKOH/g)', '胺值(mgKOH/g)', '官能度', 'Tg(℃)', '沸点(℃)', '闪点(℃)',
                   'Hansen δD', 'Hansen δP', 'Hansen δH', '极性指数', '相对挥发速率',
                   'C(%)', 'H(%)', 'O(%)', 'N(%)', 'S(%)', 'Cl(%)',
                   '环氧基(mol/100g)', '羟基(mol/100g)', '羧基(mol/100g)', '酯基(mol/100g)', '胺基(mol/100g)',
                   '酰胺(mol/100g)', '芳香环(mol/100g)', '醚键(mol/100g)', '蜡含量(%)', '颜料含量(%)', '数据来源', '备注']
    ws.append(mat_headers)
    for code, d in mat_lib.items():
        row = [code, code, '多体系', d['role'], d['rtype'], '', '已计算']
        for k in CONT_DESC:
            row.append(d.get(k, 0.0))
        row.append('类别典型值/文件信息'); row.append('')
        ws.append(row)
    style_table(ws, 1, len(mat_headers), len(mat_lib), kpi_cols=[8, 11, 12, 13, 14, 15, 30, 31, 32, 33, 34, 35, 36, 37])
    widths = [16, 20, 10, 8, 9, 14, 10] + [10] * 32 + [14, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    # 配方明细
    ws = wb.create_sheet('配方明细')
    det_headers = ['样本ID', '系列', '体系', '原料代码', '用量(g)', '角色', '树脂类型', '标签状态']
    ws.append(det_headers)
    det_rows = 0
    for s in samples.values():
        for code, amt in s['组分'].items():
            d = mat_lib.get(code)
            if d is None:
                continue
            ws.append([s['样本ID'], s['系列'], s['体系'], code, round(float(amt), 4), d['role'], d['rtype'], s['标签状态']])
            det_rows += 1
    style_table(ws, 1, len(det_headers), det_rows, kpi_cols=[4])
    for i, w in enumerate([14, 12, 16, 10, 8, 9, 10, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    # 性能结果
    ws = wb.create_sheet('性能结果')
    perf_headers = ['样本ID', '体系', '目标属性', '测试值', '单位', '标签状态', '标签来源', '测试条件']
    ws.append(perf_headers)
    perf_rows = 0
    for sid, p in perf.items():
        s = samples.get(sid, {})
        cond = f"{s.get('烘烤温度')}℃ {s.get('烘烤时间')}min" if s.get('烘烤温度') else ''
        for tgt, val in p.items():
            if val is not None and not (isinstance(val, float) and np.isnan(val)):
                unit = {'T弯': 'mm', 'MEK擦拭': '次', '水煮等级': '级'}.get(tgt, '')
                ws.append([sid, s.get('体系', ''), tgt, round(float(val), 4), unit, '实测', '实验室', cond])
                perf_rows += 1
    style_table(ws, 1, len(perf_headers), perf_rows, kpi_cols=[4])
    for i, w in enumerate([14, 12, 12, 10, 8, 10, 10, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    # 工艺条件
    ws = wb.create_sheet('工艺条件')
    proc_headers = ['样本ID', '体系', '烘烤温度(℃)', '烘烤时间(min)', '标签状态']
    ws.append(proc_headers)
    proc_rows = 0
    for sid, s in samples.items():
        p = proc.get(sid, {})
        ws.append([sid, s['体系'], p.get('烘烤温度'), p.get('烘烤时间'), s['标签状态']])
        proc_rows += 1
    style_table(ws, 1, len(proc_headers), proc_rows)
    for i, w in enumerate([14, 12, 12, 12, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    wb.save(out_path)
    return out_path


# ---------- 特征一键转换 ----------
def convert_features(mat_lib, samples, perf, proc):
    """计算特征矩阵（组分+增强描述符+比例+SMILES）→ (X, ids, series, feat_names)"""
    present_codes = sorted(set(canon(str(c).strip()) for s in samples.values() for c in s['组分']))
    rows, ids, series = [], [], []
    for sid, s in samples.items():
        p = proc.get(sid, {})
        row = build_sample_features(s['组分'], mat_lib, present_codes,
                                    bake_temp=p.get('烘烤温度'), bake_time=p.get('烘烤时间'))
        if row is None:
            continue
        rows.append(row); ids.append(sid); series.append(s.get('系列', ''))
    X = np.array(rows)
    feat_names = present_codes + ['烘烤温度', '烘烤时间'] + ENH_FEATURES + \
                 [f'r_{i}' for i in range(22)] + SMI_AGG_KEYS
    return X, ids, series, feat_names


def export_features(mat_lib, samples, perf, proc, out_path):
    """导出配方级描述符 + 建模输入 工作表到 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    X, ids, series, feat_names = convert_features(mat_lib, samples, perf, proc)
    HDR_FILL = PatternFill('solid', fgColor='1F2937')
    HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    ZEBRA_1 = PatternFill('solid', fgColor='FFFFFF')
    ZEBRA_2 = PatternFill('solid', fgColor='F7F9FC')
    THIN = Side(style='thin', color='D9DEE7')
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    BODY_FONT = Font(name='Arial', size=10)

    def style_table(ws, header_row, n_cols, n_rows):
        for c in range(1, n_cols + 1):
            cell = ws.cell(header_row, c)
            cell.fill = HDR_FILL; cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = BORDER
        for r in range(header_row + 1, header_row + n_rows + 1):
            fill = ZEBRA_1 if (r - header_row) % 2 == 0 else ZEBRA_2
            for c in range(1, n_cols + 1):
                cell = ws.cell(r, c)
                cell.fill = fill; cell.font = BODY_FONT; cell.border = BORDER
                cell.alignment = Alignment(vertical='center', wrap_text=True)

    wb = Workbook()
    ws = wb.active; ws.title = '配方级描述符'
    ws.append(['样本ID', '系列', '体系', '标签状态'] + feat_names)
    for i, sid in enumerate(ids):
        s = samples.get(sid, {})
        status = '实测' if s.get('标签状态') == '实测' else '无标签'
        ws.append([sid, series[i], s.get('体系', ''), status] +
                  [round(float(v), 6) for v in X[i]])
    style_table(ws, 1, len(feat_names) + 4, len(ids))
    for i, w in enumerate([14, 10, 12, 10] + [10] * len(feat_names), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'E2'

    ws = wb.create_sheet('建模输入')
    ws.append(['样本ID', '系列', '体系', '标签状态', 'T弯实测', 'MEK实测', '水煮实测'] + feat_names)
    for i, sid in enumerate(ids):
        s = samples.get(sid, {})
        status = '实测' if s.get('标签状态') == '实测' else '无标签'
        t_w = s.get('T弯'); m_w = s.get('MEK'); z_w = s.get('水煮')
        ws.append([sid, series[i], s.get('体系', ''), status,
                   round(float(t_w), 4) if t_w is not None else '',
                   round(float(m_w), 4) if m_w is not None else '',
                   round(float(z_w), 4) if z_w is not None else ''] +
                  [round(float(v), 6) for v in X[i]])
    style_table(ws, 1, len(feat_names) + 7, len(ids))
    for i, w in enumerate([14, 10, 12, 10, 10, 10, 10] + [10] * len(feat_names), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'H2'

    wb.save(out_path)
    return out_path, X.shape


# ---------- GUI ----------
class DataPrepApp:
    def __init__(self, root):
        self.root = root
        root.title('数据整理与特征转换工作台 v1.0')
        root.geometry('980x680')
        root.minsize(820, 560)

        self.files = []
        self.mat_lib = None
        self.samples = None
        self.perf = None
        self.proc = None

        nb = ttk.Notebook(root)
        nb.pack(fill='both', expand=True, padx=8, pady=8)

        # ===== Tab1 数据整理 =====
        tab1 = ttk.Frame(nb)
        nb.add(tab1, text='数据整理')
        top1 = ttk.Frame(tab1); top1.pack(fill='x', padx=8, pady=8)
        ttk.Label(top1, text='数据源文件（支持：模板Excel / 配料测试汇总 / 配比方案 / 聚酯金黄 / 原料数据）').pack(anchor='w')
        btns1 = ttk.Frame(tab1); btns1.pack(fill='x', padx=8)
        ttk.Button(btns1, text='添加文件', command=self.add_files).pack(side='left', padx=4)
        ttk.Button(btns1, text='移除选中', command=self.remove_selected).pack(side='left', padx=4)
        ttk.Button(btns1, text='清空', command=self.clear_files).pack(side='left', padx=4)
        ttk.Button(btns1, text='一键整理并导出模板', command=self.organize).pack(side='right', padx=4)
        self.file_list = tk.Listbox(tab1, height=7)
        self.file_list.pack(fill='x', padx=8, pady=8)
        self.log1 = scrolledtext.ScrolledText(tab1, height=16, state='disabled')
        self.log1.pack(fill='both', expand=True, padx=8, pady=8)

        # ===== Tab2 特征一键转换 =====
        tab2 = ttk.Frame(nb)
        nb.add(tab2, text='特征一键转换')
        top2 = ttk.Frame(tab2); top2.pack(fill='x', padx=8, pady=8)
        ttk.Label(top2, text='数据文件（终极版模板结构 Excel，可先用「数据整理」生成）').pack(anchor='w')
        btns2 = ttk.Frame(tab2); btns2.pack(fill='x', padx=8)
        ttk.Button(btns2, text='选择数据文件', command=self.load_data_file).pack(side='left', padx=4)
        ttk.Button(btns2, text='一键计算特征并导出', command=self.convert).pack(side='right', padx=4)
        self.data_lbl = ttk.Label(tab2, text='未加载数据文件')
        self.data_lbl.pack(anchor='w', padx=8, pady=4)
        self.log2 = scrolledtext.ScrolledText(tab2, height=20, state='disabled')
        self.log2.pack(fill='both', expand=True, padx=8, pady=8)

    def log(self, box, text):
        box.configure(state='normal')
        box.insert('end', text + '\n')
        box.see('end')
        box.configure(state='disabled')
        self.root.update_idletasks()

    # ---- Tab1 ----
    def add_files(self):
        paths = filedialog.askopenfilenames(title='选择数据源文件', filetypes=[('Excel文件', '*.xlsx'), ('所有文件', '*.*')])
        for p in paths:
            if p not in self.files:
                self.files.append(p)
                self.file_list.insert('end', os.path.basename(p))

    def remove_selected(self):
        sel = self.file_list.curselection()
        for i in reversed(sel):
            self.files.pop(i)
            self.file_list.delete(i)

    def clear_files(self):
        self.files = []
        self.file_list.delete(0, 'end')

    def organize(self):
        if not self.files:
            messagebox.showwarning('提示', '请先添加数据源文件')
            return
        self.log1.delete('1.0', 'end')
        threading.Thread(target=self._organize_worker, daemon=True).start()

    def _organize_worker(self):
        try:
            self.log(self.log1, '开始整理数据...')
            mat_lib, samples, perf, proc, report = organize_files(self.files)
            self.mat_lib, self.samples, self.perf, self.proc = mat_lib, samples, perf, proc
            for line in report:
                self.log(self.log1, line)
            n_lab = sum(1 for s in samples.values() if s.get('标签状态') == '实测')
            n_unlab = len(samples) - n_lab
            self.log(self.log1, f'整理完成：原料{len(mat_lib)}种，样本{len(samples)}个（实测{n_lab}，无标签{n_unlab}），性能{sum(len(v) for v in perf.values())}条')
            # 未登记原料检查
            unreg = sorted(set(c for s in samples.values() for c in s['组分']) - set(mat_lib.keys()))
            if unreg:
                self.log(self.log1, f'警告：{len(unreg)} 种原料未在原料主数据登记：{unreg}')
            else:
                self.log(self.log1, '原料代码全部已登记 ✓')
            save = filedialog.asksaveasfilename(defaultextension='.xlsx', initialfile='整理后数据集.xlsx',
                                                filetypes=[('Excel文件', '*.xlsx')])
            if save:
                export_template(mat_lib, samples, perf, proc, save)
                self.log(self.log1, f'已导出模板结构数据: {save}')
        except Exception as e:
            self.log(self.log1, '整理失败: ' + traceback.format_exc())
            self.root.after(0, lambda: messagebox.showerror('整理失败', str(e)))

    # ---- Tab2 ----
    def load_data_file(self):
        path = filedialog.askopenfilename(title='选择数据文件', filetypes=[('Excel文件', '*.xlsx'), ('所有文件', '*.*')])
        if not path:
            return
        try:
            self.mat_lib, self.samples, self.perf, self.proc = load_dataset(path)
            self.data_lbl.config(text=os.path.basename(path))
            self.log(self.log2, f'数据加载成功：原料{len(self.mat_lib)}种，样本{len(self.samples)}个')
        except Exception as e:
            messagebox.showerror('加载失败', str(e))
            self.log(self.log2, '加载失败: ' + traceback.format_exc())

    def convert(self):
        if self.mat_lib is None or self.samples is None:
            messagebox.showwarning('提示', '请先加载数据文件')
            return
        self.log2.delete('1.0', 'end')
        threading.Thread(target=self._convert_worker, daemon=True).start()

    def _convert_worker(self):
        try:
            self.log(self.log2, '开始计算特征矩阵...')
            X, ids, series, feat_names = convert_features(self.mat_lib, self.samples, self.perf, self.proc)
            self.log(self.log2, f'特征矩阵：{X.shape[0]} 样本 × {X.shape[1]} 特征')
            self.log(self.log2, f'特征构成：组分用量{len([c for c in feat_names if not c.startswith(("w_","s_","r_")) and c not in ("烘烤温度","烘烤时间")])} + 增强描述符{len(ENH_FEATURES)} + 比例22 + SMILES聚合{len(SMI_AGG_KEYS)}')
            save = filedialog.asksaveasfilename(defaultextension='.xlsx', initialfile='特征矩阵.xlsx',
                                                filetypes=[('Excel文件', '*.xlsx')])
            if save:
                out, shape = export_features(self.mat_lib, self.samples, self.perf, self.proc, save)
                self.log(self.log2, f'已导出特征矩阵: {out} ({shape[0]}×{shape[1]})')
        except Exception as e:
            self.log(self.log2, '特征计算失败: ' + traceback.format_exc())
            self.root.after(0, lambda: messagebox.showerror('特征计算失败', str(e)))


def main():
    root = tk.Tk()
    app = DataPrepApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()
