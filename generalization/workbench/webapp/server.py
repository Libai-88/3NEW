# -*- coding: utf-8 -*-
"""
数据整理与特征转换工作台 · Web 版 (Data Prep Workbench Web) v1.0
================================================================
配套「终极版数据集模板 v3」使用的前置自动化工具，聚焦固定且机械的流程自动化，
减少人工误差。当前阶段包含三大功能（不含模型训练）：

  1. 数据整理：导入多种数据源（模板Excel / 配料测试汇总 / 配比方案 / 聚酯金黄 / 原料数据）
     → 自动识别格式、清洗原料代码（别名映射+大小写统一）、重复测量取均值
     → 导出为终极版模板结构 Excel（原料主数据/配方明细/性能结果/工艺条件）
  2. 特征一键转换：从模板结构数据自动计算特征矩阵
     （组分用量 + 增强描述符 + 显式比例 + SMILES 分子描述符加权聚合）
     → 导出「配方级描述符」与「建模输入」工作表
  3. 辅助数据录入：表单式录入配方/原料/性能，自动校验并生成模板结构数据

启动：python server.py  →  浏览器打开 http://127.0.0.1:8765
零第三方 Web 依赖（仅需 Python 标准库 + pandas/numpy/openpyxl）。
"""
import os
import sys
import io
import re
import json
import base64
import traceback
import threading
import webbrowser
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlparse, parse_qs

import numpy as np
import pandas as pd

# 复用工作台核心特征函数（smi_desc 内嵌，避免运行时依赖 RDKit）
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from CoatingModelWorkbench import (
    load_dataset, build_sample_features, build_feature_matrix,
    enhanced_descriptors, explicit_ratios, smi_aggregate,
    canon, ENH_FEATURES, SMI_AGG_KEYS, CONT_DESC, ROLES, RTYPES,
)
from materials import MAT, ALIAS
from flow import suggest_type, validate_file, build_manifest, build_acquisition_plan, build_readiness_report, FILE_TYPES

# ---------- 数据整理：格式识别与解析（复用 DataPrepWorkbench 逻辑） ----------
NOISE = {'合计', '固含', '硬度', '刮伤', '度系数最终值', '佳仪滑度'}
NOISE_PREFIX = ('121℃', '121°C', '121℃/60min')


def clean_code(code):
    code = str(code).strip().replace('\n', '')
    if code in ALIAS:
        return ALIAS[code]
    for k in MAT.keys():
        if code.upper() == k.upper():
            return k
    return code


def is_num(v):
    return isinstance(v, (int, float)) and not (isinstance(v, float) and np.isnan(v))


def est_material(code):
    """按代码模式估算新原料描述符（未登记原料自动登记，保证模板完整性）"""
    c = code.upper()
    base = dict(role='助剂', rtype='其他', NV=50, density=1.0, Mw=1000, EEW=0, AV=0, OHV=0,
                amine=0, func=0, Tg=20, bp=200, fp=80, dD=17.5, dP=6.0, dH=8.0, pol=3.0,
                evap=0.1, C=65, H=9, O=25, N=0, S=0, Cl=0,
                fg_epoxy=0, fg_oh=0.02, fg_cooh=0.005, fg_ester=0.2, fg_amine=0,
                fg_amide=0.05, fg_arom=0.2, fg_ether=0.1, wax=0, pig=0)
    if c.startswith('IR') or c.startswith('IA') or 'R170M' in c or '50177' in c or 'FT960' in c or '818' in c:
        base.update(role='树脂', rtype='环氧', NV=55, Mw=1300, EEW=1800, OHV=32, func=2, Tg=68,
                    fg_epoxy=0.056, fg_oh=0.019, fg_arom=0.4, fg_ether=0.15)
    elif c.startswith('RF') or c.startswith('RH'):
        base.update(role='固化剂', rtype='酚醛', NV=60, Mw=900, OHV=280, func=3, Tg=70,
                    fg_oh=0.165, fg_arom=0.6, fg_ether=0.05)
    elif c.startswith('RJ'):
        base.update(role='树脂', rtype='聚酯', NV=60, Mw=3000, OHV=60, func=2, Tg=40,
                    fg_oh=0.035, fg_ester=0.35, fg_arom=0.3)
    elif c.startswith('TF'):
        base.update(role='树脂', rtype='乙烯基', NV=25, Mw=20000, OHV=20, func=1, Tg=55, Cl=30,
                    fg_oh=0.012, fg_arom=0.1)
    elif c.startswith('AL') or c.startswith('AS') or c.startswith('RX'):
        base.update(role='树脂', rtype='丙烯酸', NV=50, Mw=15000, OHV=50, func=1, Tg=45,
                    fg_oh=0.030, fg_ester=0.30, fg_arom=0.25)
    elif (c.startswith('TM') or c.startswith('TZ') or c.startswith('TT')):
        base.update(role='溶剂', rtype='其他', NV=0, Mw=120, Tg=-70, bp=150, fp=40,
                    fg_oh=0.5, fg_ether=0.5, evap=0.3)
    elif c.startswith('RY') or '白浆' in code or '炭黑' in code or 'PVC' in code:
        base.update(role='颜料', rtype='其他', NV=100, density=3.0, Mw=200, pig=100, fg_arom=0.3)
    elif c.startswith('AZ') or c.startswith('BYK') or c.startswith('FL') or c.startswith('RA') or '气硅' in code or 'CAB' in code:
        base.update(role='助剂', rtype='其他', NV=50, Mw=3000, OHV=30, fg_oh=0.02, fg_ester=0.2)
    elif 'DBE' in c or 'DMP' in c or 'DPM' in c or 'MEK' in c or 'MIBK' in c:
        base.update(role='溶剂', rtype='其他', NV=0, Mw=110, Tg=-70, bp=160, fp=40,
                    fg_ester=0.5, fg_ether=0.3, evap=0.5)
    return base


def detect_format(path):
    """识别数据源格式：template / labeled / formulation / materials / unknown"""
    try:
        xl = pd.ExcelFile(path)
        names = [str(s) for s in xl.sheet_names]
        if any('原料主数据' in s for s in names) and any('配方明细' in s for s in names):
            return 'template'
        if any('配方与结果' in s or '配料' in s for s in names):
            return 'labeled'
        if any('原料' in s and '送检' in s for s in names) or any('原材料' in s for s in names):
            return 'materials'
        df = xl.parse(xl.sheet_names[0], header=None)
        for i in range(1, min(6, len(df))):
            row = df.iloc[i]
            c0 = row[0]
            c1 = row[1] if df.shape[1] > 1 else None
            c0_txt = str(c0).strip() if pd.notna(c0) else ''
            c1_txt = str(c1).strip() if pd.notna(c1) else ''
            if (not is_num(c0) and 0 < len(c0_txt) <= 20) or \
               (is_num(c0) and c1_txt and not is_num(c1) and 0 < len(c1_txt) <= 20):
                return 'formulation'
        return 'unknown'
    except Exception:
        return 'unknown'


def parse_formulation_sheet(df):
    """解析单个配方sheet：结构A(序号|代码|用量1|用量2...) / 结构B(代码|用量1|用量2...)"""
    ing_rows = []
    for i in range(1, len(df)):
        row = df.iloc[i]
        c0 = row[0]
        c1 = row[1] if len(row) > 1 else None
        c0_txt = str(c0).strip() if pd.notna(c0) else ''
        c1_txt = str(c1).strip() if pd.notna(c1) else ''
        c0_is_num = is_num(c0)
        c1_is_num = is_num(c1)
        if not c0_txt or len(c0_txt) > 20:
            continue
        if c0_is_num and (not c1_txt or len(c1_txt) > 20):
            continue
        if not c0_is_num and len(c0_txt) > 20:
            continue
        if c0_is_num and c1_txt and not c1_is_num:
            code = c1_txt
            nums = [v for v in row[2:] if is_num(v) and v > 0]
        elif not c0_is_num and c0_txt:
            code = c0_txt
            nums = [v for v in row[1:] if is_num(v) and v > 0]
        else:
            continue
        if not nums:
            continue
        ing_rows.append((code, nums))
    if not ing_rows:
        return []
    n_var = max(len(n) for _, n in ing_rows)
    formulas = []
    for v in range(n_var):
        comp = {}
        for code, nums in ing_rows:
            if v < len(nums):
                comp[code] = comp.get(code, 0.0) + float(nums[v])
        if comp:
            formulas.append(comp)
    return formulas


def parse_formulation_file(path, system_hint=''):
    """解析配方方案文件（多sheet），返回样本列表"""
    xl = pd.ExcelFile(path)
    samples = []
    for s in xl.sheet_names:
        df = xl.parse(s, header=None)
        for comp in parse_formulation_sheet(df):
            comp2 = {}
            for code, amt in comp.items():
                cc = clean_code(code)
                if cc in NOISE or cc.startswith(NOISE_PREFIX) or cc.isdigit():
                    continue
                comp2[cc] = comp2.get(cc, 0.0) + float(amt)
            if not comp2:
                continue
            samples.append({
                '样本ID': f'{system_hint}-{s}-{len(samples)+1}', '体系': system_hint,
                '系列': system_hint, '组分': comp2,
                '烘烤温度': None, '烘烤时间': None,
                'T弯': None, 'MEK': None, '水煮': None,
                '标签状态': '无标签', '来源': os.path.basename(path),
            })
    return samples


def parse_bake(s):
    if pd.isna(s):
        return None, None
    s = str(s)
    temps = re.findall(r'(\d{3})\s*[℃°]', s)
    times = re.findall(r'(\d+)\s*min', s)
    return (int(temps[0]) if temps else None), (int(times[0]) if times else None)


def parse_labeled_file(path):
    """解析配料测试数据汇总V1（有标签）"""
    xl = pd.ExcelFile(path)
    sheet = None
    for s in xl.sheet_names:
        if '配方与结果' in s or '配料' in s:
            sheet = s
            break
    if sheet is None:
        sheet = xl.sheet_names[0]
    df = xl.parse(sheet)
    excl = ['批次', '追溯编号', '配方ID', '配方系列', '配方类型', '线棒号', '烘烤条件',
            'T弯(mm)_原始', 'MEK擦拭(次)_原始', '水煮（等级）_原始', '检测指标数量', '检测完整率',
            '检测完整性', '复核状态', '来源文件']
    comp_cols = [c for c in df.columns if c not in excl]
    bake = df['烘烤条件'].apply(parse_bake) if '烘烤条件' in df.columns else None
    samples = []
    for idx, row in df.iterrows():
        comp = {}
        for c in comp_cols:
            v = row[c]
            if pd.notna(v) and isinstance(v, (int, float)) and v > 0:
                comp[c] = float(v)
        if not comp:
            continue
        bt = bake.iloc[idx] if bake is not None else (None, None)

        def _num(col):
            if col not in df.columns:
                return None
            v = row[col]
            try:
                return float(v)
            except (ValueError, TypeError):
                return None
        t_w = _num('T弯(mm)_原始')
        m_w = _num('MEK擦拭(次)_原始')
        z_w = _num('水煮（等级）_原始')
        samples.append({
            '样本ID': str(row.get('配方ID', f'L{len(samples)+1}')), '体系': '环氧酚醛',
            '系列': str(row.get('配方系列', '')).strip(), '组分': comp,
            '烘烤温度': bt[0], '烘烤时间': bt[1],
            'T弯': t_w, 'MEK': m_w, '水煮': z_w,
            '标签状态': '实测', '来源': os.path.basename(path),
        })
    return samples


def parse_materials_file(path):
    """解析原料数据文件 → mat_lib 补充（支持 原料代码/存货编码 列）"""
    xl = pd.ExcelFile(path)
    sheet = None
    for s in xl.sheet_names:
        if '原料主数据' in s:
            sheet = s
            break
    if sheet is None:
        sheet = xl.sheet_names[0]
    df = xl.parse(sheet)
    mat_lib = {}
    code_col = None
    for cand in ['原料代码', '存货编码', '存货代码', '代码']:
        if cand in df.columns:
            code_col = cand
            break
    if code_col is None:
        return mat_lib
    for _, row in df.iterrows():
        code = str(row.get(code_col, '')).strip()
        if not code or code == 'nan':
            continue
        role = str(row.get('角色', '')).strip()
        rtype = str(row.get('树脂类型', '')).strip()
        if role not in ROLES:
            role = est_material(code)['role']
        if rtype not in RTYPES:
            rtype = est_material(code)['rtype']
        m = {'role': role, 'rtype': rtype}
        for d in CONT_DESC:
            v = row.get(d)
            m[d] = float(v) if pd.notna(v) else 0.0
        mat_lib[code] = m
    return mat_lib


def organize_files(file_paths):
    """整理多个数据源 → (mat_lib, samples, perf, proc, report)"""
    mat_lib = dict(MAT)
    samples = {}
    perf = {}
    proc = {}
    report = []
    for path in file_paths:
        kind = detect_format(path)
        base = os.path.basename(path)
        try:
            if kind == 'template':
                ml, sm, pf, pr = load_dataset(path)
                mat_lib.update(ml)
                for sid, s in sm.items():
                    samples.setdefault(sid, s)
                for sid, p in pf.items():
                    perf.setdefault(sid, {}).update(p)
                for sid, p in pr.items():
                    proc.setdefault(sid, p)
                report.append(f'[模板] {base}: 原料{len(ml)} 样本{len(sm)}')
            elif kind == 'labeled':
                for s in parse_labeled_file(path):
                    sid = s['样本ID']
                    samples[sid] = s
                    for tgt, key in [('T弯', 'T弯'), ('MEK擦拭', 'MEK'), ('水煮等级', '水煮')]:
                        v = s.get(key)
                        if v is not None and not (isinstance(v, float) and np.isnan(v)):
                            perf.setdefault(sid, {})[tgt] = float(v)
                    if s.get('烘烤温度') is not None or s.get('烘烤时间') is not None:
                        proc.setdefault(sid, {'烘烤温度': s['烘烤温度'], '烘烤时间': s['烘烤时间']})
                report.append(f'[有标签] {base}: 样本{sum(1 for s in samples.values() if s.get("来源")==base)}')
            elif kind == 'formulation':
                sys_hint = '环氧-配比方案' if '配比' in base else ('聚酯金黄' if '聚酯' in base else '配方')
                for s in parse_formulation_file(path, sys_hint):
                    samples[s['样本ID']] = s
                report.append(f'[配方方案] {base}: 样本{sum(1 for s in samples.values() if s.get("来源")==base)}')
            elif kind == 'materials':
                ml = parse_materials_file(path)
                added = [c for c in ml if c not in mat_lib]
                for c in added:
                    mat_lib[c] = ml[c]
                report.append(f'[原料数据] {base}: 新增原料{len(added)}（已有{len(ml)-len(added)}种保留原描述符）')
            else:
                report.append(f'[未识别] {base}: 无法自动识别格式，跳过')
        except Exception as e:
            report.append(f'[失败] {base}: {e}')
    # 未登记原料自动估算登记（保证模板完整性）
    all_codes = set(c for s in samples.values() for c in s['组分'])
    new_codes = sorted(c for c in all_codes if c not in mat_lib)
    for c in new_codes:
        mat_lib[c] = est_material(c)
    if new_codes:
        report.append(f'自动估算登记 {len(new_codes)} 种新原料：{new_codes[:10]}{"..." if len(new_codes) > 10 else ""}')
    # 按样本ID去重（重复测量取均值，诚实评估协议）
    def _mean(vals):
        vals = [v for v in vals if v is not None and not (isinstance(v, float) and np.isnan(v))]
        return float(np.mean(vals)) if vals else None
    agg = {}
    for sid, s in samples.items():
        if sid not in agg:
            agg[sid] = dict(s)
            agg[sid]['_T弯'] = [s.get('T弯')]; agg[sid]['_MEK'] = [s.get('MEK')]; agg[sid]['_水煮'] = [s.get('水煮')]
        else:
            agg[sid]['_T弯'].append(s.get('T弯')); agg[sid]['_MEK'].append(s.get('MEK')); agg[sid]['_水煮'].append(s.get('水煮'))
    dedup = []
    for sid, s in agg.items():
        s = dict(s)
        s['样本ID'] = sid
        s['T弯'] = _mean(s.pop('_T弯')); s['MEK'] = _mean(s.pop('_MEK')); s['水煮'] = _mean(s.pop('_水煮'))
        dedup.append(s)
    n_dup = len(samples) - len(dedup)
    if n_dup > 0:
        report.append(f'按样本ID去重: {len(samples)} → {len(dedup)}（重复测量取均值）')
    samples = {s['样本ID']: s for s in dedup}
    return mat_lib, samples, perf, proc, report


# ---------- 导出：终极版模板结构 ----------
def export_template(mat_lib, samples, perf, proc, out_path):
    """导出终极版模板结构 Excel（原料主数据/配方明细/性能结果/工艺条件）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    HDR_FILL = PatternFill('solid', fgColor='1F2937')
    HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    ZEBRA_1 = PatternFill('solid', fgColor='FFFFFF')
    ZEBRA_2 = PatternFill('solid', fgColor='F7F9FC')
    THIN = Side(style='thin', color='D9DEE7')
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    BODY_FONT = Font(name='Arial', size=10)
    TITLE_FONT = Font(name='Arial', size=14, bold=True, color='1F2937')
    BOLD_FONT = Font(name='Arial', size=10, bold=True)

    def style_table(ws, header_row, n_cols, n_rows, kpi_cols=None):
        for c in range(1, n_cols + 1):
            cell = ws.cell(header_row, c)
            cell.fill = HDR_FILL; cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = BORDER
        for r in range(header_row + 1, header_row + n_rows + 1):
            fill = ZEBRA_1 if (r - header_row) % 2 == 0 else ZEBRA_2
            for c in range(1, n_cols + 1):
                cell = ws.cell(r, c)
                cell.fill = fill; cell.font = BODY_FONT; cell.border = BORDER
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        if kpi_cols:
            for c in kpi_cols:
                for r in range(header_row + 1, header_row + n_rows + 1):
                    ws.cell(r, c).fill = PatternFill('solid', fgColor='EAF2FF')
                    ws.cell(r, c).font = BOLD_FONT

    wb = Workbook()
    ws = wb.active; ws.title = '使用说明'
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 120
    lines = [
        ('整理后涂料配方-性能数据集（终极版模板 v3 结构）', ''),
        ('', ''),
        ('一、数据构成', '本文件由「数据整理与特征转换工作台」自动生成：'),
        ('', f'  · 原料主数据 {len(mat_lib)} 种'),
        ('', f'  · 配方样本 {len(samples)} 个'),
        ('', f'  · 性能记录 {sum(len(v) for v in perf.values())} 条'),
        ('', ''),
        ('二、工作表说明', '1. 原料主数据：原料描述符（建模特征基础）。'),
        ('', '2. 配方明细：长表，每行=一个样本中的一个组分（含系列）。'),
        ('', '3. 性能结果：实测性能（T弯/MEK擦拭/水煮等级）。'),
        ('', '4. 工艺条件：烘烤温度/时间等工艺参数。'),
        ('', '5. 配方级描述符：样本级特征矩阵（由特征一键转换生成）。'),
        ('', '6. 建模输入：宽表，一行=一个样本，特征+目标。'),
        ('', ''),
        ('三、使用建议', '1. 新增原料请在「原料主数据」补充真实描述符（SDS/TDS）。'),
        ('', '2. 用「特征一键转换」生成配方级描述符与建模输入。'),
    ]
    r = 1
    for text, _ in lines:
        cell = ws.cell(r, 2, text)
        if r == 1:
            cell.font = TITLE_FONT
        elif text.startswith(('一、', '二、', '三、')):
            cell.font = BOLD_FONT
        else:
            cell.font = BODY_FONT
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        r += 1

    # 原料主数据
    ws = wb.create_sheet('原料主数据')
    mat_headers = ['原料代码', '原料名称', '所属体系', '角色', '树脂类型', 'SMILES', '描述符状态'] + \
                  ['固含NV(%)', '密度(g/cm³)', '分子量(g/mol)', '环氧当量EEW(g/eq)', '酸值AV(mgKOH/g)',
                   '羟值OHV(mgKOH/g)', '胺值(mgKOH/g)', '官能度', 'Tg(℃)', '沸点(℃)', '闪点(℃)',
                   'Hansen δD', 'Hansen δP', 'Hansen δH', '极性指数', '相对挥发速率',
                   'C(%)', 'H(%)', 'O(%)', 'N(%)', 'S(%)', 'Cl(%)',
                   '环氧基(mol/100g)', '羟基(mol/100g)', '羧基(mol/100g)', '酯基(mol/100g)', '胺基(mol/100g)',
                   '酰胺(mol/100g)', '芳香环(mol/100g)', '醚键(mol/100g)', '蜡含量(%)', '颜料含量(%)', '数据来源', '备注']
    ws.append(mat_headers)
    for code, d in mat_lib.items():
        row = [code, code, '多体系', d['role'], d['rtype'], '', '已计算']
        for k in CONT_DESC:
            row.append(d.get(k, 0.0))
        row.append('类别典型值/文件信息'); row.append('')
        ws.append(row)
    style_table(ws, 1, len(mat_headers), len(mat_lib), kpi_cols=[8, 11, 12, 13, 14, 15, 30, 31, 32, 33, 34, 35, 36, 37])
    widths = [16, 20, 10, 8, 9, 14, 10] + [10] * 32 + [14, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    # 配方明细
    ws = wb.create_sheet('配方明细')
    det_headers = ['样本ID', '系列', '体系', '原料代码', '用量(g)', '角色', '树脂类型', '标签状态']
    ws.append(det_headers)
    det_rows = 0
    for s in samples.values():
        for code, amt in s['组分'].items():
            d = mat_lib.get(code)
            if d is None:
                continue
            ws.append([s['样本ID'], s['系列'], s['体系'], code, round(float(amt), 4), d['role'], d['rtype'], s['标签状态']])
            det_rows += 1
    style_table(ws, 1, len(det_headers), det_rows, kpi_cols=[4])
    for i, w in enumerate([14, 12, 16, 10, 8, 9, 10, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    # 性能结果
    ws = wb.create_sheet('性能结果')
    perf_headers = ['样本ID', '体系', '目标属性', '测试值', '单位', '标签状态', '标签来源', '测试条件']
    ws.append(perf_headers)
    perf_rows = 0
    for sid, p in perf.items():
        s = samples.get(sid, {})
        cond = f"{s.get('烘烤温度')}℃ {s.get('烘烤时间')}min" if s.get('烘烤温度') else ''
        for tgt, val in p.items():
            if val is not None and not (isinstance(val, float) and np.isnan(val)):
                unit = {'T弯': 'mm', 'MEK擦拭': '次', '水煮等级': '级'}.get(tgt, '')
                ws.append([sid, s.get('体系', ''), tgt, round(float(val), 4), unit, '实测', '实验室', cond])
                perf_rows += 1
    style_table(ws, 1, len(perf_headers), perf_rows, kpi_cols=[4])
    for i, w in enumerate([14, 12, 12, 10, 8, 10, 10, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 工艺条件
    ws = wb.create_sheet('工艺条件')
    proc_headers = ['样本ID', '体系', '烘烤温度(℃)', '烘烤时间(min)', '标签状态']
    ws.append(proc_headers)
    proc_rows = 0
    for sid, s in samples.items():
        p = proc.get(sid, {})
        ws.append([sid, s['体系'], p.get('烘烤温度'), p.get('烘烤时间'), s['标签状态']])
        proc_rows += 1
    style_table(ws, 1, len(proc_headers), proc_rows)
    for i, w in enumerate([14, 12, 12, 12, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    wb.save(out_path)
    return out_path


# ---------- 特征一键转换 ----------
def convert_features(mat_lib, samples, perf, proc):
    """计算特征矩阵（组分+增强描述符+比例+SMILES）→ (X, ids, series, feat_names)"""
    present_codes = sorted(set(canon(str(c).strip()) for s in samples.values() for c in s['组分']))
    rows, ids, series = [], [], []
    for sid, s in samples.items():
        p = proc.get(sid, {})
        row = build_sample_features(s['组分'], mat_lib, present_codes,
                                    bake_temp=p.get('烘烤温度'), bake_time=p.get('烘烤时间'))
        if row is None:
            continue
        rows.append(row); ids.append(sid); series.append(s.get('系列', ''))
    X = np.array(rows)
    feat_names = present_codes + ['烘烤温度', '烘烤时间'] + ENH_FEATURES + \
                 [f'r_{i}' for i in range(22)] + SMI_AGG_KEYS
    return X, ids, series, feat_names


def export_features(mat_lib, samples, perf, proc, out_path):
    """导出配方级描述符 + 建模输入 工作表到 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    X, ids, series, feat_names = convert_features(mat_lib, samples, perf, proc)
    HDR_FILL = PatternFill('solid', fgColor='1F2937')
    HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    ZEBRA_1 = PatternFill('solid', fgColor='FFFFFF')
    ZEBRA_2 = PatternFill('solid', fgColor='F7F9FC')
    THIN = Side(style='thin', color='D9DEE7')
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    BODY_FONT = Font(name='Arial', size=10)

    def style_table(ws, header_row, n_cols, n_rows):
        for c in range(1, n_cols + 1):
            cell = ws.cell(header_row, c)
            cell.fill = HDR_FILL; cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = BORDER
        for r in range(header_row + 1, header_row + n_rows + 1):
            fill = ZEBRA_1 if (r - header_row) % 2 == 0 else ZEBRA_2
            for c in range(1, n_cols + 1):
                cell = ws.cell(r, c)
                cell.fill = fill; cell.font = BODY_FONT; cell.border = BORDER
                cell.alignment = Alignment(vertical='center', wrap_text=True)

    wb = Workbook()
    ws = wb.active; ws.title = '配方级描述符'
    ws.append(['样本ID', '系列', '体系', '标签状态'] + feat_names)
    for i, sid in enumerate(ids):
        s = samples.get(sid, {})
        status = '实测' if s.get('标签状态') == '实测' else '无标签'
        ws.append([sid, series[i], s.get('体系', ''), status] +
                  [round(float(v), 6) for v in X[i]])
    style_table(ws, 1, len(feat_names) + 4, len(ids))
    for i, w in enumerate([14, 10, 12, 10] + [10] * len(feat_names), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'E2'

    ws = wb.create_sheet('建模输入')
    ws.append(['样本ID', '系列', '体系', '标签状态', 'T弯实测', 'MEK实测', '水煮实测'] + feat_names)
    for i, sid in enumerate(ids):
        s = samples.get(sid, {})
        status = '实测' if s.get('标签状态') == '实测' else '无标签'
        t_w = s.get('T弯'); m_w = s.get('MEK'); z_w = s.get('水煮')
        ws.append([sid, series[i], s.get('体系', ''), status,
                   round(float(t_w), 4) if t_w is not None else '',
                   round(float(m_w), 4) if m_w is not None else '',
                   round(float(z_w), 4) if z_w is not None else ''] +
                  [round(float(v), 6) for v in X[i]])
    style_table(ws, 1, len(feat_names) + 7, len(ids))
    for i, w in enumerate([14, 10, 12, 10, 10, 10, 10] + [10] * len(feat_names), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'H2'

    wb.save(out_path)
    return out_path, X.shape


# ---------- 辅助数据录入 ----------
def build_manual_entry():
    """返回手动录入所需元数据：原料代码列表（含角色/树脂类型）、体系选项"""
    codes = sorted(MAT.keys())
    return {
        'codes': codes,
        'roles': sorted(ROLES),
        'rtypes': sorted(RTYPES),
        'systems': ['环氧酚醛', '环氧-配比方案', '聚酯金黄', '聚酯', '有机', '其他'],
    }


def add_manual_sample(mat_lib, samples, perf, proc, entry):
    """录入一个配方样本（entry: 样本ID/体系/系列/烘烤温度/烘烤时间/组分{代码:用量}/T弯/MEK/水煮）"""
    sid = str(entry.get('样本ID', '')).strip()
    if not sid:
        raise ValueError('样本ID不能为空')
    comp = {}
    for code, amt in (entry.get('组分') or {}).items():
        amt = float(amt)
        if amt > 0:
            cc = clean_code(code)
            comp[cc] = comp.get(cc, 0.0) + amt
    if not comp:
        raise ValueError('组分不能为空（至少一个原料用量>0）')
    # 未登记原料自动估算
    for c in comp:
        if c not in mat_lib:
            mat_lib[c] = est_material(c)
    system = entry.get('体系', '其他') or '其他'
    series = entry.get('系列', system) or system
    samples[sid] = {
        '样本ID': sid, '体系': system, '系列': series, '组分': comp,
        '烘烤温度': entry.get('烘烤温度'), '烘烤时间': entry.get('烘烤时间'),
        'T弯': entry.get('T弯'), 'MEK': entry.get('MEK'), '水煮': entry.get('水煮'),
        '标签状态': '实测' if (entry.get('T弯') is not None or entry.get('MEK') is not None or entry.get('水煮') is not None) else '无标签',
        '来源': '手动录入',
    }
    for tgt, key in [('T弯', 'T弯'), ('MEK擦拭', 'MEK'), ('水煮等级', '水煮')]:
        v = entry.get(key)
        if v is not None:
            perf.setdefault(sid, {})[tgt] = float(v)
    if entry.get('烘烤温度') is not None or entry.get('烘烤时间') is not None:
        proc[sid] = {'烘烤温度': entry.get('烘烤温度'), '烘烤时间': entry.get('烘烤时间')}
    return sid


# ---------- HTTP 服务 ----------
TMP = os.path.join(os.path.dirname(os.path.abspath(__file__)), '_tmp')
os.makedirs(TMP, exist_ok=True)

# 会话状态（单用户本地应用）
STATE = {
    'mat_lib': None,
    'samples': None,
    'perf': None,
    'proc': None,
    'report': [],
}


def _reset_state():
    STATE['mat_lib'] = dict(MAT)
    STATE['samples'] = {}
    STATE['perf'] = {}
    STATE['proc'] = {}
    STATE['report'] = []


def _save_upload(data, name):
    """保存上传文件到临时目录，返回路径"""
    path = os.path.join(TMP, name)
    with open(path, 'wb') as fh:
        fh.write(data)
    return path


def _json_ok(obj):
    body = json.dumps(obj, ensure_ascii=False).encode('utf-8')
    return body, 'application/json; charset=utf-8'


def _json_err(msg):
    return _json_ok({'ok': False, 'error': str(msg)})


class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        pass

    def _send(self, body, ctype, status=200):
        self.send_response(status)
        self.send_header('Content-Type', ctype)
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_file(self, path, ctype):
        with open(path, 'rb') as fh:
            body = fh.read()
        self._send(body, ctype)

    def do_GET(self):
        parsed = urlparse(self.path)
        p = parsed.path
        if p in ('/', '/index.html'):
            self._send_file(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'index.html'), 'text/html; charset=utf-8')
        elif p == '/api/state':
            n_lab = sum(1 for s in (STATE['samples'] or {}).values() if s.get('标签状态') == '实测')
            body, ctype = _json_ok({
                'ok': True,
                'mat_count': len(STATE['mat_lib'] or {}),
                'sample_count': len(STATE['samples'] or {}),
                'labeled_count': n_lab,
                'perf_count': sum(len(v) for v in (STATE['perf'] or {}).values()),
                'report': STATE['report'],
            })
            self._send(body, ctype)
        elif p == '/api/materials':
            body, ctype = _json_ok({'ok': True, 'materials': STATE['mat_lib'] or {}})
            self._send(body, ctype)
        elif p == '/api/meta':
            body, ctype = _json_ok({'ok': True, **build_manual_entry()})
            self._send(body, ctype)
        elif p == '/api/export/manifest':
            # 导出流水线清单（可复现审计）
            try:
                manifest = build_manifest(STATE.get('_files', []), STATE.get('_ftypes', {}),
                                          {'mat_count': len(STATE['mat_lib'] or {}),
                                           'sample_count': len(STATE['samples'] or {}),
                                           'labeled_count': sum(1 for s in (STATE['samples'] or {}).values() if s.get('标签状态') == '实测'),
                                           'perf_count': sum(len(v) for v in (STATE['perf'] or {}).values())})
                body, ctype = _json_ok({'ok': True, 'manifest': manifest})
                self._send(body, ctype)
            except Exception as e:
                body, ctype = _json_err(e)
                self._send(body, ctype)
        elif p == '/api/acquire':
            # 补标签排程：推荐下一批应补测标签的样本（实验 M 结论：系列分层随机采样）
            try:
                if not STATE['samples']:
                    raise ValueError('当前无数据，请先整理或录入')
                q = parse_qs(parsed.query)
                budget = int(q.get('budget', ['10'])[0])
                strategy = q.get('strategy', ['strat_random'])[0]
                seed = int(q.get('seed', ['42'])[0])
                plan = build_acquisition_plan(STATE['samples'], budget=budget, strategy=strategy, seed=seed)
                body, ctype = _json_ok({'ok': True, **plan})
                self._send(body, ctype)
            except Exception as e:
                body, ctype = _json_err(e)
                self._send(body, ctype)
        elif p == '/api/readiness':
            # 建模就绪检查：自动评估数据是否达到可训练/逼近 R²>0.9 标准（实验 J/M/N 阈值）
            try:
                if not STATE['samples']:
                    raise ValueError('当前无数据，请先整理或录入')
                rep = build_readiness_report(STATE['mat_lib'], STATE['samples'], STATE['perf'], STATE['proc'])
                body, ctype = _json_ok(rep)
                self._send(body, ctype)
            except Exception as e:
                body, ctype = _json_err(e)
                self._send(body, ctype)
        elif p == '/api/export/template':
            try:
                if not STATE['samples']:
                    raise ValueError('当前无数据，请先整理或录入')
                out = os.path.join(TMP, '整理后数据集_终极版模板.xlsx')
                export_template(STATE['mat_lib'], STATE['samples'], STATE['perf'], STATE['proc'], out)
                self._send_file(out, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            except Exception as e:
                body, ctype = _json_err(e)
                self._send(body, ctype)
        elif p == '/api/export/features':
            try:
                if not STATE['samples']:
                    raise ValueError('当前无数据，请先整理或录入')
                out = os.path.join(TMP, '特征矩阵_配方级描述符与建模输入.xlsx')
                export_features(STATE['mat_lib'], STATE['samples'], STATE['perf'], STATE['proc'], out)
                self._send_file(out, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            except Exception as e:
                body, ctype = _json_err(e)
                self._send(body, ctype)
        else:
            self._send(b'not found', 'text/plain', 404)

    def do_POST(self):
        parsed = urlparse(self.path)
        p = parsed.path
        try:
            if p == '/api/reset':
                _reset_state()
                body, ctype = _json_ok({'ok': True, 'report': STATE['report']})
                self._send(body, ctype)
            elif p == '/api/upload':
                # 仅上传到临时目录（供预校验使用，不整理）
                ctype = self.headers.get('Content-Type', '')
                if 'multipart/form-data' not in ctype:
                    raise ValueError('需要 multipart/form-data 上传')
                boundary = ctype.split('boundary=')[1].strip().strip('"')
                raw = self.rfile.read(int(self.headers.get('Content-Length', 0)))
                files = _parse_multipart(raw, boundary)
                if not files:
                    raise ValueError('未收到文件')
                paths = []
                for name, data in files:
                    paths.append(_save_upload(data, name))
                body, ctype = _json_ok({'ok': True, 'paths': paths})
                self._send(body, ctype)
            elif p == '/api/validate':
                # 预校验：文件类型声明 + 校验（写死规则），返回错误/警告/建议
                n = int(self.headers.get('Content-Length', 0))
                payload = json.loads(self.rfile.read(n).decode('utf-8'))
                paths = payload.get('paths', [])
                declared = payload.get('declared', {})
                results = []
                for pth in paths:
                    if not os.path.exists(pth):
                        results.append({'name': os.path.basename(pth), 'ok': False,
                                        'errors': ['文件不存在'], 'warnings': [], 'info': {}})
                        continue
                    ftype = declared.get(os.path.basename(pth)) or suggest_type(pth)
                    r = validate_file(pth, ftype, STATE['mat_lib'])
                    r['name'] = os.path.basename(pth)
                    r['suggested'] = suggest_type(pth)
                    r['declared'] = ftype
                    r['type_desc'] = FILE_TYPES.get(ftype, '')
                    results.append(r)
                body, ctype = _json_ok({'ok': True, 'results': results})
                self._send(body, ctype)
            elif p == '/api/organize':
                # multipart 文件上传
                ctype = self.headers.get('Content-Type', '')
                if 'multipart/form-data' not in ctype:
                    raise ValueError('需要 multipart/form-data 上传')
                boundary = ctype.split('boundary=')[1].strip().strip('"')
                raw = self.rfile.read(int(self.headers.get('Content-Length', 0)))
                files = _parse_multipart(raw, boundary)
                if not files:
                    raise ValueError('未收到文件')
                paths = []
                for name, data in files:
                    paths.append(_save_upload(data, name))
                _reset_state()
                STATE['mat_lib'], STATE['samples'], STATE['perf'], STATE['proc'], STATE['report'] = organize_files(paths)
                STATE['_files'] = paths
                STATE['_ftypes'] = {os.path.basename(p): detect_format(p) for p in paths}
                n_lab = sum(1 for s in STATE['samples'].values() if s.get('标签状态') == '实测')
                body, ctype = _json_ok({
                    'ok': True,
                    'mat_count': len(STATE['mat_lib']),
                    'sample_count': len(STATE['samples']),
                    'labeled_count': n_lab,
                    'perf_count': sum(len(v) for v in STATE['perf'].values()),
                    'report': STATE['report'],
                })
                self._send(body, ctype)
            elif p == '/api/entry':
                n = int(self.headers.get('Content-Length', 0))
                entry = json.loads(self.rfile.read(n).decode('utf-8'))
                sid = add_manual_sample(STATE['mat_lib'], STATE['samples'], STATE['perf'], STATE['proc'], entry)
                body, ctype = _json_ok({'ok': True, '样本ID': sid, 'report': STATE['report']})
                self._send(body, ctype)
            elif p == '/api/load/template':
                ctype = self.headers.get('Content-Type', '')
                if 'multipart/form-data' not in ctype:
                    raise ValueError('需要 multipart/form-data 上传')
                boundary = ctype.split('boundary=')[1].strip().strip('"')
                raw = self.rfile.read(int(self.headers.get('Content-Length', 0)))
                files = _parse_multipart(raw, boundary)
                if not files:
                    raise ValueError('未收到文件')
                path = _save_upload(files[0][1], files[0][0])
                _reset_state()
                STATE['mat_lib'], STATE['samples'], STATE['perf'], STATE['proc'] = load_dataset(path)
                STATE['report'] = [f'已加载模板: {files[0][0]}（原料{len(STATE["mat_lib"])} 样本{len(STATE["samples"])}）']
                n_lab = sum(1 for s in STATE['samples'].values() if s.get('标签状态') == '实测')
                body, ctype = _json_ok({
                    'ok': True,
                    'mat_count': len(STATE['mat_lib']),
                    'sample_count': len(STATE['samples']),
                    'labeled_count': n_lab,
                    'perf_count': sum(len(v) for v in STATE['perf'].values()),
                    'report': STATE['report'],
                })
                self._send(body, ctype)
            else:
                self._send(b'not found', 'text/plain', 404)
        except Exception as e:
            traceback.print_exc()
            body, ctype = _json_err(e)
            self._send(body, ctype)


def _parse_multipart(raw, boundary):
    """解析 multipart/form-data，返回 [(filename, bytes)]"""
    delim = b'--' + boundary.encode('utf-8')
    parts = raw.split(delim)
    files = []
    for part in parts:
        if not part or part in (b'--\r\n', b'--', b'\r\n'):
            continue
        # 分离头与内容
        header_end = part.find(b'\r\n\r\n')
        if header_end < 0:
            continue
        header = part[:header_end].decode('utf-8', errors='ignore')
        content = part[header_end + 4:]
        if content.endswith(b'\r\n'):
            content = content[:-2]
        m = re.search(r'filename="([^"]*)"', header)
        if m:
            files.append((m.group(1), content))
    return files


def main():
    port = 8765
    _reset_state()
    url = f'http://127.0.0.1:{port}'
    print('=' * 60)
    print('数据整理与特征转换工作台 · Web 版 v1.0')
    print('=' * 60)
    print(f'请在浏览器打开: {url}')
    print('关闭本窗口即停止服务。')
    print('=' * 60)
    threading.Timer(1.0, lambda: webbrowser.open(url)).start()
    srv = ThreadingHTTPServer(('127.0.0.1', port), Handler)
    try:
        srv.serve_forever()
    except KeyboardInterrupt:
        pass


if __name__ == '__main__':
    main()
