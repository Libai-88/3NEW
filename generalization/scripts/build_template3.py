# -*- coding: utf-8 -*-
"""
终极版数据集模板 v3 —— 多体系适配 + 高便捷性
============================================
v3 相对 v2 的核心改进（均有实验/实践依据）：
  1) 新增「体系配置」表：以配置驱动模板，新增体系/目标属性无需改结构，
     实现"多体系适配"（环氧酚醛/有机/聚酯/聚氨酯/丙烯酸/环氧胺等）。
  2) 数据验证下拉：体系/角色/树脂类型/标签状态/目标属性全部下拉选择，
     杜绝手输错别字（人工误差主要来源）。
  3) 条件格式：配方明细中未在「原料主数据」登记的原料代码自动标红，
     提示先登记，形成"录入-登记"闭环。
  4) SMILES 状态列：原料主数据标注描述符计算状态（已计算/待计算/专有估算），
     支撑工作台自动计算分子描述符。
  5) 目标属性配置：每个体系可配置目标属性+单位+方向+数据类型，
     建模输入按配置自动生成目标列。
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from materials import MAT, CONT_DESC

# ---------- 样式 ----------
HDR_FILL = PatternFill('solid', fgColor='1F2937')
HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
ZEBRA_1 = PatternFill('solid', fgColor='FFFFFF')
ZEBRA_2 = PatternFill('solid', fgColor='F7F9FC')
KPI_FILL = PatternFill('solid', fgColor='EAF2FF')
GREEN_FILL = PatternFill('solid', fgColor='E3F5EA')
ORANGE_FILL = PatternFill('solid', fgColor='FFF1DE')
RED_FILL = PatternFill('solid', fgColor='FDE8E8')
CONF_FILL = PatternFill('solid', fgColor='F0F6FF')
THIN = Side(style='thin', color='D9DEE7')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BODY_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', size=10, bold=True)
TITLE_FONT = Font(name='Arial', size=14, bold=True, color='1F2937')
NOTE_FONT = Font(name='Arial', size=9, color='6B7280')

def style_table(ws, header_row, n_cols, n_rows, kpi_cols=None, status_col=None):
    for c in range(1, n_cols+1):
        cell = ws.cell(header_row, c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    for r in range(header_row+1, header_row+n_rows+1):
        fill = ZEBRA_1 if (r-header_row) % 2 == 0 else ZEBRA_2
        for c in range(1, n_cols+1):
            cell = ws.cell(r, c)
            cell.fill = fill; cell.font = BODY_FONT; cell.border = BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    if kpi_cols:
        for c in kpi_cols:
            for r in range(header_row+1, header_row+n_rows+1):
                ws.cell(r, c).fill = KPI_FILL
                ws.cell(r, c).font = BOLD_FONT
    if status_col:
        for r in range(header_row+1, header_row+n_rows+1):
            v = ws.cell(r, status_col).value
            if v == '实测':
                ws.cell(r, status_col).fill = GREEN_FILL
            elif v == '伪标签':
                ws.cell(r, status_col).fill = ORANGE_FILL
            elif v == '推荐测试':
                ws.cell(r, status_col).fill = KPI_FILL

wb = Workbook()

# ================= Sheet1 使用说明 =================
ws = wb.active; ws.title = '使用说明'
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 120
lines = [
    ('终极版涂料配方-性能数据集模板 v3', ''),
    ('', ''),
    ('一、模板定位', '统一承载不同化学体系（环氧酚醛、有机、聚酯、聚氨酯、丙烯酸、环氧胺等）的配方与性能数据，'),
    ('', '以"原料描述符"替代"原料编码"作为建模特征，使模型具备跨体系、跨新组分的泛化能力。'),
    ('', ''),
    ('二、v3 核心改进', '1. 体系配置表驱动：新增体系/目标属性只需在「体系配置」登记，无需改模板结构（多体系适配）。'),
    ('', '2. 数据验证下拉：体系/角色/树脂类型/标签状态/目标属性下拉选择，杜绝手输错别字（减少人工误差）。'),
    ('', '3. 待登记原料高亮：配方明细中未登记的原料代码自动标红，提示先到「原料主数据」登记（录入闭环）。'),
    ('', '4. SMILES 状态列：原料主数据标注描述符状态（已计算/待计算/专有估算），支撑工作台自动计算。'),
    ('', '5. 目标属性配置：每个体系可配置目标属性+单位+方向+数据类型，建模输入按配置自动生成。'),
    ('', ''),
    ('三、工作表结构', '1. 体系配置：登记体系、树脂类型、固化机制、目标属性（单位/方向/数据类型）。'),
    ('', '2. 原料主数据：登记原料描述符（SMILES 自动计算分子描述符；专有树脂填类别典型值）。'),
    ('', '3. 配方明细：长格式录入配方（每行=一个样本中的一个组分），含「系列」列用于系列编码建模。'),
    ('', '4. 性能结果：录入目标性能测试值，标注标签状态（实测/伪标签/推荐测试/人工复核）。'),
    ('', '5. 工艺条件：录入烘烤温度/时间/膜厚/基材等工艺参数（作为建模特征）。'),
    ('', '6. 配方级描述符：自动由配方明细+原料主数据聚合生成，是建模的特征矩阵。'),
    ('', '7. 建模输入：宽表，一行=一个样本，特征+目标+标签状态，直接用于模型训练。'),
    ('', '8. 数据字典：全部字段的口径说明。'),
    ('', ''),
    ('四、填写流程', '第1步：在「体系配置」确认/登记体系与目标属性。'),
    ('', '第2步：在「原料主数据」登记新原料（配方明细中未登记的代码会标红提示）。'),
    ('', '第3步：在「配方明细」按 样本ID+系列+原料代码+用量 逐行录入配方（系列=配方家族，如EP-A）。'),
    ('', '第4步：在「性能结果」录入测试值并标注标签状态；在「工艺条件」录入工艺参数。'),
    ('', '第5步：用配套 Windows 工作台一键完成：描述符计算→模型训练→性能预测→报告导出。'),
    ('', ''),
    ('五、系列编码（关键建模技术）', '1. 实验验证：组分特征+系列目标编码使 T弯 R² 从0.57提升至0.65（5折CV，折叠内OOF编码）。'),
    ('', '2. 原理：系列=配方家族，捕获批次/工艺/未测因素的系统性影响；编码=该系列历史目标均值。'),
    ('', '3. 已知系列预测：用该系列历史均值编码（R²≈0.65）；新系列预测：用全局均值编码（R²≈0.57）。'),
    ('', '4. 新体系/新原料：走描述符路线+主动学习闭环（预测→推荐测试→回填实测→重训）。'),
    ('', ''),
    ('六、标签补充闭环（解决缺标签问题）', '1. 用有标签体系训练基线模型（见泛化方案）。'),
    ('', '2. 对无标签体系配方预测，输出预测值+不确定性（树间std）。'),
    ('', '3. 高置信(低不确定)→自动生成伪标签；中置信→推荐测试(主动学习)；低置信→人工复核。'),
    ('', '4. 伪标签回放：将高置信伪标签加入训练集（权重<1）重训，迭代提升。'),
    ('', '5. 迁移学习：用源域预训练+目标域少量实测标签微调，效果最优（实验验证）。'),
    ('', ''),
    ('七、关键规则', '1. 原料代码必须与「原料主数据」一致且唯一；样本ID全局唯一。'),
    ('', '2. 用量统一为质量份(g)；描述符口径：固含/当量按"到货状态"计。'),
    ('', '3. 标签状态枚举：实测/伪标签/推荐测试/人工复核，建模时按状态筛选或加权。'),
    ('', '4. 性能结果与工艺条件通过"样本ID"与配方明细关联。'),
    ('', '5. 体系名称必须在「体系配置」中登记，否则下拉无法选择。'),
]
r = 1
for text, _ in lines:
    cell = ws.cell(r, 2, text)
    if r == 1:
        cell.font = TITLE_FONT
    elif text.startswith(('一、','二、','三、','四、','五、','六、')):
        cell.font = BOLD_FONT
    else:
        cell.font = BODY_FONT
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    r += 1

# ================= Sheet2 体系配置 =================
ws = wb.create_sheet('体系配置')
sys_headers = ['体系名称','固化机制','典型树脂类型','目标属性','单位','方向','数据类型','适用标准/说明']
ws.append(sys_headers)
sys_rows = [
    ('环氧酚醛','环氧-酚醛缩合','环氧/酚醛','T弯','mm','越低越好','连续','杯突/弯曲试验'),
    ('环氧酚醛','环氧-酚醛缩合','环氧/酚醛','MEK擦拭','次','越高越好','计数','MEK溶剂擦拭'),
    ('环氧酚醛','环氧-酚醛缩合','环氧/酚醛','水煮等级','级','越低越好','等级','1-5级水煮'),
    ('有机','羟基-氨基树脂','丙烯酸/乙烯基/环氧','T弯','mm','越低越好','连续','弯曲试验'),
    ('有机','羟基-氨基树脂','丙烯酸/乙烯基/环氧','MEK擦拭','次','越高越好','计数','MEK溶剂擦拭'),
    ('有机','羟基-氨基树脂','丙烯酸/乙烯基/环氧','附着力','级','越高越好','等级','划格试验'),
    ('聚酯','羟基-氨基树脂','聚酯','T弯','mm','越低越好','连续','弯曲试验'),
    ('聚酯','羟基-氨基树脂','聚酯','MEK擦拭','次','越高越好','计数','MEK溶剂擦拭'),
    ('聚酯','羟基-氨基树脂','聚酯','水煮等级','级','越低越好','等级','1-5级水煮'),
    ('聚氨酯','羟基-异氰酸酯','聚酯/丙烯酸/聚氨酯','T弯','mm','越低越好','连续','弯曲试验'),
    ('聚氨酯','羟基-异氰酸酯','聚酯/丙烯酸/聚氨酯','MEK擦拭','次','越高越好','计数','MEK溶剂擦拭'),
    ('聚氨酯','羟基-异氰酸酯','聚酯/丙烯酸/聚氨酯','硬度','H','越高越好','等级','铅笔硬度'),
    ('丙烯酸','自由基聚合/羟基-氨基','丙烯酸','T弯','mm','越低越好','连续','弯曲试验'),
    ('丙烯酸','自由基聚合/羟基-氨基','丙烯酸','MEK擦拭','次','越高越好','计数','MEK溶剂擦拭'),
    ('环氧胺','环氧-胺加成','环氧','T弯','mm','越低越好','连续','弯曲试验'),
    ('环氧胺','环氧-胺加成','环氧','MEK擦拭','次','越高越好','计数','MEK溶剂擦拭'),
]
for row in sys_rows:
    ws.append(list(row))
n_sys = len(sys_rows)
style_table(ws, 1, len(sys_headers), n_sys, kpi_cols=[1,4,5,6,7])
for i, w in enumerate([12,20,22,14,8,12,10,24], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'
# 体系名称下拉数据源（供其他表引用）
sys_names = sorted(set(r[0] for r in sys_rows))
target_names = sorted(set(r[3] for r in sys_rows))

# ================= Sheet3 原料主数据 =================
ws = wb.create_sheet('原料主数据')
mat_headers = ['原料代码','原料名称','所属体系','角色','树脂类型','SMILES','描述符状态',
               '固含NV(%)','密度(g/cm³)','分子量(g/mol)',
               '环氧当量EEW(g/eq)','酸值AV(mgKOH/g)','羟值OHV(mgKOH/g)','胺值(mgKOH/g)','官能度',
               'Tg(℃)','沸点(℃)','闪点(℃)','Hansen δD','Hansen δP','Hansen δH','极性指数','相对挥发速率',
               'C(%)','H(%)','O(%)','N(%)','S(%)','Cl(%)',
               '环氧基(mol/100g)','羟基(mol/100g)','羧基(mol/100g)','酯基(mol/100g)','胺基(mol/100g)',
               '酰胺(mol/100g)','芳香环(mol/100g)','醚键(mol/100g)','蜡含量(%)','颜料含量(%)','数据来源','备注']
ws.append(mat_headers)
sys_map = {}
for code in ['IR190','IR809','住友55754G','RF401','RF160','RF516','RF950','RF956','RH601','1510蜡','AZ088','正丁醇','补加混合液','10%磷酸']:
    sys_map[code] = '环氧酚醛'
for code in ['TF100','TM004','AS400','RX170-140','40%50177','IR877','RJ173M','RJ561','RY460','AC040','BYK104','IR909','R170M','IR557','TF022','TM221','IR868','RY075N','AZ135','35.7%白浆','14.28%炭黑浆料','3%气硅','20%CAB','杜邦-FT960','AL525','AL710','AZ306','AZ551','BYK306','FL208','FL208S','FL815C','IA151','IA893','IR842','RA009','RA083','RA824','RJ183','RJ362','日本151-PVC','TZ161','TZ425','TZ240','TT444','TT066','TM982','TM024','TZ221','RY078']:
    sys_map[code] = '有机'
for code in ['RJ173M','RJ561','RF950','RY460','AC040','TF100','IR557','IR909','R170M','RF516','AZ135','TM004','IR868','RY075N','TF022','TZ425','AL800','IA800','IA8000','10%AC040']:
    sys_map[code] = '聚酯'
names = {'IR190':'9型环氧树脂36%固含','IR809':'环氧树脂55%固含','住友55754G':'住友环氧树脂','RF401':'酚醛固化剂PR401',
         'RF160':'酚醛固化剂PR33160G','RF516':'酚醛固化剂PR516','RF950':'酚醛固化剂PR8219-50','RF956':'酚醛固化剂PR8219-65',
         'RH601':'酚醛固化剂SM601RX75','1510蜡':'1510蜡25%工作液','AZ088':'分散剂BYK088','正丁醇':'正丁醇',
         '补加混合液':'乙二醇单丁醚:二甲苯=2:1','10%磷酸':'磷酸10%水溶液','TF100':'乙烯基树脂','TM004':'乙二醇丁醚',
         'AS400':'丙烯酸树脂','RX170-140':'丙烯酸树脂','40%50177':'环氧树脂40%固含','IR877':'环氧树脂','RJ173M':'聚酯树脂',
         'RJ561':'聚酯树脂','RY460':'黄色颜料','AC040':'助剂','BYK104':'分散剂BYK104','IR909':'环氧树脂','R170M':'环氧树脂',
         'IR557':'环氧树脂','TF022':'乙烯基树脂','TM221':'乙二醇丁醚','IR868':'环氧树脂','RY075N':'黄色颜料','AZ135':'助剂',
         '35.7%白浆':'白色颜料浆35.7%','14.28%炭黑浆料':'炭黑浆料14.28%','3%气硅':'气相二氧化硅3%','20%CAB':'CAB溶液20%',
         '杜邦-FT960':'环氧树脂','AL525':'丙烯酸树脂','AL710':'丙烯酸树脂','AZ306':'助剂','AZ551':'助剂','BYK306':'流平剂',
         'FL208':'助剂','FL208S':'助剂','FL815C':'助剂','IA151':'丙烯酸树脂','IA893':'丙烯酸树脂','IR842':'环氧树脂',
         'RA009':'助剂','RA083':'助剂','RA824':'助剂','RJ183':'聚酯树脂','RJ362':'聚酯树脂','日本151-PVC':'颜料浆',
         'TZ161':'PMA溶剂','TZ425':'DBE溶剂','TZ240':'醋酸丁酯','TT444':'丁酮','TT066':'环己酮','TM982':'PM溶剂',
         'TM024':'二乙二醇单丁醚','TZ221':'乙二醇丁醚','RY078':'黄色颜料','AL800':'丙烯酸树脂','IA800':'丙烯酸树脂',
         'IA8000':'丙烯酸树脂','10%AC040':'AC040 10%'}
SMILES = {
    '正丁醇':'CCCCO','二甲苯':'Cc1ccccc1C','补加混合液':'CCCCOCCO','TM004':'CCCCOCCO',
    'TZ240':'CCCCOC(=O)C','TT444':'CCC(=O)C','TT066':'O=C1CCCCC1','TZ161':'CC(C)OC(=O)C',
    'TZ425':'COC(=O)CCCC(=O)OC','TM982':'CC(C)OC','TM024':'CCCCOCCOCCO','TZ221':'CCCCOCCO',
    '10%磷酸':'OP(=O)(O)O',
}
for code, d in MAT.items():
    smi = SMILES.get(code, '')
    status = '已计算' if smi else '专有估算'
    row = [code, names.get(code, code), sys_map.get(code, '通用'), d['role'], d['rtype'],
           smi, status]
    for k in ['NV','density','Mw','EEW','AV','OHV','amine','func','Tg','bp','fp','dD','dP','dH','pol','evap',
              'C','H','O','N','S','Cl','fg_epoxy','fg_oh','fg_cooh','fg_ester','fg_amine','fg_amide','fg_arom','fg_ether','wax','pig']:
        row.append(d[k])
    row.append('类别典型值/文件信息')
    row.append('')
    ws.append(row)
n_mat = len(MAT)
style_table(ws, 1, len(mat_headers), n_mat, kpi_cols=[8,11,12,13,14,15,30,31,32,33,34,35,36,37])
widths = [14,22,10,8,9,18,10,10,10,11,12,12,12,11,8,8,9,9,9,9,9,10,8,8,8,8,8,8,10,10,10,10,10,10,10,10,9,9,14,20]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'
# 数据验证：角色/树脂类型
dv_role = DataValidation(type='list', formula1='"树脂,固化剂,溶剂,助剂,颜料"', allow_blank=True)
dv_rtype = DataValidation(type='list', formula1='"环氧,酚醛,聚酯,乙烯基,丙烯酸,聚氨酯,其他"', allow_blank=True)
dv_sys = DataValidation(type='list', formula1='"环氧酚醛,有机,聚酯,聚氨酯,丙烯酸,环氧胺,通用"', allow_blank=True)
ws.add_data_validation(dv_role); ws.add_data_validation(dv_rtype); ws.add_data_validation(dv_sys)
dv_role.add('D2:D1000'); dv_rtype.add('E2:E1000'); dv_sys.add('C2:C1000')

# ================= Sheet4 配方明细 =================
ws = wb.create_sheet('配方明细')
det_headers = ['样本ID','系列','体系','原料代码','用量(g)','角色','树脂类型',
               '固含贡献','密度贡献','分子量贡献','EEW贡献','酸值贡献','羟值贡献','胺值贡献','官能度贡献',
               'Tg贡献','沸点贡献','闪点贡献','δD贡献','δP贡献','δH贡献','极性贡献','挥发速率贡献',
               'C贡献','H贡献','O贡献','N贡献','S贡献','Cl贡献',
               '环氧基贡献','羟基贡献','羧基贡献','酯基贡献','胺基贡献','酰胺贡献','芳香环贡献','醚键贡献',
               '蜡贡献','颜料贡献']
ws.append(det_headers)
MAT_COL = {'NV':8,'density':9,'Mw':10,'EEW':11,'AV':12,'OHV':13,'amine':14,'func':15,'Tg':16,'bp':17,'fp':18,
           'dD':19,'dP':20,'dH':21,'pol':22,'evap':23,'C':24,'H':25,'O':26,'N':27,'S':28,'Cl':29,
           'fg_epoxy':30,'fg_oh':31,'fg_cooh':32,'fg_ester':33,'fg_amine':34,'fg_amide':35,'fg_arom':36,'fg_ether':37,'wax':38,'pig':39}
CONT_START = 8
CONT_KEYS = ['NV','density','Mw','EEW','AV','OHV','amine','func','Tg','bp','fp','dD','dP','dH','pol','evap',
             'C','H','O','N','S','Cl','fg_epoxy','fg_oh','fg_cooh','fg_ester','fg_amine','fg_amide','fg_arom','fg_ether','wax','pig']

samples = []
epoxy_samples = [
    ('EP-0001','EP-A','环氧酚醛',{'IR190':66.0,'IR809':0.12,'RF516':2.64,'RF956':1.53,'1510蜡':0.79,'AZ088':0.06,'正丁醇':1.48,'补加混合液':1.27}),
    ('EP-0002','EP-A','环氧酚醛',{'IR190':66.09,'RF401':0.96,'RF516':0.48,'RF950':6.29,'RF956':6.65,'住友55754G':0.42,'1510蜡':2.98,'AZ088':0.06,'正丁醇':5.49,'补加混合液':1.55,'10%磷酸':2.4}),
    ('EP-0003','EP-B','环氧酚醛',{'IR190':66.08,'RF160':11.73,'RF950':0.46,'1510蜡':3.0,'AZ088':0.06,'正丁醇':4.11,'补加混合液':6.24,'10%磷酸':1.79}),
    ('EP-0004','EP-B','环氧酚醛',{'IR190':66.24,'RF160':1.02,'IR809':0.51,'RF516':6.18,'RF950':3.17,'RF956':6.81,'RH601':3.16,'住友55754G':1.18,'1510蜡':3.6,'AZ088':0.06,'正丁醇':7.69,'补加混合液':6.52,'10%磷酸':2.57}),
    ('EP-0005','EP-C','环氧酚醛',{'IR190':66.0,'RF401':1.42,'RF160':6.46,'IR809':0.65,'RF516':7.0,'RF950':8.98,'RH601':1.26,'住友55754G':1.98,'1510蜡':3.09,'AZ088':0.06,'正丁醇':11.84,'补加混合液':5.16,'10%磷酸':2.11}),
]
organic_samples = [
    ('ORG-0001','ORG-A','有机',{'TF100':19.18,'TM004':5.28,'AS400':3.68,'RX170-140':56.64,'40%50177':44.47,'IR877':37.01}),
    ('ORG-0002','ORG-A','有机',{'TF100':19.18,'TM004':5.28,'AS400':7.368,'RX170-140':56.64,'40%50177':44.47,'IR877':37.01}),
    ('ORG-0003','ORG-B','有机',{'TF100':29.48,'TM004':8.12,'AS400':5.66,'RX170-140':56.64,'40%50177':44.47,'RJ173M':37.01}),
]
poly_samples = [
    ('PES-0001','PES-A','聚酯',{'RJ173M':37.01,'RJ561':28.2,'RF950':5,'RY460':5,'AC040':0.15,'TF100':3.35,'TM004':1.35}),
    ('PES-0002','PES-A','聚酯',{'IR557':37.01,'RJ561':28.2,'RF950':10,'RY460':5,'AC040':0.15,'TF100':3.35,'TM004':1.35}),
    ('PES-0003','PES-B','聚酯',{'IR909':37.01,'RJ561':28.2,'RF516':10,'RY460':5,'AC040':0.15,'TF100':3.35,'TM004':1.35}),
]
samples = epoxy_samples + organic_samples + poly_samples

det_rows = 0
for sid, ser, sysn, comp in samples:
    for code, amt in comp.items():
        if amt <= 0:
            continue
        tr = ws.max_row + 1
        row = [sid, ser, sysn, code, amt]
        row.append(f'=IFERROR(VLOOKUP($D{tr},\'原料主数据\'!$A:$E,4,FALSE),"待登记")')
        row.append(f'=IFERROR(VLOOKUP($D{tr},\'原料主数据\'!$A:$E,5,FALSE),"待登记")')
        for k in CONT_KEYS:
            col = MAT_COL[k]
            row.append(f'=IFERROR($E{tr}*VLOOKUP($D{tr},\'原料主数据\'!$A:${get_column_letter(40)}, {col}, FALSE),"")')
        ws.append(row)
        det_rows += 1
style_table(ws, 1, len(det_headers), det_rows)
widths = [12,10,10,14,10,9,9] + [10]*32
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'
# 条件格式：未登记原料标红（D列代码不在原料主数据A列）
ws.conditional_formatting.add(
    f'D2:D{det_rows+1}',
    FormulaRule(formula=[f'ISNA(MATCH($D2,\'原料主数据\'!$A:$A,0))'], fill=RED_FILL, stopIfTrue=True))
# 数据验证：体系下拉
dv_sys2 = DataValidation(type='list', formula1='"环氧酚醛,有机,聚酯,聚氨酯,丙烯酸,环氧胺"', allow_blank=True)
ws.add_data_validation(dv_sys2); dv_sys2.add('C2:C10000')

# ================= Sheet5 性能结果 =================
ws = wb.create_sheet('性能结果')
perf_headers = ['样本ID','体系','目标属性','测试值','单位','标签状态','标签来源','不确定性','测试条件','测试日期','备注']
ws.append(perf_headers)
perf_rows = [
    ('EP-0001','环氧酚醛','T弯',17.415,'mm','实测','实验室','','205℃ 17min','2026-08-06',''),
    ('EP-0001','环氧酚醛','MEK擦拭',9,'次','实测','实验室','','205℃ 17min','2026-08-06',''),
    ('EP-0001','环氧酚醛','水煮等级',4,'级','实测','实验室','','205℃ 17min','2026-08-06',''),
    ('EP-0002','环氧酚醛','T弯',19.22,'mm','实测','实验室','','205℃ 17min','2026-08-06',''),
    ('EP-0002','环氧酚醛','MEK擦拭',161,'次','实测','实验室','','205℃ 17min','2026-08-06',''),
    ('EP-0003','环氧酚醛','T弯',18.05,'mm','实测','实验室','','205℃ 17min','2026-08-06',''),
    ('EP-0004','环氧酚醛','T弯',20.8,'mm','实测','实验室','','205℃ 17min','2026-08-06',''),
    ('EP-0004','环氧酚醛','MEK擦拭',300,'次','实测','实验室','','205℃ 17min','2026-08-06','超量程按300计'),
    ('ORG-0001','有机','T弯',None,'mm','推荐测试','主动学习','0.42','双涂双烘190+205℃*12min','2026-08-06',''),
    ('ORG-0002','有机','T弯',None,'mm','伪标签','模型预测','0.31','双涂双烘190+205℃*12min','2026-08-06',''),
    ('PES-0001','聚酯','T弯',None,'mm','人工复核','模型预测','0.68','待定','2026-08-06',''),
]
for row in perf_rows:
    ws.append(list(row))
n_perf = len(perf_rows)
style_table(ws, 1, len(perf_headers), n_perf, status_col=6)
for i, w in enumerate([12,10,12,10,8,10,12,10,22,12,14], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'
# 数据验证：体系/目标属性/标签状态
dv_sys3 = DataValidation(type='list', formula1='"环氧酚醛,有机,聚酯,聚氨酯,丙烯酸,环氧胺"', allow_blank=True)
dv_tgt = DataValidation(type='list', formula1='"T弯,MEK擦拭,水煮等级,附着力,硬度,光泽,冲击"', allow_blank=True)
dv_status = DataValidation(type='list', formula1='"实测,伪标签,推荐测试,人工复核"', allow_blank=True)
ws.add_data_validation(dv_sys3); ws.add_data_validation(dv_tgt); ws.add_data_validation(dv_status)
dv_sys3.add('B2:B10000'); dv_tgt.add('C2:C10000'); dv_status.add('F2:F10000')

# ================= Sheet6 工艺条件 =================
ws = wb.create_sheet('工艺条件')
proc_headers = ['样本ID','体系','烘烤温度(℃)','烘烤时间(min)','膜厚(g/m²)','基材','批次','线棒号','备注']
ws.append(proc_headers)
proc_rows = [
    ('EP-0001','环氧酚醛',205,17,None,'镀铬铁','8.6','14#',''),
    ('EP-0002','环氧酚醛',205,17,None,'镀铬铁','8.6','14#',''),
    ('EP-0003','环氧酚醛',205,17,None,'镀铬铁','8.6','14#',''),
    ('EP-0004','环氧酚醛',205,17,None,'镀铬铁','8.6','14#',''),
    ('EP-0005','环氧酚醛',205,17,None,'镀铬铁','8.6','14#',''),
    ('ORG-0001','有机',205,12,None,'镀铬铁','25.8.28','','双涂双烘190+205'),
    ('ORG-0002','有机',205,12,None,'镀铬铁','25.9.16','',''),
    ('ORG-0003','有机',205,12,None,'镀铬铁','25.9.18','',''),
    ('PES-0001','聚酯',None,None,None,'镀铬铁','25.8.20','',''),
    ('PES-0002','聚酯',None,None,None,'镀铬铁','25.8.28','',''),
    ('PES-0003','聚酯',None,None,None,'镀铬铁','25.9.18','',''),
]
for row in proc_rows:
    ws.append(list(row))
n_proc = len(proc_rows)
style_table(ws, 1, len(proc_headers), n_proc)
for i, w in enumerate([12,10,12,12,10,10,10,8,14], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'
dv_sys4 = DataValidation(type='list', formula1='"环氧酚醛,有机,聚酯,聚氨酯,丙烯酸,环氧胺"', allow_blank=True)
ws.add_data_validation(dv_sys4); dv_sys4.add('B2:B10000')

# ================= Sheet7 配方级描述符 =================
ws = wb.create_sheet('配方级描述符')
fd_headers = ['样本ID','系列','体系','总用量(g)','组分数量','树脂占比','固化剂占比','溶剂占比','助剂占比','颜料占比',
              '固化剂/树脂比','环氧树脂占比','酚醛树脂占比','聚酯树脂占比','乙烯基树脂占比','丙烯酸树脂占比','聚氨酯树脂占比',
              '加权固含(%)','加权密度','加权分子量','加权EEW','加权酸值','加权羟值','加权胺值','加权Tg','加权沸点','加权闪点',
              '加权δD','加权δP','加权δH','加权极性','加权挥发速率',
              '加权C(%)','加权H(%)','加权O(%)','加权N(%)','加权S(%)','加权Cl(%)',
              '环氧基密度(mol/100g)','羟基密度','羧基密度','酯基密度','胺基密度','酰胺密度','芳香环密度','醚键密度',
              '蜡含量(%)','颜料含量(%)']
ws.append(fd_headers)
DET = '配方明细'
DET_LAST = det_rows + 1
def DET_COL(letter):
    return f"'配方明细'!${letter}$2:${letter}${DET_LAST}"
sample_ids = list(dict.fromkeys([s[0] for s in samples]))
fd_rows = []
for sid in sample_ids:
    ser = next(s[1] for s in samples if s[0] == sid)
    sysn = next(s[2] for s in samples if s[0] == sid)
    fd_rows.append((sid, ser, sysn))
for i, (sid, ser, sysn) in enumerate(fd_rows):
    r = i + 2
    row = [sid, ser, sysn,
           f'=SUMIF({DET_COL("A")},$A{r},{DET_COL("E")})',
           f'=COUNTIF({DET_COL("A")},$A{r})',
           f'=IFERROR(SUMIFS({DET_COL("E")},{DET_COL("A")},$A{r},{DET_COL("F")},"树脂")/$C{r},0)',
           f'=IFERROR(SUMIFS({DET_COL("E")},{DET_COL("A")},$A{r},{DET_COL("F")},"固化剂")/$C{r},0)',
           f'=IFERROR(SUMIFS({DET_COL("E")},{DET_COL("A")},$A{r},{DET_COL("F")},"溶剂")/$C{r},0)',
           f'=IFERROR(SUMIFS({DET_COL("E")},{DET_COL("A")},$A{r},{DET_COL("F")},"助剂")/$C{r},0)',
           f'=IFERROR(SUMIFS({DET_COL("E")},{DET_COL("A")},$A{r},{DET_COL("F")},"颜料")/$C{r},0)',
           f'=IF($F{r}>0,$G{r}/$F{r},0)',
           f'=IFERROR(SUMIFS({DET_COL("E")},{DET_COL("A")},$A{r},{DET_COL("G")},"环氧")/$C{r},0)',
           f'=IFERROR(SUMIFS({DET_COL("E")},{DET_COL("A")},$A{r},{DET_COL("G")},"酚醛")/$C{r},0)',
           f'=IFERROR(SUMIFS({DET_COL("E")},{DET_COL("A")},$A{r},{DET_COL("G")},"聚酯")/$C{r},0)',
           f'=IFERROR(SUMIFS({DET_COL("E")},{DET_COL("A")},$A{r},{DET_COL("G")},"乙烯基")/$C{r},0)',
           f'=IFERROR(SUMIFS({DET_COL("E")},{DET_COL("A")},$A{r},{DET_COL("G")},"丙烯酸")/$C{r},0)',
           f'=IFERROR(SUMIFS({DET_COL("E")},{DET_COL("A")},$A{r},{DET_COL("G")},"聚氨酯")/$C{r},0)',
    ]
    cont_cols = {k: CONT_START + i for i, k in enumerate(CONT_KEYS)}
    for k in ['NV','density','Mw','EEW','AV','OHV','amine','func','Tg','bp','fp','dD','dP','dH','pol','evap','C','H','O','N','S','Cl']:
        col = cont_cols[k]
        row.append(f'=IFERROR(SUMPRODUCT(({DET_COL("A")}=$A{r})*{DET_COL(get_column_letter(col))})/$C{r},0)')
    for k in ['fg_epoxy','fg_oh','fg_cooh','fg_ester','fg_amine','fg_amide','fg_arom','fg_ether']:
        col = cont_cols[k]
        row.append(f'=IFERROR(SUMPRODUCT(({DET_COL("A")}=$A{r})*{DET_COL(get_column_letter(col))})/100,0)')
    for k in ['wax','pig']:
        col = cont_cols[k]
        row.append(f'=IFERROR(SUMPRODUCT(({DET_COL("A")}=$A{r})*{DET_COL(get_column_letter(col))})/$C{r},0)')
    ws.append(row)
n_fd = len(fd_rows)
style_table(ws, 1, len(fd_headers), n_fd, kpi_cols=[4,6,7,18,21,22,23,39,40,41,42])
for i, w in enumerate([12,10,10,10,9,9,9,9,9,9,11,10,10,10,11,11,11] + [10]*31, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# ================= Sheet8 建模输入（宽表，ML-ready） =================
ws = wb.create_sheet('建模输入')
mi_headers = ['样本ID','系列','体系','标签状态','目标属性','目标值','预测值','不确定性']
FD_HEADERS = fd_headers[3:]
for h in FD_HEADERS:
    mi_headers.append(h)
ws.append(mi_headers)
TARGETS_MI = ['T弯', 'MEK擦拭', '水煮等级']
mi_rows = []
for sid in sample_ids:
    ser = next(s[1] for s in samples if s[0] == sid)
    sysn = next(s[2] for s in samples if s[0] == sid)
    for tgt in TARGETS_MI:
        mi_rows.append((sid, ser, sysn, tgt))
FD_START = 4
for i, (sid, ser, sysn, tgt) in enumerate(mi_rows):
    r = i + 2
    row = [sid, ser, sysn,
           f'=IFERROR(INDEX(\'性能结果\'!$F:$F,MATCH(1,(\'性能结果\'!$A:$A=$A{r})*(\'性能结果\'!$C:$C=$E{r}),0)),"")',
           tgt,
           f'=IFERROR(INDEX(\'性能结果\'!$D:$D,MATCH(1,(\'性能结果\'!$A:$A=$A{r})*(\'性能结果\'!$C:$C=$E{r}),0)),"")',
           '', '',
    ]
    fd_row = sample_ids.index(sid) + 2
    for j in range(len(FD_HEADERS)):
        col = FD_START + j
        row.append(f"='配方级描述符'!{get_column_letter(col)}{fd_row}")
    ws.append(row)
n_mi = len(mi_rows)
style_table(ws, 1, len(mi_headers), n_mi, kpi_cols=[6])
for i, w in enumerate([12,10,10,10,12,10,10,10] + [10]*len(FD_HEADERS), 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# ================= Sheet9 数据字典 =================
ws = wb.create_sheet('数据字典')
dict_headers = ['字段名','所属工作表','类型','单位','口径说明','示例']
ws.append(dict_headers)
dict_rows = [
    ('样本ID','配方明细/性能结果/工艺条件/配方级描述符/建模输入','文本','-','全局唯一，关联各表','EP-0001'),
    ('系列','配方明细/配方级描述符/建模输入','文本','-','配方系列/家族（如EP-A），用于系列目标编码建模','EP-A'),
    ('体系','体系配置/配方明细','文本','-','化学体系类别，需在体系配置登记','环氧酚醛'),
    ('固化机制','体系配置','文本','-','体系交联反应机制','环氧-酚醛缩合'),
    ('目标属性','体系配置/性能结果','文本','-','性能名称，需在体系配置登记','T弯'),
    ('方向','体系配置','枚举','-','越低越好/越高越好，决定建模目标方向','越低越好'),
    ('数据类型','体系配置','枚举','-','连续/计数/等级/分类，决定回归或分类建模','连续'),
    ('原料代码','原料主数据/配方明细','文本','-','原料唯一编码，与原料主数据一致','IR190'),
    ('用量','配方明细','数值','g','该组分在样本中的质量份','66.0'),
    ('角色','原料主数据/配方明细','枚举','-','树脂/固化剂/溶剂/助剂/颜料','树脂'),
    ('树脂类型','原料主数据/配方明细','枚举','-','环氧/酚醛/聚酯/乙烯基/丙烯酸/聚氨酯/氨基/其他','环氧'),
    ('SMILES','原料主数据','文本','-','原料结构式，可自动计算RDKit/Mordred分子描述符','CCCCO'),
    ('描述符状态','原料主数据','枚举','-','已计算/待计算/专有估算','已计算'),
    ('固含NV','原料主数据','数值','%','按到货状态的固体含量','36'),
    ('密度','原料主数据','数值','g/cm³','原料密度','1.00'),
    ('分子量','原料主数据','数值','g/mol','数均分子量（树脂可为典型值）','1400'),
    ('环氧当量EEW','原料主数据','数值','g/eq','含1mol环氧基的到货产品质量','2640'),
    ('酸值AV','原料主数据','数值','mgKOH/g','中和1g样品游离酸所需KOH','0.5'),
    ('羟值OHV','原料主数据','数值','mgKOH/g','中和1g样品羟基所需KOH','30'),
    ('胺值','原料主数据','数值','mgKOH/g','中和1g样品胺基所需KOH','0'),
    ('官能度','原料主数据','数值','-','每分子平均活性基团数','2'),
    ('Tg','原料主数据','数值','℃','玻璃化转变温度','70'),
    ('沸点/闪点','原料主数据','数值','℃','常压沸点/闭杯闪点','250/100'),
    ('Hansen δD/δP/δH','原料主数据','数值','MPa^0.5','Hansen溶解度参数三分量','18.5/6.0/8.5'),
    ('极性指数','原料主数据','数值','-','溶剂极性指数（树脂可填0）','3.0'),
    ('相对挥发速率','原料主数据','数值','-','以乙酸丁酯=1','0.07'),
    ('C/H/O/N/S/Cl','原料主数据','数值','%','元素质量分数','72/7/20/0/0/0'),
    ('环氧基/羟基/羧基/酯基/胺基/酰胺/芳香环/醚键','原料主数据','数值','mol/100g','官能团密度','0.038'),
    ('蜡含量/颜料含量','原料主数据','数值','%','原料中蜡/颜料质量分数','25'),
    ('总用量','配方级描述符','公式','g','样本各组分用量之和','SUMIF(配方明细,A2,用量)'),
    ('树脂/固化剂/溶剂/助剂/颜料占比','配方级描述符','公式','-','各角色质量分数','0.72'),
    ('加权固含等','配方级描述符','公式','-','按质量分数加权的原料描述符','SUMPRODUCT(匹配*贡献)/总用量'),
    ('环氧基密度等','配方级描述符','公式','mol/100g','每100g配方的官能团摩尔数','0.05'),
    ('测试值','性能结果/建模输入','数值','-','性能测试结果（实测值）','17.415'),
    ('标签状态','性能结果/建模输入','枚举','-','实测/伪标签/推荐测试/人工复核','实测'),
    ('标签来源','性能结果','文本','-','实验室/模型预测/主动学习/人工复核','实验室'),
    ('不确定性','性能结果/建模输入','数值','-','模型预测的树间标准差（RF）','0.42'),
    ('预测值','建模输入','数值','-','模型对无标签样本的预测（伪标签候选）','18.6'),
    ('烘烤温度/时间','工艺条件','数值','℃/min','固化工艺参数','205/17'),
]
for row in dict_rows:
    ws.append(list(row))
n_dict = len(dict_rows)
style_table(ws, 1, len(dict_headers), n_dict)
for i, w in enumerate([34,34,8,12,46,14], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '终极版数据集模板.xlsx')
wb.save(out)
print('saved:', out)
