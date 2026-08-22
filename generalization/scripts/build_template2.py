# -*- coding: utf-8 -*-
"""
通用型数据集模板 v2 —— 既利于人录入，也利于模型训练
====================================================
v2 优化点（依据：MVP v2 验证 + 文献调研）：
  1) 原料主数据增加 SMILES 列：可接入 RDKit/Mordred 自动计算分子描述符，
     替代人工录入，减少人工误差（依据：Mordred/RDKit 自动描述符工具链）。
  2) 性能结果增加"标签状态/标签来源/不确定性"列：
     支撑半监督伪标签回放 + 主动学习推荐 + 人工复核闭环
     （依据：XRDMatch 伪标签、环氧胶主动学习、钢材缺陷预测+人工修缮案例）。
  3) 新增"建模输入"工作表：宽表格式，一行=一个样本，特征列+目标列+标签状态
     直接可导入 pandas/sklearn，无需任何加工即可训练（利于模型训练）。
  4) 配方级描述符保持自动聚合（SUMIFS/SUMPRODUCT），零人工计算。
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from materials import MAT, CONT_DESC

# ---------- 样式 ----------
HDR_FILL = PatternFill('solid', fgColor='1F2937')
HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
ZEBRA_1 = PatternFill('solid', fgColor='FFFFFF')
ZEBRA_2 = PatternFill('solid', fgColor='F7F9FC')
KPI_FILL = PatternFill('solid', fgColor='EAF2FF')
GREEN_FILL = PatternFill('solid', fgColor='E3F5EA')
ORANGE_FILL = PatternFill('solid', fgColor='FFF1DE')
THIN = Side(style='thin', color='D9DEE7')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BODY_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', size=10, bold=True)
TITLE_FONT = Font(name='Arial', size=14, bold=True, color='1F2937')
NOTE_FONT = Font(name='Arial', size=9, color='6B7280')

def style_table(ws, header_row, n_cols, n_rows, kpi_cols=None, status_col=None):
    for c in range(1, n_cols+1):
        cell = ws.cell(header_row, c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    for r in range(header_row+1, header_row+n_rows+1):
        fill = ZEBRA_1 if (r-header_row) % 2 == 0 else ZEBRA_2
        for c in range(1, n_cols+1):
            cell = ws.cell(r, c)
            cell.fill = fill; cell.font = BODY_FONT; cell.border = BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    if kpi_cols:
        for c in kpi_cols:
            for r in range(header_row+1, header_row+n_rows+1):
                ws.cell(r, c).fill = KPI_FILL
                ws.cell(r, c).font = BOLD_FONT
    if status_col:
        for r in range(header_row+1, header_row+n_rows+1):
            v = ws.cell(r, status_col).value
            if v == '实测':
                ws.cell(r, status_col).fill = GREEN_FILL
            elif v == '伪标签':
                ws.cell(r, status_col).fill = ORANGE_FILL
            elif v == '推荐测试':
                ws.cell(r, status_col).fill = KPI_FILL

wb = Workbook()

# ================= Sheet1 使用说明 =================
ws = wb.active; ws.title = '使用说明'
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 120
lines = [
    ('通用型涂料配方-性能数据集模板 v2', ''),
    ('', ''),
    ('一、模板定位', '统一承载不同化学体系（环氧酚醛、有机、聚酯、聚氨酯、丙烯酸等）的配方与性能数据，'),
    ('', '以"原料描述符"替代"原料编码"作为建模特征，使模型具备跨体系、跨新组分的泛化能力。'),
    ('', ''),
    ('二、v2 优化点', '1. 原料主数据新增 SMILES 列：可接入 RDKit/Mordred 自动计算分子描述符，替代人工录入，减少人工误差。'),
    ('', '2. 性能结果新增"标签状态/标签来源/不确定性"列：支撑半监督伪标签回放、主动学习推荐、人工复核闭环。'),
    ('', '3. 新增"建模输入"工作表：宽表格式，一行=一个样本，特征+目标+标签状态，可直接导入 pandas/sklearn 训练。'),
    ('', '4. 配方级描述符保持自动聚合（SUMIFS/SUMPRODUCT），零人工计算，降低误差变量。'),
    ('', ''),
    ('三、工作表结构', '1. 原料主数据：登记原料描述符（SMILES 可自动计算分子描述符；专有树脂填类别典型值）。'),
    ('', '2. 配方明细：长格式录入配方（每行=一个样本中的一个组分），配方级描述符自动计算。'),
    ('', '3. 性能结果：录入目标性能测试值，标注标签状态（实测/伪标签/推荐测试/人工复核）。'),
    ('', '4. 工艺条件：录入烘烤温度/时间/膜厚/基材等工艺参数（作为建模特征）。'),
    ('', '5. 配方级描述符：自动由配方明细+原料主数据聚合生成，是建模的特征矩阵。'),
    ('', '6. 建模输入：宽表，一行=一个样本，特征+目标+标签状态，直接用于模型训练。'),
    ('', '7. 数据字典：全部字段的口径说明。'),
    ('', ''),
    ('四、填写流程', '第1步：在"原料主数据"登记新原料。有 SMILES 的原料由脚本自动计算分子描述符；'),
    ('', '       专有树脂填化学类别典型值并注明"估算"。'),
    ('', '第2步：在"配方明细"按 样本ID+原料代码+用量 逐行录入配方。'),
    ('', '第3步：在"性能结果"录入测试值并标注标签状态；在"工艺条件"录入工艺参数。'),
    ('', '第4步：检查"配方级描述符"与"建模输入"自动生成完整，即可导出训练。'),
    ('', ''),
    ('五、标签补充闭环（解决缺标签问题）', '1. 冷启动：源域(环氧酚醛)预训练 + 目标域少量实测标签微调，建立初始模型（迁移学习，MVP实验H/I）。'),
    ('', '2. 对无标签体系配方预测，输出预测值+不确定性（深度集成方差）。'),
    ('', '3. 质量门控：源域可预测性高(随机划分CV R²≥0.2)才启用伪标签；高置信→伪标签(权重0.5)回放；否则跳过伪标签。'),
    ('', '4. 主动学习为主：对高不确定样本优先安排实测(推荐测试)，实测标签权重1.0回填；人工复核高置信伪标签与高不确定样本。'),
    ('', '5. 最终标签以实测为主、伪标签为辅；迭代至性能平台期或标注预算耗尽（双重停止准则）。'),
    ('', ''),
    ('六、关键规则', '1. 原料代码必须与"原料主数据"一致且唯一；样本ID全局唯一。'),
    ('', '2. 用量统一为质量份(g)；描述符口径：固含/当量按"到货状态"计。'),
    ('', '3. 标签状态枚举：实测/伪标签/推荐测试/人工复核，建模时按状态筛选或加权。'),
    ('', '4. 性能结果与工艺条件通过"样本ID"与配方明细关联。'),
]
r = 1
for text, _ in lines:
    cell = ws.cell(r, 2, text)
    if r == 1:
        cell.font = TITLE_FONT
    elif text.startswith(('一、','二、','三、','四、','五、','六、')):
        cell.font = BOLD_FONT
    else:
        cell.font = BODY_FONT
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    r += 1

# ================= Sheet2 原料主数据 =================
ws = wb.create_sheet('原料主数据')
mat_headers = ['原料代码','原料名称','所属体系','角色','树脂类型','SMILES',
               '固含NV(%)','密度(g/cm³)','分子量(g/mol)',
               '环氧当量EEW(g/eq)','酸值AV(mgKOH/g)','羟值OHV(mgKOH/g)','胺值(mgKOH/g)','官能度',
               'Tg(℃)','沸点(℃)','闪点(℃)','Hansen δD','Hansen δP','Hansen δH','极性指数','相对挥发速率',
               'C(%)','H(%)','O(%)','N(%)','S(%)','Cl(%)',
               '环氧基(mol/100g)','羟基(mol/100g)','羧基(mol/100g)','酯基(mol/100g)','胺基(mol/100g)',
               '酰胺(mol/100g)','芳香环(mol/100g)','醚键(mol/100g)','蜡含量(%)','颜料含量(%)','数据来源','备注']
ws.append(mat_headers)
sys_map = {}
for code in ['IR190','IR809','住友55754G','RF401','RF160','RF516','RF950','RF956','RH601','1510蜡','AZ088','正丁醇','补加混合液','10%磷酸']:
    sys_map[code] = '环氧酚醛'
for code in ['TF100','TM004','AS400','RX170-140','40%50177','IR877','RJ173M','RJ561','RY460','AC040','BYK104','IR909','R170M','IR557','TF022','TM221','IR868','RY075N','AZ135','35.7%白浆','14.28%炭黑浆料','3%气硅','20%CAB','杜邦-FT960','AL525','AL710','AZ306','AZ551','BYK306','FL208','FL208S','FL815C','IA151','IA893','IR842','RA009','RA083','RA824','RJ183','RJ362','日本151-PVC','TZ161','TZ425','TZ240','TT444','TT066','TM982','TM024','TZ221','RY078']:
    sys_map[code] = '有机'
for code in ['RJ173M','RJ561','RF950','RY460','AC040','TF100','IR557','IR909','R170M','RF516','AZ135','TM004','IR868','RY075N','TF022','TZ425','AL800','IA800','IA8000','10%AC040']:
    sys_map[code] = '聚酯'
names = {'IR190':'9型环氧树脂36%固含','IR809':'环氧树脂55%固含','住友55754G':'住友环氧树脂','RF401':'酚醛固化剂PR401',
         'RF160':'酚醛固化剂PR33160G','RF516':'酚醛固化剂PR516','RF950':'酚醛固化剂PR8219-50','RF956':'酚醛固化剂PR8219-65',
         'RH601':'酚醛固化剂SM601RX75','1510蜡':'1510蜡25%工作液','AZ088':'分散剂BYK088','正丁醇':'正丁醇',
         '补加混合液':'乙二醇单丁醚:二甲苯=2:1','10%磷酸':'磷酸10%水溶液','TF100':'乙烯基树脂','TM004':'乙二醇丁醚',
         'AS400':'丙烯酸树脂','RX170-140':'丙烯酸树脂','40%50177':'环氧树脂40%固含','IR877':'环氧树脂','RJ173M':'聚酯树脂',
         'RJ561':'聚酯树脂','RY460':'黄色颜料','AC040':'助剂','BYK104':'分散剂BYK104','IR909':'环氧树脂','R170M':'环氧树脂',
         'IR557':'环氧树脂','TF022':'乙烯基树脂','TM221':'乙二醇丁醚','IR868':'环氧树脂','RY075N':'黄色颜料','AZ135':'助剂',
         '35.7%白浆':'白色颜料浆35.7%','14.28%炭黑浆料':'炭黑浆料14.28%','3%气硅':'气相二氧化硅3%','20%CAB':'CAB溶液20%',
         '杜邦-FT960':'环氧树脂','AL525':'丙烯酸树脂','AL710':'丙烯酸树脂','AZ306':'助剂','AZ551':'助剂','BYK306':'流平剂',
         'FL208':'助剂','FL208S':'助剂','FL815C':'助剂','IA151':'丙烯酸树脂','IA893':'丙烯酸树脂','IR842':'环氧树脂',
         'RA009':'助剂','RA083':'助剂','RA824':'助剂','RJ183':'聚酯树脂','RJ362':'聚酯树脂','日本151-PVC':'颜料浆',
         'TZ161':'PMA溶剂','TZ425':'DBE溶剂','TZ240':'醋酸丁酯','TT444':'丁酮','TT066':'环己酮','TM982':'PM溶剂',
         'TM024':'二乙二醇单丁醚','TZ221':'乙二醇丁醚','RY078':'黄色颜料','AL800':'丙烯酸树脂','IA800':'丙烯酸树脂',
         'IA8000':'丙烯酸树脂','10%AC040':'AC040 10%'}
# SMILES 映射（可扩展；溶剂/单体有明确结构，树脂填"专有"）
SMILES = {
    '正丁醇':'CCCCO','二甲苯':'Cc1ccccc1C','补加混合液':'CCCCOCCO','TM004':'CCCCOCCO',
    'TZ240':'CCCCOC(=O)C','TT444':'CCC(=O)C','TT066':'O=C1CCCCC1','TZ161':'CC(C)OC(=O)C',
    'TZ425':'COC(=O)CCCC(=O)OC','TM982':'CC(C)OC','TM024':'CCCCOCCOCCO','TZ221':'CCCCOCCO',
    '10%磷酸':'OP(=O)(O)O',
}
for code, d in MAT.items():
    row = [code, names.get(code, code), sys_map.get(code, '通用'), d['role'], d['rtype'],
           SMILES.get(code, '专有(无SMILES)')]
    for k in ['NV','density','Mw','EEW','AV','OHV','amine','func','Tg','bp','fp','dD','dP','dH','pol','evap',
              'C','H','O','N','S','Cl','fg_epoxy','fg_oh','fg_cooh','fg_ester','fg_amine','fg_amide','fg_arom','fg_ether','wax','pig']:
        row.append(d[k])
    row.append('类别典型值/文件信息')
    row.append('')
    ws.append(row)
n_mat = len(MAT)
style_table(ws, 1, len(mat_headers), n_mat, kpi_cols=[7,10,11,12,13,14,29,30,31,32,33,34,35,36])
widths = [14,22,10,8,9,16,10,10,11,12,12,12,11,8,8,9,9,9,9,9,10,8,8,8,8,8,8,10,10,10,10,10,10,10,10,9,9,14,20]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# ================= Sheet3 配方明细 =================
ws = wb.create_sheet('配方明细')
det_headers = ['样本ID','体系','原料代码','用量(g)','角色','树脂类型',
               '固含贡献','密度贡献','分子量贡献','EEW贡献','酸值贡献','羟值贡献','胺值贡献','官能度贡献',
               'Tg贡献','沸点贡献','闪点贡献','δD贡献','δP贡献','δH贡献','极性贡献','挥发速率贡献',
               'C贡献','H贡献','O贡献','N贡献','S贡献','Cl贡献',
               '环氧基贡献','羟基贡献','羧基贡献','酯基贡献','胺基贡献','酰胺贡献','芳香环贡献','醚键贡献',
               '蜡贡献','颜料贡献']
ws.append(det_headers)
MAT_COL = {'NV':7,'density':8,'Mw':9,'EEW':10,'AV':11,'OHV':12,'amine':13,'func':14,'Tg':15,'bp':16,'fp':17,
           'dD':18,'dP':19,'dH':20,'pol':21,'evap':22,'C':23,'H':24,'O':25,'N':26,'S':27,'Cl':28,
           'fg_epoxy':29,'fg_oh':30,'fg_cooh':31,'fg_ester':32,'fg_amine':33,'fg_amide':34,'fg_arom':35,'fg_ether':36,'wax':37,'pig':38}
CONT_START = 7
CONT_KEYS = ['NV','density','Mw','EEW','AV','OHV','amine','func','Tg','bp','fp','dD','dP','dH','pol','evap',
             'C','H','O','N','S','Cl','fg_epoxy','fg_oh','fg_cooh','fg_ester','fg_amine','fg_amide','fg_arom','fg_ether','wax','pig']

samples = []
epoxy_samples = [
    ('EP-0001','环氧酚醛',{'IR190':66.0,'IR809':0.12,'RF516':2.64,'RF956':1.53,'1510蜡':0.79,'AZ088':0.06,'正丁醇':1.48,'补加混合液':1.27}),
    ('EP-0002','环氧酚醛',{'IR190':66.09,'RF401':0.96,'RF516':0.48,'RF950':6.29,'RF956':6.65,'住友55754G':0.42,'1510蜡':2.98,'AZ088':0.06,'正丁醇':5.49,'补加混合液':1.55,'10%磷酸':2.4}),
    ('EP-0003','环氧酚醛',{'IR190':66.08,'RF160':11.73,'RF950':0.46,'1510蜡':3.0,'AZ088':0.06,'正丁醇':4.11,'补加混合液':6.24,'10%磷酸':1.79}),
    ('EP-0004','环氧酚醛',{'IR190':66.24,'RF160':1.02,'IR809':0.51,'RF516':6.18,'RF950':3.17,'RF956':6.81,'RH601':3.16,'住友55754G':1.18,'1510蜡':3.6,'AZ088':0.06,'正丁醇':7.69,'补加混合液':6.52,'10%磷酸':2.57}),
    ('EP-0005','环氧酚醛',{'IR190':66.0,'RF401':1.42,'RF160':6.46,'IR809':0.65,'RF516':7.0,'RF950':8.98,'RH601':1.26,'住友55754G':1.98,'1510蜡':3.09,'AZ088':0.06,'正丁醇':11.84,'补加混合液':5.16,'10%磷酸':2.11}),
]
organic_samples = [
    ('ORG-0001','有机',{'TF100':19.18,'TM004':5.28,'AS400':3.68,'RX170-140':56.64,'40%50177':44.47,'IR877':37.01}),
    ('ORG-0002','有机',{'TF100':19.18,'TM004':5.28,'AS400':7.368,'RX170-140':56.64,'40%50177':44.47,'IR877':37.01}),
    ('ORG-0003','有机',{'TF100':29.48,'TM004':8.12,'AS400':5.66,'RX170-140':56.64,'40%50177':44.47,'RJ173M':37.01}),
]
poly_samples = [
    ('PES-0001','聚酯',{'RJ173M':37.01,'RJ561':28.2,'RF950':5,'RY460':5,'AC040':0.15,'TF100':3.35,'TM004':1.35}),
    ('PES-0002','聚酯',{'IR557':37.01,'RJ561':28.2,'RF950':10,'RY460':5,'AC040':0.15,'TF100':3.35,'TM004':1.35}),
    ('PES-0003','聚酯',{'IR909':37.01,'RJ561':28.2,'RF516':10,'RY460':5,'AC040':0.15,'TF100':3.35,'TM004':1.35}),
]
samples = epoxy_samples + organic_samples + poly_samples

det_rows = 0
for sid, sysn, comp in samples:
    for code, amt in comp.items():
        if amt <= 0:
            continue
        tr = ws.max_row + 1
        row = [sid, sysn, code, amt]
        row.append(f'=VLOOKUP($C{tr},\'原料主数据\'!$A:$E,4,FALSE)')
        row.append(f'=VLOOKUP($C{tr},\'原料主数据\'!$A:$E,5,FALSE)')
        for k in CONT_KEYS:
            col = MAT_COL[k]
            row.append(f'=$D{tr}*VLOOKUP($C{tr},\'原料主数据\'!$A:${get_column_letter(38)}, {col}, FALSE)')
        ws.append(row)
        det_rows += 1
style_table(ws, 1, len(det_headers), det_rows)
widths = [12,10,14,10,9,9] + [10]*32
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# ================= Sheet4 性能结果 =================
ws = wb.create_sheet('性能结果')
perf_headers = ['样本ID','体系','目标属性','测试值','单位','标签状态','标签来源','不确定性','测试条件','测试日期','备注']
ws.append(perf_headers)
perf_rows = [
    ('EP-0001','环氧酚醛','T弯',17.415,'mm','实测','实验室','','205℃ 17min','2026-08-06',''),
    ('EP-0001','环氧酚醛','MEK擦拭',9,'次','实测','实验室','','205℃ 17min','2026-08-06',''),
    ('EP-0001','环氧酚醛','水煮等级',4,'级','实测','实验室','','205℃ 17min','2026-08-06',''),
    ('EP-0002','环氧酚醛','T弯',19.22,'mm','实测','实验室','','205℃ 17min','2026-08-06',''),
    ('EP-0002','环氧酚醛','MEK擦拭',161,'次','实测','实验室','','205℃ 17min','2026-08-06',''),
    ('EP-0003','环氧酚醛','T弯',18.05,'mm','实测','实验室','','205℃ 17min','2026-08-06',''),
    ('EP-0004','环氧酚醛','T弯',20.8,'mm','实测','实验室','','205℃ 17min','2026-08-06',''),
    ('EP-0004','环氧酚醛','MEK擦拭',300,'次','实测','实验室','','205℃ 17min','2026-08-06','超量程按300计'),
    ('ORG-0001','有机','T弯',None,'mm','推荐测试','主动学习','0.42','双涂双烘190+205℃*12min','2026-08-06',''),
    ('ORG-0002','有机','T弯',None,'mm','伪标签','模型预测','0.31','双涂双烘190+205℃*12min','2026-08-06',''),
    ('PES-0001','聚酯','T弯',None,'mm','人工复核','模型预测','0.68','待定','2026-08-06',''),
]
for row in perf_rows:
    ws.append(list(row))
n_perf = len(perf_rows)
style_table(ws, 1, len(perf_headers), n_perf, status_col=6)
for i, w in enumerate([12,10,12,10,8,10,12,10,22,12,14], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# ================= Sheet5 工艺条件 =================
ws = wb.create_sheet('工艺条件')
proc_headers = ['样本ID','体系','烘烤温度(℃)','烘烤时间(min)','膜厚(g/m²)','基材','批次','线棒号','备注']
ws.append(proc_headers)
proc_rows = [
    ('EP-0001','环氧酚醛',205,17,None,'镀铬铁','8.6','14#',''),
    ('EP-0002','环氧酚醛',205,17,None,'镀铬铁','8.6','14#',''),
    ('EP-0003','环氧酚醛',205,17,None,'镀铬铁','8.6','14#',''),
    ('EP-0004','环氧酚醛',205,17,None,'镀铬铁','8.6','14#',''),
    ('EP-0005','环氧酚醛',205,17,None,'镀铬铁','8.6','14#',''),
    ('ORG-0001','有机',205,12,None,'镀铬铁','25.8.28','','双涂双烘190+205'),
    ('ORG-0002','有机',205,12,None,'镀铬铁','25.9.16','',''),
    ('ORG-0003','有机',205,12,None,'镀铬铁','25.9.18','',''),
    ('PES-0001','聚酯',None,None,None,'镀铬铁','25.8.20','',''),
    ('PES-0002','聚酯',None,None,None,'镀铬铁','25.8.28','',''),
    ('PES-0003','聚酯',None,None,None,'镀铬铁','25.9.18','',''),
]
for row in proc_rows:
    ws.append(list(row))
n_proc = len(proc_rows)
style_table(ws, 1, len(proc_headers), n_proc)
for i, w in enumerate([12,10,12,12,10,10,10,8,14], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# ================= Sheet6 配方级描述符 =================
ws = wb.create_sheet('配方级描述符')
fd_headers = ['样本ID','体系','总用量(g)','组分数量','树脂占比','固化剂占比','溶剂占比','助剂占比','颜料占比',
              '固化剂/树脂比','环氧树脂占比','酚醛树脂占比','聚酯树脂占比','乙烯基树脂占比','丙烯酸树脂占比','聚氨酯树脂占比',
              '加权固含(%)','加权密度','加权分子量','加权EEW','加权酸值','加权羟值','加权胺值','加权Tg','加权沸点','加权闪点',
              '加权δD','加权δP','加权δH','加权极性','加权挥发速率',
              '加权C(%)','加权H(%)','加权O(%)','加权N(%)','加权S(%)','加权Cl(%)',
              '环氧基密度(mol/100g)','羟基密度','羧基密度','酯基密度','胺基密度','酰胺密度','芳香环密度','醚键密度',
              '蜡含量(%)','颜料含量(%)']
ws.append(fd_headers)
DET = '配方明细'
DET_LAST = det_rows + 1
def DET_COL(letter):
    return f"'配方明细'!${letter}$2:${letter}${DET_LAST}"
sample_ids = list(dict.fromkeys([s[0] for s in samples]))
fd_rows = []
for sid in sample_ids:
    sysn = next(s[1] for s in samples if s[0] == sid)
    fd_rows.append((sid, sysn))
for i, (sid, sysn) in enumerate(fd_rows):
    r = i + 2
    row = [sid, sysn,
           f'=SUMIF({DET_COL("A")},$A{r},{DET_COL("D")})',
           f'=COUNTIF({DET_COL("A")},$A{r})',
           f'=SUMIFS({DET_COL("D")},{DET_COL("A")},$A{r},{DET_COL("E")},"树脂")/$C{r}',
           f'=SUMIFS({DET_COL("D")},{DET_COL("A")},$A{r},{DET_COL("E")},"固化剂")/$C{r}',
           f'=SUMIFS({DET_COL("D")},{DET_COL("A")},$A{r},{DET_COL("E")},"溶剂")/$C{r}',
           f'=SUMIFS({DET_COL("D")},{DET_COL("A")},$A{r},{DET_COL("E")},"助剂")/$C{r}',
           f'=SUMIFS({DET_COL("D")},{DET_COL("A")},$A{r},{DET_COL("E")},"颜料")/$C{r}',
           f'=IF($E{r}>0,$F{r}/$E{r},0)',
           f'=SUMIFS({DET_COL("D")},{DET_COL("A")},$A{r},{DET_COL("F")},"环氧")/$C{r}',
           f'=SUMIFS({DET_COL("D")},{DET_COL("A")},$A{r},{DET_COL("F")},"酚醛")/$C{r}',
           f'=SUMIFS({DET_COL("D")},{DET_COL("A")},$A{r},{DET_COL("F")},"聚酯")/$C{r}',
           f'=SUMIFS({DET_COL("D")},{DET_COL("A")},$A{r},{DET_COL("F")},"乙烯基")/$C{r}',
           f'=SUMIFS({DET_COL("D")},{DET_COL("A")},$A{r},{DET_COL("F")},"丙烯酸")/$C{r}',
           f'=SUMIFS({DET_COL("D")},{DET_COL("A")},$A{r},{DET_COL("F")},"聚氨酯")/$C{r}',
    ]
    cont_cols = {k: CONT_START + i for i, k in enumerate(CONT_KEYS)}
    for k in ['NV','density','Mw','EEW','AV','OHV','amine','func','Tg','bp','fp','dD','dP','dH','pol','evap','C','H','O','N','S','Cl']:
        col = cont_cols[k]
        row.append(f'=SUMPRODUCT(({DET_COL("A")}=$A{r})*{DET_COL(get_column_letter(col))})/$C{r}')
    for k in ['fg_epoxy','fg_oh','fg_cooh','fg_ester','fg_amine','fg_amide','fg_arom','fg_ether']:
        col = cont_cols[k]
        row.append(f'=SUMPRODUCT(({DET_COL("A")}=$A{r})*{DET_COL(get_column_letter(col))})/100')
    for k in ['wax','pig']:
        col = cont_cols[k]
        row.append(f'=SUMPRODUCT(({DET_COL("A")}=$A{r})*{DET_COL(get_column_letter(col))})/$C{r}')
    ws.append(row)
n_fd = len(fd_rows)
style_table(ws, 1, len(fd_headers), n_fd, kpi_cols=[3,5,6,17,20,21,22,38,39,40,41])
for i, w in enumerate([12,10,10,9,9,9,9,9,9,11,10,10,10,11,11,11] + [10]*31, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# ================= Sheet7 建模输入（宽表，ML-ready） =================
ws = wb.create_sheet('建模输入')
mi_headers = ['样本ID','体系','标签状态','目标属性','目标值','预测值','不确定性']
# 追加配方级描述符列（引用配方级描述符工作表）
FD_HEADERS = fd_headers[2:]  # 去掉 样本ID/体系
for h in FD_HEADERS:
    mi_headers.append(h)
ws.append(mi_headers)
# 建模输入行：每个样本 × 每个目标属性
TARGETS_MI = ['T弯', 'MEK擦拭', '水煮等级']
mi_rows = []
for sid in sample_ids:
    sysn = next(s[1] for s in samples if s[0] == sid)
    for tgt in TARGETS_MI:
        mi_rows.append((sid, sysn, tgt))
# 配方级描述符列在"配方级描述符"工作表中的列号 (C=3 起)
FD_START = 3  # 总用量
for i, (sid, sysn, tgt) in enumerate(mi_rows):
    r = i + 2
    row = [sid, sysn,
           f'=IFERROR(INDEX(\'性能结果\'!$F:$F,MATCH(1,(\'性能结果\'!$A:$A=$A{r})*(\'性能结果\'!$C:$C=$D{r}),0)),"")',
           tgt,
           f'=IFERROR(INDEX(\'性能结果\'!$D:$D,MATCH(1,(\'性能结果\'!$A:$A=$A{r})*(\'性能结果\'!$C:$C=$D{r}),0)),"")',
           '',  # 预测值（由脚本回填）
           '',  # 不确定性（由脚本回填）
    ]
    # 引用配方级描述符的对应行（同一样本）
    fd_row = sample_ids.index(sid) + 2
    for j in range(len(FD_HEADERS)):
        col = FD_START + j
        row.append(f"='配方级描述符'!{get_column_letter(col)}{fd_row}")
    ws.append(row)
n_mi = len(mi_rows)
style_table(ws, 1, len(mi_headers), n_mi, kpi_cols=[5])
for i, w in enumerate([12,10,10,12,10,10,10] + [10]*len(FD_HEADERS), 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# ================= Sheet8 数据字典 =================
ws = wb.create_sheet('数据字典')
dict_headers = ['字段名','所属工作表','类型','单位','口径说明','示例']
ws.append(dict_headers)
dict_rows = [
    ('样本ID','配方明细/性能结果/工艺条件/配方级描述符/建模输入','文本','-','全局唯一，关联各表','EP-0001'),
    ('体系','配方明细','文本','-','化学体系类别：环氧酚醛/有机/聚酯/聚氨酯/丙烯酸等','环氧酚醛'),
    ('原料代码','原料主数据/配方明细','文本','-','原料唯一编码，与原料主数据一致','IR190'),
    ('用量','配方明细','数值','g','该组分在样本中的质量份','66.0'),
    ('角色','原料主数据/配方明细','枚举','-','树脂/固化剂/溶剂/助剂/颜料','树脂'),
    ('树脂类型','原料主数据/配方明细','枚举','-','环氧/酚醛/聚酯/乙烯基/丙烯酸/聚氨酯/其他','环氧'),
    ('SMILES','原料主数据','文本','-','原料结构式，可自动计算RDKit/Mordred分子描述符','CCCCO'),
    ('固含NV','原料主数据','数值','%','按到货状态的固体含量','36'),
    ('密度','原料主数据','数值','g/cm³','原料密度','1.00'),
    ('分子量','原料主数据','数值','g/mol','数均分子量（树脂可为典型值）','1400'),
    ('环氧当量EEW','原料主数据','数值','g/eq','含1mol环氧基的到货产品质量','2640'),
    ('酸值AV','原料主数据','数值','mgKOH/g','中和1g样品游离酸所需KOH','0.5'),
    ('羟值OHV','原料主数据','数值','mgKOH/g','中和1g样品羟基所需KOH','30'),
    ('胺值','原料主数据','数值','mgKOH/g','中和1g样品胺基所需KOH','0'),
    ('官能度','原料主数据','数值','-','每分子平均活性基团数','2'),
    ('Tg','原料主数据','数值','℃','玻璃化转变温度','70'),
    ('沸点/闪点','原料主数据','数值','℃','常压沸点/闭杯闪点','250/100'),
    ('Hansen δD/δP/δH','原料主数据','数值','MPa^0.5','Hansen溶解度参数三分量','18.5/6.0/8.5'),
    ('极性指数','原料主数据','数值','-','溶剂极性指数（树脂可填0）','3.0'),
    ('相对挥发速率','原料主数据','数值','-','以乙酸丁酯=1','0.07'),
    ('C/H/O/N/S/Cl','原料主数据','数值','%','元素质量分数','72/7/20/0/0/0'),
    ('环氧基/羟基/羧基/酯基/胺基/酰胺/芳香环/醚键','原料主数据','数值','mol/100g','官能团密度','0.038'),
    ('蜡含量/颜料含量','原料主数据','数值','%','原料中蜡/颜料质量分数','25'),
    ('总用量','配方级描述符','公式','g','样本各组分用量之和','SUMIF(配方明细,A2,用量)'),
    ('树脂/固化剂/溶剂/助剂/颜料占比','配方级描述符','公式','-','各角色质量分数','0.72'),
    ('加权固含等','配方级描述符','公式','-','按质量分数加权的原料描述符','SUMPRODUCT(匹配*贡献)/总用量'),
    ('环氧基密度等','配方级描述符','公式','mol/100g','每100g配方的官能团摩尔数','0.05'),
    ('目标属性','性能结果/建模输入','文本','-','性能名称：T弯/MEK擦拭/水煮等级等','T弯'),
    ('测试值','性能结果/建模输入','数值','-','性能测试结果（实测值）','17.415'),
    ('标签状态','性能结果/建模输入','枚举','-','实测/伪标签/推荐测试/人工复核','实测'),
    ('标签来源','性能结果','文本','-','实验室/模型预测/主动学习/人工复核','实验室'),
    ('不确定性','性能结果/建模输入','数值','-','模型预测的树间标准差（RF）','0.42'),
    ('预测值','建模输入','数值','-','模型对无标签样本的预测（伪标签候选）','18.6'),
    ('烘烤温度/时间','工艺条件','数值','℃/min','固化工艺参数','205/17'),
]
for row in dict_rows:
    ws.append(list(row))
n_dict = len(dict_rows)
style_table(ws, 1, len(dict_headers), n_dict)
for i, w in enumerate([34,34,8,12,46,14], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '通用型数据集模板.xlsx')
wb.save(out)
print('saved:', out)
