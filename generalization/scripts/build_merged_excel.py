# -*- coding: utf-8 -*-
"""
合并版数据集 Excel 生成：终极版模板结构 + 全部数据
==================================================
- 原料主数据：80 种原料（同物异名已合并；占位行按公开手册值/名称自证修正，余量标「待确认」）
- 配方明细：486 样本（371 有标签 + 115 无标签）
- 性能结果：有标签样本的 T弯/MEK/水煮
- 工艺条件：烘烤参数
- 配方级描述符：数值填充（不依赖公式）
- 配方级机理特征：workbench/mech_desc.py 计算的当量/计量比/νe/Fox Tg/固化度等
- 建模输入：ML-ready 宽表
"""
import pickle, warnings, os
import numpy as np
import pandas as pd
warnings.filterwarnings('ignore')
import sys; sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'workbench'))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from materials import CONT_DESC, ALIAS
import handbook_fixes as HF
import tds_sds

D = pickle.load(open(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'data', 'merged_data.pkl'),'rb'))
full_mat = D['full_mat']; new_mats = D['new_mats']
all_samples = D['all_samples']

# 占位描述符修正：同物合并 + 名称自证固含 + 公开手册常数 + 族内一致
_fix_changed, _MERGE, _PENDING = HF.apply(full_mat)
for _c in _MERGE:
    full_mat.pop(_c, None)
new_mats = [c for c in new_mats if c not in _MERGE]

# 供应商 TDS/SDS 实测层（最高优先级，逐字段替换类别典型值并登记来源）
_tds_changed, _TDS_PROV = tds_sds.apply(full_mat)


def clean_code(code):
    """原始代码 → 清洗代码（应用ALIAS映射与同物合并表）"""
    key = str(code).strip()
    key = ALIAS.get(key, key)
    return _MERGE.get(key, key)


def _num(v):
    v = _amount(v)
    return 0.0 if v is None else v


def _desc_row(comp, mat):
    """配方级描述符（线性口径），与 scripts/descriptors.py 同式，但用当前原料库重算。"""
    items, total = [], 0.0
    for code, amt in comp.items():
        a = _amount(amt)
        if a is None or a <= 0:
            continue
        key = clean_code(code)
        if key not in mat:
            continue
        items.append((key, a))
        total += a
    if total <= 0 or not items:
        return None
    w = [a / total for _, a in items]
    roles = ['树脂', '固化剂', '溶剂', '助剂', '颜料']
    rt = ['环氧', '酚醛', '聚酯', '乙烯基', '丙烯酸', '聚氨酯', '氨基', '其他']
    role_frac = {r: 0.0 for r in roles}
    rtype_frac = {r: 0.0 for r in rt}
    for (k, _), wi in zip(items, w):
        role_frac[mat[k]['role']] += wi
        rtype_frac[mat[k]['rtype']] += wi
    d = {}
    for dk in CONT_DESC:
        d['w_' + dk] = sum(_num(mat[k][dk]) * wi for (k, _), wi in zip(items, w))
    for fg in ['fg_epoxy', 'fg_oh', 'fg_cooh', 'fg_ester', 'fg_amine', 'fg_amide', 'fg_arom', 'fg_ether']:
        d['s_' + fg] = sum(_num(mat[k][fg]) * a for k, a in items)
    resin = sum(a for (k, a) in items if mat[k]['role'] == '树脂')
    xlink = sum(a for (k, a) in items if mat[k]['role'] == '固化剂')
    ep = d['s_fg_epoxy']
    oh = d['s_fg_oh']
    d.update(resin_frac=role_frac['树脂'], xlink_frac=role_frac['固化剂'], solvent_frac=role_frac['溶剂'],
             additive_frac=role_frac['助剂'], pigment_frac=role_frac['颜料'],
             xlink_resin_ratio=xlink / resin if resin > 0 else 0,
             oh_epoxy_eq_ratio=oh / ep if ep > 0 else 0,
             epoxy_eq_100g=ep, oh_eq_100g=oh, n_components=len(items), avg_func=d['w_func'])
    for r in rt:
        d['rtype_' + r] = rtype_frac[r]
    return d


def _amount(v):
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return None if np.isnan(f) else f


DESC_ORDER = (['resin_frac', 'xlink_frac', 'solvent_frac', 'additive_frac', 'pigment_frac',
               'xlink_resin_ratio', 'oh_epoxy_eq_ratio', 'epoxy_eq_100g', 'oh_eq_100g',
               'n_components', 'avg_func',
               'rtype_环氧', 'rtype_酚醛', 'rtype_聚酯', 'rtype_乙烯基', 'rtype_丙烯酸',
               'rtype_聚氨酯', 'rtype_氨基', 'rtype_其他']
              + ['w_' + k for k in CONT_DESC]
              + ['s_' + k for k in ['fg_epoxy', 'fg_oh', 'fg_cooh', 'fg_ester', 'fg_amine',
                                    'fg_amide', 'fg_arom', 'fg_ether']])

# 配方级描述符：按当前原料库（含 TDS/SDS 实测层）重算，保证与原料主数据一致
_rows, _ids = [], []
for s in all_samples:
    d = _desc_row(s['组分'], full_mat)
    if d is None:
        continue
    _rows.append([d.get(k, 0.0) for k in DESC_ORDER])
    _ids.append(s['样本ID'])
desc_df = pd.DataFrame(_rows, columns=DESC_ORDER)
desc_df.insert(0, '样本ID', _ids)
desc_df.insert(1, '体系', [next(x['体系'] for x in all_samples if x['样本ID'] == i) for i in _ids])


# ---------- 样式 ----------
HDR_FILL = PatternFill('solid', fgColor='1F2937')
HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
ZEBRA_1 = PatternFill('solid', fgColor='FFFFFF')
ZEBRA_2 = PatternFill('solid', fgColor='F7F9FC')
KPI_FILL = PatternFill('solid', fgColor='EAF2FF')
GREEN_FILL = PatternFill('solid', fgColor='E3F5EA')
ORANGE_FILL = PatternFill('solid', fgColor='FFF1DE')
THIN = Side(style='thin', color='D9DEE7')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BODY_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', size=10, bold=True)
TITLE_FONT = Font(name='Arial', size=14, bold=True, color='1F2937')
NOTE_FONT = Font(name='Arial', size=9, color='6B7280')

def style_table(ws, header_row, n_cols, n_rows, kpi_cols=None):
    for c in range(1, n_cols+1):
        cell = ws.cell(header_row, c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    for r in range(header_row+1, header_row+n_rows+1):
        fill = ZEBRA_1 if (r-header_row) % 2 == 0 else ZEBRA_2
        for c in range(1, n_cols+1):
            cell = ws.cell(r, c)
            cell.fill = fill; cell.font = BODY_FONT; cell.border = BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    if kpi_cols:
        for c in kpi_cols:
            for r in range(header_row+1, header_row+n_rows+1):
                ws.cell(r, c).fill = KPI_FILL
                ws.cell(r, c).font = BOLD_FONT

wb = Workbook()

# ================= Sheet1 使用说明 =================
ws = wb.active; ws.title = '使用说明'
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 120
lines = [
    ('合并版涂料配方-性能数据集（终极版模板 v3 填充）', ''),
    ('', ''),
    ('一、数据构成', '本文件将现有全部数据源整理并填入终极版模板 v3：'),
    ('', '  · 有标签样本 371 条（环氧酚醛 345 + 聚酯金黄 26，含 T弯/MEK/水煮 实测值）'),
    ('', '  · 无标签配方 115 条（环氧配比方案 112 + 聚酯金黄 3）'),
    ('', '  · 原料主数据 80 种（同物异名已合并；49 种描述符已按供应商 TDS/SDS 逐字段实测替换）'),
    ('', ''),
    ('二、数据来源', '1. 配料测试数据汇总V1.xlsx（有标签）'),
    ('', '2. AI研发26.7.22配比方案.xlsx（无标签-环氧）'),
    ('', '3. 聚酯金黄-AI(1).xlsx（无标签-聚酯）'),
    ('', '4. AI项目原料送检、部分实验数据.xlsx（原料信息）'),
    ('', '5. TDS-SDS/ 供应商技术数据表与安全说明书 291 份（原料描述符实测值与档案出处）'),
    ('', ''),
    ('三、工作表说明', '1. 原料主数据：80 种原料描述符；「描述符状态」区分 TDS实测/送检组成/手册值/类别典型值/待确认，「数据来源」标注档案实测字段数，「备注」给出牌号对应依据与档案文件名。'),
    ('', '2. 配方明细：486 样本的配方长表（每行=一个组分）。'),
    ('', '3. 性能结果：实测样本（含聚酯金黄 26 条补全）的性能数据。'),
    ('', '4. 工艺条件：烘烤温度/时间等。'),
    ('', '5. 配方级描述符：按质量分数加权的原料描述符（线性口径）。'),
    ('', '6. 配方级机理特征：当量浓度、化学计量比、交联密度、Fox Tg、固化度、'
         'Hansen 距离、PVC 等非线性机理量（由 workbench/mech_desc.py 计算）。'),
    ('', '7. 建模输入：宽表，一行=一个样本，特征+目标，可直接导入训练。'),
    ('', '8. 数据字典：各字段的类型、单位与口径说明。'),
    ('', ''),
    ('四、使用建议', '1. 建模时按"标签状态"筛选：实测样本用于训练/验证，无标签样本用于预测。'),
    ('', '2. 水煮指标：聚酯金黄 26 条实测全部为 2 级，该体系对水煮不提供判别信息，'
         '评估须按体系拆分报告，避免常量层抬高综合准确率。'),
    ('', '3. 聚酯金黄 29 条配方无烘烤工艺记录，「配方级机理特征」表中固化类特征'
         '（烘烤依赖）对该部分样本留空，表示不可知而非零固化。'),
    ('', '4. 标为「类别典型值」的原料按 TDS-SDS/ 补档清单逐条替换（牌号未识别的树脂集中在聚酯金黄与配比方案）；「待确认」行（DMP、209-基料）需 SDS 核定。'),
    ('', '5. 用配套 Windows 工作台可一键完成建模与预测。'),
]
r = 1
for text, _ in lines:
    cell = ws.cell(r, 2, text)
    if r == 1:
        cell.font = TITLE_FONT
    elif text.startswith(('一、','二、','三、','四、')):
        cell.font = BOLD_FONT
    else:
        cell.font = BODY_FONT
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    r += 1

# ================= Sheet2 体系配置 =================
ws = wb.create_sheet('体系配置')
sys_headers = ['体系名称','固化机制','典型树脂类型','目标属性','单位','方向','数据类型','适用标准/说明']
ws.append(sys_headers)
sys_rows = [
    ('环氧酚醛','环氧-酚醛缩合','环氧/酚醛','T弯','mm','越低越好','连续','杯突/弯曲试验'),
    ('环氧酚醛','环氧-酚醛缩合','环氧/酚醛','MEK擦拭','次','越高越好','计数','MEK溶剂擦拭'),
    ('环氧酚醛','环氧-酚醛缩合','环氧/酚醛','水煮等级','级','越低越好','等级','1-5级水煮'),
    ('环氧配比方案','环氧-酚醛缩合','环氧/酚醛','T弯','mm','越低越好','连续','弯曲试验'),
    ('环氧配比方案','环氧-酚醛缩合','环氧/酚醛','MEK擦拭','次','越高越好','计数','MEK溶剂擦拭'),
    ('聚酯金黄','羟基-氨基树脂','聚酯','T弯','mm','越低越好','连续','弯曲试验'),
    ('聚酯金黄','羟基-氨基树脂','聚酯','MEK擦拭','次','越高越好','计数','MEK溶剂擦拭'),
    ('有机','羟基-氨基树脂','丙烯酸/乙烯基/环氧','T弯','mm','越低越好','连续','弯曲试验'),
    ('有机','羟基-氨基树脂','丙烯酸/乙烯基/环氧','MEK擦拭','次','越高越好','计数','MEK溶剂擦拭'),
    ('聚氨酯','羟基-异氰酸酯','聚酯/丙烯酸/聚氨酯','T弯','mm','越低越好','连续','弯曲试验'),
    ('聚氨酯','羟基-异氰酸酯','聚酯/丙烯酸/聚氨酯','MEK擦拭','次','越高越好','计数','MEK溶剂擦拭'),
    ('丙烯酸','自由基聚合/羟基-氨基','丙烯酸','T弯','mm','越低越好','连续','弯曲试验'),
    ('丙烯酸','自由基聚合/羟基-氨基','丙烯酸','MEK擦拭','次','越高越好','计数','MEK溶剂擦拭'),
    ('环氧胺','环氧-胺加成','环氧','T弯','mm','越低越好','连续','弯曲试验'),
    ('环氧胺','环氧-胺加成','环氧','MEK擦拭','次','越高越好','计数','MEK溶剂擦拭'),
]
for row in sys_rows:
    ws.append(list(row))
n_sys = len(sys_rows)
style_table(ws, 1, len(sys_headers), n_sys, kpi_cols=[1,4,5,6,7])
for i, w in enumerate([14,20,22,14,8,12,10,24], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# ================= Sheet3 原料主数据 =================
ws = wb.create_sheet('原料主数据')
mat_headers = ['原料代码','原料名称','所属体系','角色','树脂类型','SMILES','描述符状态',
               '固含NV(%)','密度(g/cm³)','分子量(g/mol)',
               '环氧当量EEW(g/eq)','酸值AV(mgKOH/g)','羟值OHV(mgKOH/g)','胺值(mgKOH/g)','官能度',
               'Tg(℃)','沸点(℃)','闪点(℃)','Hansen δD','Hansen δP','Hansen δH','极性指数','相对挥发速率',
               'C(%)','H(%)','O(%)','N(%)','S(%)','Cl(%)',
               '环氧基(mol/100g)','羟基(mol/100g)','羧基(mol/100g)','酯基(mol/100g)','胺基(mol/100g)',
               '酰胺(mol/100g)','芳香环(mol/100g)','醚键(mol/100g)','蜡含量(%)','颜料含量(%)','数据来源','备注']
ws.append(mat_headers)
names = {'IR190':'9型环氧树脂36%固含','IR809':'环氧树脂55%固含','住友55754G':'住友环氧树脂','RF401':'酚醛固化剂PR401',
         'RF160':'酚醛固化剂PR33160G','RF516':'酚醛固化剂PR516','RF950':'酚醛固化剂PR8219-50','RF956':'酚醛固化剂PR8219-65',
         'RH601':'酚醛固化剂SM601RX75','1510蜡':'1510蜡25%工作液','AZ088':'分散剂BYK088','正丁醇':'正丁醇',
         '补加混合液':'乙二醇单丁醚:二甲苯=2:1','10%磷酸':'磷酸10%水溶液','TF100':'乙烯基树脂','TM004':'乙二醇丁醚',
         'AS400':'丙烯酸树脂','RX170-140':'丙烯酸树脂','40%50177':'环氧树脂40%固含','IR877':'环氧树脂','RJ173M':'聚酯树脂',
         'RJ561':'聚酯树脂','RY460':'黄色颜料','AC040':'助剂','BYK104':'分散剂BYK104','IR909':'环氧树脂','R170M':'环氧树脂',
         'IR557':'环氧树脂','TF022':'乙烯基树脂','TM221':'乙二醇丁醚','IR868':'环氧树脂','RY075N':'黄色颜料','AZ135':'助剂',
         '35.7%白浆':'白色颜料浆35.7%','14.28%炭黑浆料':'炭黑浆料14.28%','3%气硅':'气相二氧化硅3%','20%CAB':'CAB溶液20%',
         '杜邦-FT960':'环氧树脂','AL525':'丙烯酸树脂','AL710':'丙烯酸树脂','AZ306':'助剂','AZ551':'助剂','BYK306':'流平剂',
         'FL208':'助剂','FL208S':'助剂','FL815C':'助剂','IA151':'丙烯酸树脂','IA893':'丙烯酸树脂','IR842':'环氧树脂',
         'RA009':'助剂','RA083':'助剂','RA824':'助剂','RJ183':'聚酯树脂','RJ362':'聚酯树脂','日本151-PVC':'颜料浆',
         'TZ161':'PMA溶剂','TZ425':'DBE溶剂','TZ240':'醋酸丁酯','TT444':'丁酮','TT066':'环己酮','TM982':'PM溶剂',
         'TM024':'二乙二醇单丁醚','TZ221':'乙二醇丁醚','RY078':'黄色颜料','AL800':'丙烯酸树脂','IA800':'丙烯酸树脂',
         'IA8000':'丙烯酸树脂','10%AC040':'AC040 10%'}
for code, d in full_mat.items():
    is_new = code in new_mats
    pv = d.get('prov') or {}
    n_tds = sum(1 for k in CONT_DESC if pv.get(k) in ('tds', 'sds', 'formula', 'tds_carry', 'name'))
    n_any = sum(1 for k in CONT_DESC if pv.get(k) in
                ('tds', 'sds', 'formula', 'tds_carry', 'name', 'compo', 'handbook'))
    src = d.get('数据来源', '')
    if code in _PENDING:
        status = '待确认'
    elif src.startswith('TDS'):
        status = 'TDS实测'
    elif src.startswith('handbook'):
        status = '手册值'
    elif src == 'COMPO_RULES':
        status = '送检组成'
    else:
        status = '专有估算' if is_new else '类别典型值'
    row = [code, names.get(code, code), '多体系', d['role'], d['rtype'], '', status]
    for k in ['NV','density','Mw','EEW','AV','OHV','amine','func','Tg','bp','fp','dD','dP','dH','pol','evap',
              'C','H','O','N','S','Cl','fg_epoxy','fg_oh','fg_cooh','fg_ester','fg_amine','fg_amide','fg_arom','fg_ether','wax','pig']:
        row.append(d[k])
    row.append('估算' if is_new else '类别典型值/文件信息')
    row.append('')
    if src:
        row[-2] = src
        row[-1] = d.get('TDS依据') or d.get('备注', '')
    if n_tds:
        row[-2] = f'{src}({n_tds}/{len(CONT_DESC)}字段档案实测)'
    elif n_any:
        row[-2] = f'{src or "类别典型值"}({n_any}/{len(CONT_DESC)}字段有据可依)'
    if d.get('TDS档案'):
        row[-1] = f"{row[-1]}｜档案：{'、'.join(d['TDS档案'][:3])}"
    ws.append(row)
n_mat = len(full_mat)
style_table(ws, 1, len(mat_headers), n_mat, kpi_cols=[8,11,12,13,14,15,30,31,32,33,34,35,36,37])
widths = [16,20,10,8,9,14,10,10,10,11,12,12,12,11,8,8,9,9,9,9,9,10,8,8,8,8,8,8,10,10,10,10,10,10,10,10,9,9,14,20]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# ================= Sheet3 配方明细 =================
ws = wb.create_sheet('配方明细')
det_headers = ['样本ID','系列','体系','原料代码','用量(g)','角色','树脂类型','标签状态']
ws.append(det_headers)
det_rows = 0
for s in all_samples:
    for code, amt in s['组分'].items():
        ccode = clean_code(code)
        d = full_mat.get(ccode)
        if d is None: continue
        ws.append([s['样本ID'], s['系列'], s['体系'], ccode, round(float(amt),4), d['role'], d['rtype'], s['标签状态']])
        det_rows += 1
style_table(ws, 1, len(det_headers), det_rows, kpi_cols=[4])
for i, w in enumerate([14,12,16,10,8,9,10], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# ================= Sheet4 性能结果 =================
ws = wb.create_sheet('性能结果')
perf_headers = ['样本ID','体系','目标属性','测试值','单位','标签状态','标签来源','测试条件']
ws.append(perf_headers)
perf_rows = 0
for s in all_samples:
    if s['标签状态'] != '实测': continue
    cond = f"{s['烘烤温度']}℃ {s['烘烤时间']}min" if s['烘烤温度'] else ''
    for tgt, val, unit in [('T弯', s['T弯'], 'mm'), ('MEK擦拭', s['MEK'], '次'), ('水煮等级', s['水煮'], '级')]:
        if val is not None and not (isinstance(val, float) and np.isnan(val)):
            ws.append([s['样本ID'], s['体系'], tgt, round(float(val),4), unit, '实测', '实验室', cond])
            perf_rows += 1
style_table(ws, 1, len(perf_headers), perf_rows, kpi_cols=[4])
for i, w in enumerate([14,12,12,10,8,10,10,18], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# ================= Sheet5 工艺条件 =================
ws = wb.create_sheet('工艺条件')
proc_headers = ['样本ID','体系','烘烤温度(℃)','烘烤时间(min)','标签状态']
ws.append(proc_headers)
proc_rows = 0
for s in all_samples:
    ws.append([s['样本ID'], s['体系'], s['烘烤温度'], s['烘烤时间'], s['标签状态']])
    proc_rows += 1
style_table(ws, 1, len(proc_headers), proc_rows)
for i, w in enumerate([14,12,12,12,10], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# ================= Sheet6 配方级描述符 =================
ws = wb.create_sheet('配方级描述符')
ser_map = {s['样本ID']: s['系列'] for s in all_samples}
fd_headers = ['样本ID','系列','体系','标签状态'] + list(desc_df.columns[2:])
ws.append(fd_headers)
for _, row in desc_df.iterrows():
    ws.append([row['样本ID'], ser_map.get(row['样本ID'], ''), row['体系'], '实测' if row['样本ID'] in {s['样本ID'] for s in all_samples if s['标签状态']=='实测'} else '无标签'] + [round(float(v),6) if isinstance(v,(int,float)) else v for v in row[2:]])
n_fd = len(desc_df)
style_table(ws, 1, len(fd_headers), n_fd, kpi_cols=[5,6,7])
for i, w in enumerate([14,10,12,10] + [10]*len(desc_df.columns[2:]), 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'D2'

# ================= Sheet7 建模输入 =================
ws = wb.create_sheet('建模输入')
mi_headers = ['样本ID','系列','体系','标签状态','T弯实测','MEK实测','水煮实测'] + list(desc_df.columns[2:])
ws.append(mi_headers)
lab_map = {s['样本ID']: s for s in all_samples}
for _, row in desc_df.iterrows():
    sid = row['样本ID']
    s = lab_map.get(sid, {})
    t_w = s.get('T弯'); m_w = s.get('MEK'); z_w = s.get('水煮')
    status = '实测' if s.get('标签状态') == '实测' else '无标签'
    ws.append([sid, ser_map.get(sid, ''), row['体系'], status,
               round(float(t_w),4) if t_w is not None else '', 
               round(float(m_w),4) if m_w is not None else '',
               round(float(z_w),4) if z_w is not None else ''] + [round(float(v),6) if isinstance(v,(int,float)) else v for v in row[2:]])
n_mi = len(desc_df)
style_table(ws, 1, len(mi_headers), n_mi, kpi_cols=[5,6,7])
for i, w in enumerate([14,10,12,10,10,10,10] + [10]*len(desc_df.columns[2:]), 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'H2'

# ================= Sheet8 配方级机理特征 =================
# 由 workbench/mech_desc.py 现场计算（不依赖 pkl 中的历史 desc_df），
# 羟基/羧基当量采用羟值/酸值标准换算口径（oh_source='ohv'，单位自洽）。
# 无烘烤记录样本的固化类机理量输出为空单元格（NaN 口径：未记录≠零固化，实验 T）。
ws = wb.create_sheet('配方级机理特征')
from mech_desc import mech_features, MECH_FEATURES
mech_headers = ['样本ID', '系列', '体系', '标签状态'] + MECH_FEATURES
ws.append(mech_headers)
n_mech = 0
for s in all_samples:
    d, _err = mech_features(s['组分'], full_mat, s.get('烘烤温度'), s.get('烘烤时间'),
                            oh_source='ohv', nan_no_bake=True)
    if d is None:
        continue
    status = '实测' if s['标签状态'] == '实测' else '无标签'
    vals = []
    for f in MECH_FEATURES:
        v = float(d.get(f, 0.0))
        vals.append('' if np.isnan(v) else round(v, 6))
    ws.append([s['样本ID'], s.get('系列', ''), s['体系'], status] + vals)
    n_mech += 1
style_table(ws, 1, len(mech_headers), n_mech, kpi_cols=[5, 13, 19, 24, 30])
for i, w in enumerate([14, 10, 12, 10] + [12] * len(MECH_FEATURES), 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'E2'

# ================= Sheet9 数据字典 =================
ws = wb.create_sheet('数据字典')
dict_headers = ['字段名','所属工作表','类型','单位','口径说明','示例']
ws.append(dict_headers)
dict_rows = [
    ('样本ID','配方明细/性能结果/工艺条件/配方级描述符/建模输入','文本','-','全局唯一，关联各表','D1-1'),
    ('系列','配方明细/配方级描述符/建模输入','文本','-','配方系列/家族（如D1、R01），用于系列目标编码建模','D1'),
    ('体系','配方明细/性能结果','文本','-','化学体系类别：环氧酚醛/环氧-配比方案/聚酯金黄','环氧酚醛'),
    ('原料代码','原料主数据/配方明细','文本','-','原料唯一编码，与原料主数据一致','IR190'),
    ('用量','配方明细','数值','g','该组分在样本中的质量份','66.0'),
    ('角色','原料主数据/配方明细','枚举','-','树脂/固化剂/溶剂/助剂/颜料','树脂'),
    ('树脂类型','原料主数据/配方明细','枚举','-','环氧/酚醛/聚酯/乙烯基/丙烯酸/聚氨酯/氨基/其他','环氧'),
    ('标签状态','配方明细/性能结果/工艺条件/建模输入','枚举','-','实测/无标签/伪标签/推荐测试','实测'),
    ('T弯实测','建模输入','数值','mm','杯突/弯曲试验结果（越低越好）','17.415'),
    ('MEK实测','建模输入','数值','次','MEK溶剂擦拭次数（越高越好，封顶300）','9'),
    ('水煮实测','建模输入','数值','级','水煮等级1-5（越低越好）','4'),
    ('固含NV','原料主数据','数值','%','按到货状态的固体含量','36'),
    ('密度','原料主数据','数值','g/cm³','原料密度','1.00'),
    ('分子量','原料主数据','数值','g/mol','数均分子量（树脂可为典型值）','1400'),
    ('环氧当量EEW','原料主数据','数值','g/eq','含1mol环氧基的到货产品质量','2640'),
    ('酸值AV','原料主数据','数值','mgKOH/g','中和1g样品游离酸所需KOH','0.5'),
    ('羟值OHV','原料主数据','数值','mgKOH/g','中和1g样品羟基所需KOH','30'),
    ('Tg','原料主数据','数值','℃','玻璃化转变温度','70'),
    ('Hansen δD/δP/δH','原料主数据','数值','MPa^0.5','Hansen溶解度参数三分量','18.5/6.0/8.5'),
    ('C/H/O/N/S/Cl','原料主数据','数值','%','元素质量分数','72/7/20/0/0/0'),
    ('环氧基/羟基/羧基/酯基/胺基/酰胺/芳香环/醚键','原料主数据','数值','mol/100g','官能团密度','0.038'),
    ('树脂/固化剂/溶剂/助剂/颜料占比','配方级描述符','数值','-','各角色质量分数','0.72'),
    ('加权固含等','配方级描述符','数值','-','按质量分数加权的原料描述符','36.5'),
    ('环氧基密度等','配方级描述符','数值','mol/100g','每100g配方的官能团摩尔数','0.05'),
    ('烘烤温度/时间','工艺条件','数值','℃/min','固化工艺参数','205/17'),
    ('描述符状态','原料主数据','枚举','-','TDS实测=供应商技术档案逐字段替换；送检组成/手册值/类别典型值/待确认 可信度依次降低','TDS实测'),
    ('数据来源','原料主数据','文本','-','TDS/SDS(n/32字段档案实测)=供应商技术档案覆盖；COMPO_RULES=送检组成；handbook:*=公开手册或名称自证；空=类别典型值估算','TDS/SDS(11/32字段档案实测)'),
    # ---- 机理特征（配方级机理特征表）----
    ('solids_frac / binder_solids_frac','配方级机理特征','数值','-','全配方固体分与结合料（树脂+固化剂）固体分质量分数','0.36'),
    ('eq_epoxy / eq_oh_phenol / eq_oh_ali / eq_oh_all','配方级机理特征','数值','mol/100g','环氧当量、酚羟基、脂肪族羟基、全部活性氢当量浓度；羟基按羟值标准换算（OHV/56.1）','0.034'),
    ('eq_cooh / eq_nco / eq_amine / eq_cat','配方级机理特征','数值','mol/100g','羧基、异氰酸酯、氨基树脂活性氢、催化当量浓度','0.001'),
    ('r_phenol_epoxy / r_oh_epoxy','配方级机理特征','数值','-','环氧当量对活性氢当量的化学计量比','0.56'),
    ('r_nco_oh / r_amino_oh','配方级机理特征','数值','-','异氰酸酯/氨基对羟基化学计量比（无该机制时为0）','0.21'),
    ('stoich_dev_epoxy / stoich_dev_nco','配方级机理特征','数值','-','|1−r| 当量偏离度，衡量配比是否接近化学计量点','0.44'),
    ('f_bar','配方级机理特征','数值','-','按当量加权的平均官能度','2.3'),
    ('ne_potential / ne_effective','配方级机理特征','数值','mol/100g结合料','Flory-Stockmayer 量级潜在/有效交联密度（有效值含固化度折扣；无烘烤记录时 ne_effective 留空）','0.016'),
    ('tg_fox_solids','配方级机理特征','数值','℃','结合料固体分的 Fox 共混玻璃化转变（1/T 加权，非线性）','70.1'),
    ('cure_margin / cure_margin_eff','配方级机理特征','数值','℃','烘烤温度与共混Tg/固化后Tg之差，为负表示玻璃化受限；无烘烤记录时留空','90.7'),
    ('t_eff_min','配方级机理特征','数值','min','以200℃为参考的Arrhenius等效固化时间（Ea=90kJ/mol）；无烘烤记录时留空','21.6'),
    ('h_d_resin_solvent / h_d_min_pair','配方级机理特征','数值','MPa^0.5','溶剂相与树脂相的Hansen距离（Ra）及最不利配对距离','5.4'),
    ('pvc','配方级机理特征','数值','-','颜料体积浓度=颜料体积/(颜料+结合料体积)，按密度换算','0.18'),
    ('cat_per_epoxy_eq','配方级机理特征','数值','-','催化当量/环氧当量，环氧-酚醛固化速率主变量','0.02'),
]
for row in dict_rows:
    ws.append(list(row))
n_dict = len(dict_rows)
style_table(ws, 1, len(dict_headers), n_dict)
for i, w in enumerate([34,34,8,12,46,14], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '合并版数据集.xlsx')
wb.save(out)
print('saved:', out)
_st = {}
for _c, _m in full_mat.items():
    _src = _m.get('数据来源') or '类别典型值'
    _st[_src] = _st.get(_src, 0) + 1
print(f"原料 {n_mat} 种（来源分布 {_st} / 待确认 {len(_PENDING)} / 同物合并 {len(_MERGE)}）, "
      f"配方明细 {det_rows} 行, 性能 {perf_rows} 行, 描述符 {n_fd} 样本×{len(desc_df.columns)-2} 特征, "
      f"机理特征 {n_mech} 样本×{len(MECH_FEATURES)} 特征")
