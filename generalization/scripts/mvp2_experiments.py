# -*- coding: utf-8 -*-
"""
MVP 全流程验证 v2 —— 解决局限 + 更严谨验证
==========================================
实验E: 半监督伪标签回放（模拟缺失标签，评估回放是否提升泛化）
实验F: 主动学习模拟（不确定性采样 vs 随机采样，评估"最少实验补齐标签"）
实验G: 跨体系标签补充应用（环氧酚醛训练 -> 有机/聚酯预测 -> 伪标签+主动学习推荐）
实验H: 跨体系迁移学习（环氧酚醛预训练 -> 有机/聚酯少量标签微调）
"""
import json, pickle, warnings
import numpy as np
import pandas as pd
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import r2_score, mean_absolute_error
from sklearn.model_selection import train_test_split
from descriptors import formulation_descriptors, DESC_FEATURES, resolve
from materials import MAT
import openpyxl
warnings.filterwarnings('ignore')

SEED = 42
np.random.seed(SEED)

# 数据源路径可通过命令行参数覆盖：python mvp2_experiments.py f2=xxx.xlsx f3=yyy.xlsx f4=zzz.xlsx out=结果.pkl
import sys, os
_ARGS = dict(a.split('=', 1) for a in sys.argv[1:] if '=' in a)
def _p(key, default):
    return _ARGS.get(key, default)

# ============ 数据载入（环氧酚醛） ============
f2 = _p('f2', "/workspace/.uploads/54d2ddcf-cfaf-4714-9a8f-f96748dc0968_配料测试数据汇总V1.xlsx")
df = pd.read_excel(f2, sheet_name="配方与结果数据集V1")
comp_cols = [c for c in df.columns if c not in [
    '批次','追溯编号','配方ID','配方系列','配方类型','线棒号','烘烤条件',
    'T弯(mm)_原始','MEK擦拭(次)_原始','水煮（等级）_原始','检测指标数量','检测完整率',
    '检测完整性','复核状态','来源文件']]
df['T弯'] = pd.to_numeric(df['T弯(mm)_原始'], errors='coerce')
df['MEK'] = pd.to_numeric(df['MEK擦拭(次)_原始'].astype(str).str.replace('300+','300'), errors='coerce')
df['水煮'] = pd.to_numeric(df['水煮（等级）_原始'], errors='coerce')

desc_rows = []
for _, row in df.iterrows():
    comp = {c: row[c] for c in comp_cols}
    d = formulation_descriptors(comp)
    if d:
        desc_rows.append(d)
desc_df = pd.DataFrame(desc_rows)[DESC_FEATURES]
raw_df = df[comp_cols].reset_index(drop=True)
print("环氧酚醛: 配方", len(desc_df), "描述符", desc_df.shape[1])

TARGETS = {'T弯': 'T弯', 'MEK': 'MEK', '水煮': '水煮'}
results = {}

def rf():
    return RandomForestRegressor(n_estimators=300, random_state=SEED, n_jobs=-1)

# ============ 实验E: 半监督伪标签回放 ============
print("\n===== 实验E: 半监督伪标签回放 =====")
for tname, col in TARGETS.items():
    mask = df[col].notna()
    X = desc_df.loc[mask].values
    y = df.loc[mask, col].values
    n = len(y)
    rng = np.random.RandomState(SEED)
    idx = rng.permutation(n)
    n_test = max(20, int(n*0.2))
    n_pseudo = int((n-n_test)*0.3)
    test_idx, pseudo_idx = idx[:n_test], idx[n_test:n_test+n_pseudo]
    lab_idx = idx[n_test+n_pseudo:]
    X_lab, y_lab = X[lab_idx], y[lab_idx]
    X_ps, y_ps_true = X[pseudo_idx], y[pseudo_idx]
    X_te, y_te = X[test_idx], y[test_idx]

    base = rf().fit(X_lab, y_lab)
    base_r2 = r2_score(y_te, base.predict(X_te))
    base_mae = mean_absolute_error(y_te, base.predict(X_te))

    # 半监督：预测缺失标签 -> 按不确定性分层
    preds = np.array([t.predict(X_ps) for t in base.estimators_])
    mean, std = preds.mean(0), preds.std(0)
    rel_std = std/(np.abs(mean)+1e-6)
    q = np.quantile(rel_std, 0.5)
    conf = rel_std < q
    n_conf = conf.sum()

    # 伪标签回放（权重0.5）
    if n_conf >= 5:
        X_all = np.vstack([X_lab, X_ps[conf]])
        y_all = np.concatenate([y_lab, mean[conf]])
        w_all = np.concatenate([np.ones(len(y_lab)), np.full(n_conf, 0.5)])
        semi = rf().fit(X_all, y_all, sample_weight=w_all)
        semi_r2 = r2_score(y_te, semi.predict(X_te))
        semi_mae = mean_absolute_error(y_te, semi.predict(X_te))
    else:
        semi_r2, semi_mae = base_r2, base_mae

    # 上界：真实标签回放
    X_all = np.vstack([X_lab, X_ps])
    y_all = np.concatenate([y_lab, y_ps_true])
    oracle = rf().fit(X_all, y_all)
    oracle_r2 = r2_score(y_te, oracle.predict(X_te))

    pseudo_corr = float(np.corrcoef(mean, y_ps_true)[0,1]) if n_conf>1 else 0.0
    results[f'E_{tname}'] = {
        'n_total': n, 'n_lab': len(lab_idx), 'n_pseudo': n_pseudo, 'n_conf': int(n_conf),
        'base_r2': base_r2, 'base_mae': base_mae,
        'semi_r2': semi_r2, 'semi_mae': semi_mae,
        'oracle_r2': oracle_r2, 'pseudo_corr': pseudo_corr}
    print(f"{tname}: 基线R2={base_r2:.3f} -> 半监督R2={semi_r2:.3f} (上界={oracle_r2:.3f}) | 伪标签相关={pseudo_corr:.3f} | 高置信{n_conf}")

# ============ 实验F: 主动学习模拟 ============
print("\n===== 实验F: 主动学习模拟 =====")
for tname, col in TARGETS.items():
    mask = df[col].notna()
    X = desc_df.loc[mask].values
    y = df.loc[mask, col].values
    n = len(y)
    rng = np.random.RandomState(SEED)
    pool = rng.permutation(n)
    n_init = max(10, int(n*0.1))
    n_test = max(20, int(n*0.2))
    lab = pool[:n_init].tolist()
    test = pool[n_init:n_init+n_test].tolist()
    pool = pool[n_init+n_test:].tolist()

    def eval_r2(lab_set):
        m = rf().fit(X[lab_set], y[lab_set])
        return r2_score(y[test], m.predict(X[test]))

    al_r2s, rand_r2s = [eval_r2(lab)], [eval_r2(lab)]
    n_per = 8
    for _ in range(5):
        # 主动学习：不确定性采样
        m = rf().fit(X[lab], y[lab])
        preds = np.array([t.predict(X[pool]) for t in m.estimators_])
        std = preds.std(0)
        pick_al = np.argsort(std)[-n_per:]
        lab_al = [pool[i] for i in pick_al]
        lab_new = lab + lab_al
        pool_new = [p for p in pool if p not in lab_al]
        al_r2s.append(eval_r2(lab_new))

        # 随机采样（独立副本）
        rng2 = np.random.RandomState(SEED*2 + len(al_r2s))
        rand_pick = rng2.choice(pool, n_per, replace=False)
        lab_rand = lab + list(rand_pick)
        rand_r2s.append(eval_r2(lab_rand))

        lab, pool = lab_new, pool_new

    results[f'F_{tname}'] = {'al_r2s': al_r2s, 'rand_r2s': rand_r2s,
                             'n_init': n_init, 'n_per_round': n_per, 'n_rounds': 5}
    print(f"{tname}: 初始R2={al_r2s[0]:.3f} | 主动学习终值={al_r2s[-1]:.3f} | 随机终值={rand_r2s[-1]:.3f}")

# ============ 实验G: 跨体系标签补充应用 ============
print("\n===== 实验G: 跨体系标签补充应用 =====")
def extract_formulations(f):
    wb = openpyxl.load_workbook(f, data_only=True)
    forms = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        for r_idx in range(min(len(rows), 12)):
            row = rows[r_idx]
            for c_idx, cell in enumerate(row):
                if isinstance(cell, str) and resolve(cell) and c_idx > 0:
                    comp = {}
                    for rr in range(r_idx, len(rows)):
                        code = rows[rr][c_idx]
                        if isinstance(code, str) and resolve(code):
                            for cc in range(c_idx+1, len(rows[rr])):
                                v = rows[rr][cc]
                                if isinstance(v, (int, float)) and v > 0:
                                    comp[code] = v
                                    break
                    if len(comp) >= 3:
                        forms.append((sn, comp))
                    break
    return forms

f3 = _p('f3', "/workspace/.uploads/e657c941-a24b-44f8-839c-89668666da39_fbfc94091fa4955969e5d0fff7df0fd6_789507228604997242_m_3NX240913-6C--AI研发26.7.22配比方案.xlsx")
f4 = _p('f4', "/workspace/.uploads/f43046fb-7411-4ef6-9e84-4af758d8419e_聚酯金黄-AI(1).xlsx")
forms3 = extract_formulations(f3)
forms4 = extract_formulations(f4)
print("有机体系配方:", len(forms3), "聚酯体系配方:", len(forms4))

# 编码
def encode(forms):
    rows, meta = [], []
    for sn, comp in forms:
        d = formulation_descriptors(comp)
        if d:
            rows.append([d[f] for f in DESC_FEATURES])
            meta.append((sn, comp))
    return np.array(rows), meta

X3, meta3 = encode(forms3)
X4, meta4 = encode(forms4)
print("有机可编码:", X3.shape, "聚酯可编码:", X4.shape)

# 用环氧酚醛全量训练，预测有机/聚酯
label_results = {}
for tname, col in TARGETS.items():
    mask = df[col].notna()
    m = rf().fit(desc_df.loc[mask].values, df.loc[mask, col].values)
    for sysn, Xs, metas in [('有机', X3, meta3), ('聚酯', X4, meta4)]:
        if len(Xs) == 0:
            continue
        preds = np.array([t.predict(Xs) for t in m.estimators_])
        mean, std = preds.mean(0), preds.std(0)
        rel_std = std/(np.abs(mean)+1e-6)
        q_conf, q_act = np.quantile(rel_std, 0.5), np.quantile(rel_std, 0.8)
        status = np.where(rel_std < q_conf, '伪标签',
                 np.where(rel_std < q_act, '推荐测试', '人工复核'))
        key = f'G_{sysn}_{tname}'
        label_results[key] = {
            'n': len(Xs),
            'n_pseudo': int((status=='伪标签').sum()),
            'n_active': int((status=='推荐测试').sum()),
            'n_review': int((status=='人工复核').sum()),
            'mean_pred': float(np.mean(mean)),
            'mean_std': float(np.mean(std)),
        }
        print(f"{sysn}-{tname}: n={len(Xs)} | 伪标签{(status=='伪标签').sum()} 推荐测试{(status=='推荐测试').sum()} 人工复核{(status=='人工复核').sum()} | 预测均值{np.mean(mean):.2f}")

# ============ 实验H: 跨体系迁移学习（模拟） ============
print("\n===== 实验H: 跨体系迁移学习模拟 =====")
# 模拟：环氧酚醛为源域，有机/聚酯为目标域（用环氧酚醛内部模拟跨域）
# 用"配方系列"作为域划分：训练在部分系列，微调在另一些系列
series_list = sorted(df['配方系列'].unique())
for tname, col in TARGETS.items():
    mask = df[col].notna()
    X = desc_df.loc[mask].values
    y = df.loc[mask, col].values
    ser = df.loc[mask, '配方系列'].values
    # 源域: 前70%系列, 目标域: 后30%系列
    n_ser = len(series_list)
    src_series = set(series_list[:int(n_ser*0.7)])
    src = np.array([s in src_series for s in ser])
    X_src, y_src = X[src], y[src]
    X_tgt, y_tgt = X[~src], y[~src]
    if len(X_tgt) < 20:
        continue
    # 目标域再拆: 少量标签(20%) + 测试(80%)
    rng = np.random.RandomState(SEED)
    n_tgt = len(y_tgt)
    idx = rng.permutation(n_tgt)
    n_lab = max(5, int(n_tgt*0.2))
    tgt_lab, tgt_test = idx[:n_lab], idx[n_lab:]
    X_tl, y_tl = X_tgt[tgt_lab], y_tgt[tgt_lab]
    X_tt, y_tt = X_tgt[tgt_test], y_tgt[tgt_test]

    # 1) 只用目标域少量标签训练
    m_tl = rf().fit(X_tl, y_tl)
    tl_r2 = r2_score(y_tt, m_tl.predict(X_tt))
    # 2) 源域预训练 + 目标域微调（用源域+目标标签混合，目标加权）
    X_all = np.vstack([X_src, X_tl])
    y_all = np.concatenate([y_src, y_tl])
    w_all = np.concatenate([np.ones(len(y_src)), np.full(len(y_tl), 3.0)])
    m_tl2 = rf().fit(X_all, y_all, sample_weight=w_all)
    tl2_r2 = r2_score(y_tt, m_tl2.predict(X_tt))

    results[f'H_{tname}'] = {'src_n': len(X_src), 'tgt_n': len(X_tgt),
                             'tgt_lab_n': len(y_tl), 'tl_only_r2': tl_r2, 'transfer_r2': tl2_r2}
    print(f"{tname}: 源域{len(X_src)} 目标域{len(X_tgt)}(标签{len(y_tl)}) | 仅目标标签R2={tl_r2:.3f} -> 迁移R2={tl2_r2:.3f}")

# ============ 保存 ============
out = _p('out', os.path.join(os.path.dirname(os.path.abspath(__file__)), 'mvp2_results.pkl'))
with open(out, 'wb') as fh:
    pickle.dump({'results': results, 'label_results': label_results,
                 'X3': X3, 'X4': X4, 'meta3': meta3, 'meta4': meta4,
                 'desc_df': desc_df, 'df': df}, fh)
print("\n全部实验完成，结果已保存", out)
