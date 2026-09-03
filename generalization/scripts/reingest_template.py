# -*- coding: utf-8 -*-
"""
终极版数据集模板 · 原始数据重录入
==========================================
把 5 份原始实验记录解析为「配方-性能-工艺」记录，写入 generalization/终极版数据集模板.xlsx：

  1. 7.26配料测试汇总.xlsx         宽表（列=原料，行=配方×线棒号）    烘烤 200℃/10min
  2. 8.6配料测试汇总.xlsx          宽表 + 换算明细 + 测试原始数据     烘烤 205℃/17min
  3. 8.14配料测试汇总.xlsx         宽表 + 测试原始数据（同批配方以更晚的 8.14 为准，
                                   C7 系列仅 8.6 记录）
  4. 3NX240913-6C--AI研发26.7.22配比方案.xlsx   矩阵表（行=原料/性能项，列=配方）
  5. 聚酯金黄-AI(1).xlsx                        矩阵表

矩阵表解析口径：
  · 每个「配方名称行」定义一个块；样本列 = 名称行上的配方列，排除 原始配方/百分比/1000KG/
    500克/250克/物料比例/固含 等折算列（其性能回填到与之成正比的配方列，即同配方放大批次）。
  · 原料代码取自该样本列左侧最近的「代码列」；代码列单元格仅在同行有数值时生效，用于同一
    配方块内牌号不一致（如 RX170-140 与 日本151-PVC 并列）。
  · 原料行中出现「已定义配方名」时按该配方的组成等比展开（如 13B-2A = 6G-13B-2 100g + 助剂）。
  · 合计行按各原料行之和识别并逐列校验，差异 >0.5g 进入审计报告。
  · 性能块按标签语义检测整块错位（个别表的数值列相对标签列上移一行），逐表报告。
  · 组成块与性能块的配方列下标不一致时按配方名对齐；名称行失效（陈旧名）时按列下标对齐并报告。
  · 同一配方名跨表重复：组成一致则合并为一条样本、性能按表分别记录；不一致则「名称@表名」分列。

性能取值沿用仓库既有口径（scripts/extract_perf_labels.py）：字段带 MEK/T弯/水煮 即视为同一种
测试；T弯G 冲击-5%硫酸铜腐蚀判定按字面档位换算、开区间按配方脆性定档；点状/有改善等无法定级
者留空。原始写法写入「性能结果.备注」。

用法：
  python scripts/reingest_template.py --report    仅打印解析审计，不写文件
  python scripts/reingest_template.py             解析并写出模板
"""
import os
import re
import sys
import glob
import pickle
import hashlib
import unicodedata
from collections import OrderedDict, defaultdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, '..', '..'))
WORKBENCH = os.path.join(HERE, '..', 'workbench')
UPLOADS = os.path.join(HERE, '..', 'raw')
OUT = os.path.join(HERE, '..', '终极版数据集模板.xlsx')
LEGACY_PKL = os.path.join(HERE, '..', 'data', 'merged_data.pkl')
sys.path.insert(0, HERE)
sys.path.insert(0, WORKBENCH)

from materials import MAT, ALIAS, CONT_DESC                                    # noqa: E402
from compo_rules import COMPO                                                   # noqa: E402
from DataPrepWorkbench import est_material                                     # noqa: E402
import handbook_fixes as HF                                                    # noqa: E402
import tds_sds                                                                 # noqa: E402
from extract_perf_labels import classify, parse_target_value, _assign_bands    # noqa: E402

TARGETS = ['T弯', 'MEK擦拭', '水煮等级']
UNIT = {'T弯': 'mm', 'MEK擦拭': '次', '水煮等级': '级'}
SYS_EPOXY, SYS_PEIBI, SYS_JINHUANG = '环氧酚醛', '环氧配比方案', '聚酯金黄'

SYSTEM_CONFIG = OrderedDict([
    (SYS_EPOXY, ('环氧-酚醛缩合', '环氧/酚醛',
                 [('T弯', 'mm', '越低越好', '连续', '杯突/弯曲试验'),
                  ('MEK擦拭', '次', '越高越好', '计数', 'MEK溶剂擦拭'),
                  ('水煮等级', '级', '越低越好', '等级', '121℃/60min 水煮，1-4 级')])),
    (SYS_PEIBI, ('环氧-酚醛缩合', '环氧/酚醛/乙烯基/丙烯酸/聚酯',
                 [('T弯', 'mm', '越低越好', '连续', 'T弯G冲击-5%硫酸铜*30min 腐蚀判定定档'),
                  ('MEK擦拭', '次', '越高越好', '计数', 'MEK溶剂擦拭'),
                  ('水煮等级', '级', '越低越好', '等级', '121℃/60min 杀菌水煮')])),
    (SYS_JINHUANG, ('羟基-氨基树脂', '聚酯/环氧/丙烯酸/乙烯基',
                    [('T弯', 'mm', '越低越好', '连续', '杯突/弯曲试验'),
                     ('MEK擦拭', '次', '越高越好', '计数', 'MEK溶剂擦拭'),
                     ('水煮等级', '级', '越低越好', '等级', '121℃*60min 水煮，1-4 级')])),
])

# 配料测试汇总：批次 → (文件关键字, 配料日期, 烘烤温度, 烘烤时间)
WIDE_SOURCES = [
    ('7.26', '7.26配料测试汇总', '2025-07-26', 200, 10),
    ('8.6', '8.6配料测试汇总', '2025-08-06', 205, 17),
    ('8.14', '8.14配料测试汇总', '2025-08-14', 205, 17),
]
WIDE_PERF = {'T弯': r'^T弯\(mm\)$', 'MEK': r'^MEK擦拭\(次\)$', '水煮': r'^水煮[（(]等级[）)]$'}
WIDE_SKIP = ('序号', '配方ID', '线棒号', '烘烤条件', '批次')

# 矩阵表：文件关键字 → 体系
MATRIX_SOURCES = [('配比方案', SYS_PEIBI), ('聚酯金黄', SYS_JINHUANG)]
# 代码名自证的描述符：名称已指明类别与到货状态，无需按类别典型值估算
SELF_EVIDENT = {'100#溶剂油': dict(role='溶剂', rtype='其他', NV=0, func=0, wax=0, pig=0)}
# 代码列写法 → 登记代码（原始表中以文字而非数值记录的补加料等）
CODE_FIX = {'补加100#': '100#溶剂油', 'PVC-浆料': '日本151-PVC', '阿克苏-6#炭黑': '6#炭黑-阿克苏'}
SCALE_WORDS = ('百分比', '占比', '物料比例', '原始配方', '固含', 'KG', '克')
METRIC_WORDS = ('腐蚀', '合格', '外观', '说明', '背景', '应用', '工艺', '基材', '要求', '煮', '盐',
                '酸', '硫', 'BOX', '冲击', '刮', '硬度', '粘度', '滑度', '粒子', '流平', '名称',
                '铁', '棒', '蒸汽', '圈', '花架', '气泡', '照片', '日期', '完好', '泛白', '杀菌')
NON_CODE = {'MEK', 'T弯', '合计', '序号', '备注', '粘度', '固含', '1000KG', '500克', '250克', '1KG'}
NUM_RE = re.compile(r'^-?\d+(?:\.\d+)?$')


# ---------------------------------------------------------------- 基础工具
def txt(v):
    if v is None:
        return ''
    return unicodedata.normalize('NFKC', str(v)).replace('\n', '').replace(' ', '').strip()


def show(v):
    """备注用文本：保留原始写法（仅去换行）。"""
    return '' if v is None else str(v).replace('\n', '').strip()


def num(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        return None if f != f else f
    s = txt(v)
    if not s:
        return None
    if NUM_RE.match(s):
        return float(s)
    m = re.match(r'^(\d+(?:\.\d+)?)\s*(?:克|g)$', s)
    return float(m.group(1)) if m else None


def md5(path):
    return hashlib.md5(open(path, 'rb').read()).hexdigest()


def find_source(keyword):
    hits = [p for p in sorted(glob.glob(os.path.join(UPLOADS, '*.xlsx')))
            if keyword in os.path.basename(p) and not os.path.basename(p).startswith('~$')]
    if not hits:
        raise FileNotFoundError(f'未找到原始文件：{keyword}')
    uniq = OrderedDict((md5(p), p) for p in hits)
    if len(uniq) > 1:
        p = max(hits, key=os.path.getmtime)
        print(f'  [提示] 「{keyword}」有 {len(uniq)} 份内容不同的副本，取最后修改的一份：{os.path.basename(p)}')
        return p
    return next(iter(uniq.values()))


def canon(code):
    """原始写法 → 登记代码：按归一化文本（去空格/换行、全角转半角）匹配别名表与原料库键。"""
    key = txt(code)
    key = _ALIAS_N.get(key, key)
    if key in MAT:
        return key
    return _MAT_N.get(key.upper(), key)


VOCAB = set()
NAME_CELLS = set()



def mat_like(s):
    if not s or len(s) > 24 or s in NON_CODE or s == '合计':
        return False
    if NUM_RE.match(s) or '℃' in s or '°C' in s or re.search(r'\d+\s*min', s, re.I):
        return False
    if any(w in s for w in METRIC_WORDS):
        return False
    return classify(s) is None


_ALIAS_N = {txt(k): txt(v) for k, v in ALIAS.items()}
_MAT_N = {txt(k).upper(): k for k in MAT}


KNOWN_CODES = {txt(k) for k in MAT} | {txt(v) for v in ALIAS.values()}


def is_code(cell):
    """是否原料代码：已登记原料直接认定；未登记的按形状+词表认定。

    （形状判断会把「10%磷酸」这类含「酸」字的牌号误判为性能项，故先查已登记原料。）"""
    s = txt(cell)
    if not s or s == '合计':
        return False
    c = canon(CODE_FIX.get(s, s))
    if c in KNOWN_CODES or s in KNOWN_CODES:
        return True
    return mat_like(s) and c in VOCAB


def build_vocab(paths):
    """矩阵表代码列里出现过的写法 → 原料代码词表（含跨表复用与文字计量的补加料）。

    先收集「配方名称行」上的配方名：它们可能同时出现在代码列（如 6G-13B-2 被别的配方引用），
    属于配方而非原料，不进入原料词表。"""
    for p in paths:
        wb0 = openpyxl.load_workbook(p, data_only=True)
        for ws0 in wb0.worksheets:
            g0 = grid(ws0)
            for row in g0:
                if any(num(row[j]) is not None for j in range(1, len(row))):
                    continue
                for j in range(1, len(row)):
                    s0 = txt(row[j])
                    if s0 and formula_name(s0):
                        NAME_CELLS.update({s0, canon(s0)})
    for p in paths:
        wb = openpyxl.load_workbook(p, data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for idx in (0, 1):
                    if len(row) <= idx or not isinstance(row[idx], str):
                        continue
                    s = txt(row[idx])
                    if not mat_like(s):
                        continue
                    vals = [num(v) for v in row[idx + 1:]]
                    if not any(v and v > 0 for v in vals):
                        continue
                    if s in ('50KG', '2H', '3H', 'S', 'X', 'OK') or s in NAME_CELLS:
                        continue
                    if canon(CODE_FIX.get(s, s)) in NAME_CELLS:
                        continue
                    VOCAB.add(canon(CODE_FIX.get(s, s)))
    VOCAB.update(canon(k) for k in MAT)
    VOCAB.update(canon(v) for v in ALIAS.values())
    VOCAB.update(canon(v) for v in CODE_FIX.values())
    VOCAB.difference_update({'', 'MEK', 'T弯', '合计', '序号'})


def grid(ws, max_col=40):
    mc = min(ws.max_column or 0, max_col)
    if mc < 2:
        return []
    return [[ws.cell(r, c).value for c in range(1, mc + 1)] for r in range(1, (ws.max_row or 0) + 1)]


def sheet_date(name):
    m = re.match(r'^(\d{2})\.(\d{1,2})\.(\d{1,2})', name)
    if not m:
        return ''
    y, mo, d = m.groups()
    return f'20{y}-{int(mo):02d}-{int(d):02d}'


def proportional(va, vb):
    """两条列向量（按原料行下标对齐）是否严格成正比 → 识别百分比/放大批次等折算列。"""
    ks = [k for k, x in va.items() if x and x > 0 and vb.get(k)]
    if len(ks) < 3:
        return False
    r0 = va[ks[0]] / vb[ks[0]]
    return r0 > 0 and all(abs(va[k] / vb[k] - r0) <= 1e-6 * max(1.0, abs(r0)) for k in ks)


# ---------------------------------------------------------------- 宽表：配料测试汇总
def parse_wide(path, batch, date, bake):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[[x for x in wb.sheetnames if '配料' in x or x == 'Sheet1'][0]]
    hdr = [txt(c.value) for c in ws[1]]
    col = {name: i for i, name in enumerate(hdr) if name}
    perf_col = {k: next((n for n in hdr if n and re.match(p, n)), None) for k, p in WIDE_PERF.items()}
    mat_cols = [(n, canon(CODE_FIX.get(n, n))) for n in hdr
                if n and n not in WIDE_SKIP and n not in [v for v in perf_col.values() if v] and is_code(n)]
    recs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sid = txt(row[col['配方ID']])
        if not sid:
            continue
        comp = OrderedDict()
        for _, code in mat_cols:
            v = num(row[col[[n for n, c in mat_cols if c == code][0]]])
            if v and v > 0:
                comp[code] = round(comp.get(code, 0.0) + float(v), 4)
        perf = OrderedDict()
        for key, hname in perf_col.items():
            raw = txt(row[col[hname]]) if hname and hname in col else ''
            if raw and raw not in ('——', '-', '#DIV/0!', '#VALUE!', '#REF!'):
                perf[key] = raw
        recs.append({'样本ID': sid, '系列': re.match(r'(.+)-\d+$', sid).group(1),
                     '线棒': txt(row[col['线棒号']]), '组分': comp, '性能': perf,
                     '批次': batch, '日期': date, '烘烤': bake, '体系': SYS_EPOXY,
                     '文件': os.path.basename(path)})
    return recs


# ---------------------------------------------------------------- 矩阵表：块解析
FORMULA_EXTRA = {'PPG', '8E', '7E-D-1-2'}
NOT_NAME_WORDS = METRIC_WORDS + ('BOX', '单涂', '双涂', 'MEK', '三圈', '杀菌', '涂4', 'CuSO4',
                                 '硫酸铜', '原始配方', '百分比', '放大')


NAME_RE = re.compile(r"^[‘'\"《]{0,2}[0-9A-Za-z]{1,8}[-#](?=.{0,12}[0-9A-Za-z])[-0-9A-Za-z+#]{1,12}$")


def formula_name(s):
    """配方列名：形如 6G-13B-2A / ‘10#-1 / 808-4-A 的牌号（或已知无符号牌号）。

    性能行的取值写法（2- / 22# / 标准模拟液-1H / 15-20mm）与原料代码
    （35.7%白浆-新 / 日本151-PVC / 209-白浆）都不满足该形状。"""
    if not s or len(s) > 24:
        return False
    if s in FORMULA_EXTRA or s.startswith('聚酯金黄-'):
        return True
    if '℃' in s or '°' in s or re.search(r'\d+\s*min', s, re.I) or '<' in s or '>' in s:
        return False
    if any(w in s for w in NOT_NAME_WORDS):
        return False
    if re.search(r'(mm|MM|级|次|目|角|完好|腐蚀|点|合格|min)$', s):
        return False
    return bool(NAME_RE.match(s))


def find_name_rows(g):
    """配方名称行：该行右侧没有数字、且含 ≥1 个配方列名；带 原料代码 表头的行一并识别。"""
    out = []
    for i, row in enumerate(g):
        if any(isinstance(row[j], (int, float)) and num(row[j]) is not None
               for j in range(1, len(row))):
            continue
        texts = {j: txt(row[j]) for j in range(1, len(row)) if txt(row[j])}
        if not texts:
            continue
        code_at = next((j for j, s in texts.items() if s == '原料代码'), None)
        names = {j: s for j, s in texts.items() if formula_name(s)}
        if code_at is not None:
            texts = {j: s for j, s in texts.items() if j > code_at}
            names = dict(texts)
        elif not names:
            continue
        out.append({'row': i, 'names': names, 'cols': texts})
    return out


def block_code_col(g, hi, sample_cols, limit):
    lo = min(sample_cols)
    best, best_n = None, 0
    for j in range(0, lo):
        n = sum(1 for i in range(hi + 1, limit) if j < len(g[i]) and is_code(g[i][j]))
        if n > best_n:
            best, best_n = j, n
    return best if best is not None and best_n >= 2 else max(lo - 1, 0)


def aux_code_cols(g, rows, cc, sample_cols):
    """样本列之间「只有配料标签文本、无数值」的列：对其右侧样本列生效。"""
    stat = defaultdict(lambda: [0, 0])
    for i in rows:
        for j in range(cc + 1, len(g[i])):
            if j in sample_cols or not txt(g[i][j]):
                continue
            if num(g[i][j]) is not None:
                stat[j][1] += 1
            elif is_code(g[i][j]):
                stat[j][0] += 1
    return sorted({cc} | {j for j, (nc, nn) in stat.items() if nc >= 1 and nn == 0})


def valid_label(s, names_all):
    if not s:
        return False
    return is_code(s) or s in names_all or canon(s) in names_all


def cell_label(g, i, j, codes, names_all):
    """样本列 j 在第 i 行的配料标签：优先最近的左侧代码列中「有效的原料/配方写法」。"""
    for c in sorted([x for x in codes if x < j], reverse=True):
        if c < len(g[i]) and valid_label(txt(g[i][c]), names_all):
            return txt(g[i][c])
    if codes and codes[0] < len(g[i]) and valid_label(txt(g[i][codes[0]]), names_all):
        return txt(g[i][codes[0]])
    return None


def parse_matrix_blocks(path, system, formula_defs):
    """矩阵表 → (样本记录, 审计条目)。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    records, audit = [], []
    for ws in wb.worksheets:
        g = grid(ws)
        if len(g) < 5:
            continue
        nrows = find_name_rows(g)
        if not nrows:
            continue
        names_all = {txt(n) for b in nrows for n in b['names'].values()}
        names_all |= {canon(n) for n in names_all}
        names_all |= set(formula_defs)
        title = txt(g[0][0]) if g and g[0] else ''
        tcode = re.match(r'^([0-9A-Za-z][0-9A-Za-z\-]{3,})', title)

        def disp(name):
            """单配方表保留的折算列（250克 等）用表标题里的产品号作为样本名。"""
            if any(w in name for w in SCALE_WORDS):
                return (tcode.group(1).rstrip('-') if tcode else '') or (ws.title + '-配方')
            return name

        sheet_rec = {'文件': os.path.basename(path)[:24], 'sheet': ws.title, '块': []}
        comps = OrderedDict()
        perf_by_name = OrderedDict()
        proc_bits = []
        for bi, blk in enumerate(nrows):
            hi, names = blk['row'], blk['names']
            end = nrows[bi + 1]['row'] if bi + 1 < len(nrows) else len(g)
            cols = dict(blk['cols'])
            sub = {}
            if hi + 1 < len(g):
                sub = {j: txt(g[hi + 1][j]) for j in range(1, len(g[hi + 1]))
                       if txt(g[hi + 1][j]) and num(g[hi + 1][j]) is None}
            if any(v == '原始配方' for v in sub.values()):
                cols = {j: n for j, n in cols.items() if sub.get(j) == '原始配方'}
                names = {j: n for j, n in names.items() if sub.get(j) == '原始配方'}
            if not names:
                continue
            cc = block_code_col(g, hi, list(names), end)
            ing_rows = [i for i in range(hi + 1, end)
                        if cc < len(g[i]) and is_code(g[i][cc])
                        and any((num(g[i][j]) or 0) > 0 for j in cols if j < len(g[i]))]
            codes = aux_code_cols(g, ing_rows, cc, set(cols))
            for i in range(hi + 1, end):
                if i in ing_rows or not any((num(g[i][j]) or 0) > 0 for j in cols if j < len(g[i])):
                    continue
                if any(valid_label(cell_label(g, i, j, codes, names_all), names_all)
                       for j in cols if (num(g[i][j]) if j < len(g[i]) else None)):
                    ing_rows.append(i)
            ing_rows = sorted(set(ing_rows))
            vecs = {j: {i: ((num(g[i][j]) or 0.0) if j < len(g[i]) else 0.0) for i in ing_rows}
                    for j in cols}
            sums = {j: [v for v in vecs[j].values() if v > 0] for j in cols}
            keep = sorted(names)
            if ing_rows:
                keep, parent = resolve_cols(cols, sums, vecs, keep)
            else:
                parent = {}
            for j in keep:
                comp = OrderedDict()
                for i in ing_rows:
                    v = num(g[i][j]) if j < len(g[i]) else None
                    if not v or v <= 0:
                        continue
                    lab = cell_label(g, i, j, codes, names_all)
                    if lab is None:
                        continue
                    key = txt(lab)
                    if key in names_all or canon(key) in names_all:
                        ref = formula_defs.get(key) or formula_defs.get(canon(key))
                        if ref:
                            base, btot = ref
                            for k2, a2 in base.items():
                                comp[k2] = round(comp.get(k2, 0.0) + a2 * (v / btot), 6)
                        continue
                    code = canon(CODE_FIX.get(key, key))
                    if not mat_like(code):
                        continue
                    comp[code] = round(comp.get(code, 0.0) + float(v), 4)
                if comp:
                    dn = disp(names[j])
                    comps[dn] = {'列': j, '组分': comp}
                    formula_defs[txt(dn)] = (comp, sum(comp.values()))
            labels, metric_rows = {}, []
            for i in range(hi + 1, end):
                if i in ing_rows or cc >= len(g[i]) or txt(g[i][cc]) in ('合计', ''):
                    continue
                lab = txt(g[i][cc])
                if is_code(g[i][cc]) or NUM_RE.match(lab):
                    continue
                vals = {j: g[i][j] for j in range(cc + 1, len(g[i])) if g[i][j] not in (None, '')}
                labels[i] = lab
                if vals:
                    metric_rows.append((i, lab, vals))
            shift = detect_shift(metric_rows, labels)
            for i, lab, vals in metric_rows:
                lab2 = labels.get(i + shift)
                if lab2 is None:
                    continue
                for j, v in vals.items():
                    name = names.get(j)
                    if name not in comps:
                        key = j if j in [d['列'] for d in comps.values()] else parent.get(j)
                        name = next((n for n, d in comps.items() if d['列'] == key), None) if key else None
                    if name is None:
                        continue
                    perf_by_name.setdefault(name, OrderedDict()).setdefault(lab2, []).append(v)
            proc_bits += process_texts(g)
            bad, tot = check_totals(g, hi, end, keep, ing_rows, comps)
            sheet_rec['块'].append({'名称行': hi + 1, '代码列': chr(65 + cc),
                                    '配方列': [names[j] for j in keep],
                                    '剔除列': {cols[j]: ('父列=' + cols[parent[j]]) if j in parent else '折算列'
                                               for j in cols if j not in keep},
                                    '原料行': len(ing_rows), '性能行': len(metric_rows),
                                    '错位取标': shift, '合计差异': bad,
                                    '合计可校验列': f'{len(tot)}/{len(keep)}'})
        for name, d in comps.items():
            records.append({'体系': system, '样本ID': name, 'sheet': ws.title, '组分': d['组分'],
                            '性能': perf_by_name.get(name, OrderedDict()), '日期': sheet_date(ws.title),
                            '批次': ws.title, '工艺文本': ' / '.join(OrderedDict.fromkeys(proc_bits)),
                            '文件': os.path.basename(path)})
            formula_defs[txt(name)] = (d['组分'], sum(d['组分'].values()))
        for name in perf_by_name:
            if name not in comps:
                sheet_rec.setdefault('仅性能无组成', []).append(name)
        audit.append(sheet_rec)
    return records, audit


def process_texts(g):
    out = []
    for row in g:
        for v in row:
            s = show(v)
            if s and len(s) > 5 and not is_code(s) \
                    and re.search(r'(单涂|双涂|两涂|一烘|烘烤|基材|膜厚|涂布|涂4|镀(铬|锡)|铬铁|线棒|刮棒)', s) \
                    and not re.search(r'(问题|应用|要求|满足|加工性能|客户)', s):
                out.append(s)
    return out


def resolve_cols(cols, sums, vecs, keep):
    """剔除折算列（表头含折算词，或与配方列严格成正比的百分比/放大批次列），并记录父列。

    整表只有一个配方、各列都是折算口径（物料比例/固含/250克）时，保留实际投料量最大的一列。"""
    drop = {}
    for j in keep:
        if any(w in cols[j] for w in SCALE_WORDS if w != '原始配方'):
            drop[j] = None
    for j in list(keep):
        if not (95 <= sum(sums.get(j) or [0]) <= 105):
            continue
        for k in keep:
            if k != j and k not in drop and sum(sums.get(k) or [0]) > 120 \
                    and proportional(vecs[j], vecs[k]):
                drop[j] = k
                break
    keep2 = [j for j in keep if j not in drop]
    if not keep2:
        cand = max(cols, key=lambda x: sum(sums.get(x) or [0]))
        keep2 = [cand]
        drop.pop(cand, None)
    parent = {}
    for j in cols:
        if j in keep2:
            continue
        cand = drop.get(j)
        if cand is None:
            cand = next((kk for kk in keep2 if proportional(vecs.get(j, {}), vecs.get(kk, {}))), None)
        if cand is not None:
            parent[j] = cand
            drop.pop(j, None)
    return keep2, parent


def check_totals(g, hi, end, keep, ing_rows, comps):
    """合计行校验：某行数值 ≈ 此前原料行之和 → 视为合计行；逐列比对与组成之和。"""
    tot, tot_row = {}, {}
    for i in range(hi + 1, end):
        part = defaultdict(float)
        for i2 in ing_rows:
            if i2 >= i:
                continue
            for j in keep:
                if j < len(g[i2]) and num(g[i2][j]):
                    part[j] += float(num(g[i2][j]))
        for j in keep:
            v = num(g[i][j]) if j < len(g[i]) else None
            if v and v > 50 and abs(v - part[j]) <= 0.5:
                tot[j] = v
                tot_row[j] = i
    bad = {}
    for name, d in comps.items():
        j = d['列']
        if j not in tot:
            continue  # 该列无合计行可校验
        s = sum(d['组分'].values())
        if abs(s - tot[j]) <= 0.5:
            continue
        tail = sum(float(num(g[i][j])) for i in ing_rows
                   if i > tot_row.get(j, len(g)) and num(g[i][j]))
        if tail and abs(s - tot[j] - tail) <= 0.5:
            continue                       # 合计行之后又补记了原料行，以补记后的和为准
        bad[name] = f'Σ{round(s, 2)} vs 合计{round(tot[j], 2)}'
    return bad, tot


# ---------------------------------------------------------------- 性能块错位检测
PATTERNS = {
    'MEK': re.compile(r'^[<>]?(\d+(?:\.\d+)?)(C|次)?$'),
    'T弯': re.compile(r'^(X|\d+(\.\d+)?mm|\d+-\d+mm|\d+-\d+|点状|零星点状|有改善|0腐蚀|<?\d+mm腐蚀|>\d+mm腐蚀|\d+(\.\d+)?)$'),
    '硬度': re.compile(r'^\d?H$'),
    '刮伤': re.compile(r'^\d+(\.\d+)?$'),
    'BOX': re.compile(r'^(保)?\d+角完好'),
    '固含': re.compile(r'^(0\.\d+|\d+%|\d+\.\d+)$'),
    '粘度': re.compile(r'^\d+(\.\d+)?S$'),
    '滑度': re.compile(r'^0?\.\d+$'),
    '50KG': re.compile(r'^(X|少许.*|合格|不合格|点状|无腐蚀点|个别点状|OK|\d+级?)$'),
    '级': re.compile(r'^\d(-\d)?级?-?$'),
}
ANCHOR_W = {'硬度': 2, '刮伤': 2, 'MEK': 2, 'BOX': 2, '固含': 2, '粘度': 2, 'T弯': 2, '滑度': 2}


def anchor_for(lab):
    s = txt(lab)
    if 'T弯' in s:
        return 'T弯'
    if 'MEK' in s:
        return 'MEK'
    if 'BOX' in s:
        return 'BOX'
    if '佳仪滑度' in s:
        return '滑度'
    if '硬度' in s:
        return '硬度'
    if '刮伤' in s:
        return '刮伤'
    if '固含' in s:
        return '固含'
    if '粘度' in s:
        return '粘度'
    if '50KG' in s or '50Kg' in s:
        return '50KG'
    if re.search(r'煮|盐|蒸汽|H$|S$', s):
        return '级'
    return None


def shift_score(rows, labels, shift):
    """给定错位量下的锚点一致性得分：命中加分、冲突减分（按锚点权重）。"""
    score, hit = 0.0, 0
    for i, _, vals in rows:
        lab = labels.get(i + shift)
        if lab is None:
            continue
        k = anchor_for(lab)
        if not k:
            continue
        vv = [v for v in vals.values() if show(v) != '']
        if not vv:
            continue
        w = ANCHOR_W.get(k, 1)
        ok = sum(1 for v in vv if PATTERNS[k].match(txt(v)))
        score += w * (ok - (len(vv) - ok)) / len(vv)
        hit += w
    return score, hit


def detect_shift(metric_rows, labels):
    """仅当某一错位量明显优于「不错位」时才判为错位（需明显更高的锚点一致性）。"""
    if len(metric_rows) < 5:
        return 0
    base, _ = shift_score(metric_rows, labels, 0)
    best, best_score = 0, base
    for shift in (1, 2, -1, -2):
        score, hit = shift_score(metric_rows, labels, shift)
        if hit >= 5 and score > best_score + 4:
            best, best_score = shift, score
    return best


# ---------------------------------------------------------------- 性能量化
def quantify(rec):
    """矩阵表性能项 → {目标: (值, 备注)}；无法定级者进 无法定级。"""
    got, unq = {}, []
    for lab, vals in rec['性能'].items():
        key = classify(lab)
        if key is None:
            continue
        order = sorted(vals, key=lambda v: 0 if '121' in txt(v) else 1) if key == '水煮' else list(vals)
        v = None
        src = None
        for raw in order:
            v = parse_target_value(key, raw)
            if v is not None:
                src = raw
                break
        if v is None:
            unq.append(f'{lab}={show(vals[0])}')
            continue
        band = v if isinstance(v, str) else None
        if key in got and got[key][0] == v:
            continue
        got[key] = [v, band, f'{lab}：{show(src)}']
    return got, unq


# ---------------------------------------------------------------- 原料名称/SMILES
NAMES = {
    'IR190': '9型环氧树脂36%固含', 'IR809': '环氧树脂55%固含(PR309稀释)', '住友55754G': '住友环氧树脂',
    'RF401': '酚醛固化剂PR401', 'RF160': '酚醛固化剂PR33160G', 'RF516': '酚醛固化剂PR516',
    'RF950': '酚醛固化剂PR8219-50', 'RF956': '酚醛固化剂PR8219-65', 'RH601': '酚醛固化剂SM601RX75',
    '1510蜡': '1510蜡25%工作液', 'AZ088': '分散剂BYK088', '正丁醇': '正丁醇',
    '补加混合液': '乙二醇单丁醚:二甲苯=2:1', '10%磷酸': '磷酸10%水溶液', 'TF100': '乙烯基树脂',
    'TM004': '乙二醇丁醚', 'AS400': '丙烯酸树脂', 'RX170-140': '丙烯酸树脂', '40%50177': '环氧树脂40%固含',
    'IR877': '环氧树脂', 'RJ173M': '聚酯树脂50173-M-40', 'RJ561': '聚酯树脂50561-R-60', 'RY460': '黄色颜料',
    'AC040': '助剂AC040', 'BYK104': '分散剂BYK104', 'IR909': '环氧树脂', 'R170M': '环氧树脂',
    'IR557': '环氧树脂', 'TF022': '乙烯基树脂', 'TM221': '乙二醇丁醚', 'IR868': '环氧树脂',
    'RY075N': '黄色颜料', 'AZ135': '助剂AZ135', '35.7%白浆': '钛白颜料浆35.7%', '14.28%炭黑浆料': '炭黑浆料14.28%',
    '3%气硅': '气相二氧化硅3%', '20%CAB': 'CAB溶液20%', '杜邦-FT960': '环氧树脂', 'AL525': '丙烯酸树脂',
    'AL710': '丙烯酸树脂', 'AZ306': '助剂AZ306', 'AZ551': '流平剂BYK-3550', 'BYK306': '流平剂BYK-306',
    'FL208': '助剂FL208', 'FL208S': '助剂FL208S', 'FL815C': '助剂FL815C', 'IA151': '丙烯酸多元醇',
    'IA893': '丙烯酸树脂', 'IR842': '环氧树脂', 'RA009': '助剂RA009', 'RA083': '助剂RA083',
    'RA824': '助剂RA824', 'RJ183': '聚酯树脂', 'RJ362': '聚酯树脂', '日本151-PVC': 'PVC浆料(日本151)',
    'TZ161': 'PMA溶剂', 'TZ425': 'DBE溶剂', 'TZ240': '醋酸丁酯', 'TT444': '丁酮', 'TT066': '环己酮',
    'TM982': 'PM溶剂', 'TM024': '二乙二醇单丁醚', 'TZ221': '乙二醇丁醚', 'RY078': '黄色颜料',
    'AL800': '丙烯酸树脂', 'IA800': '丙烯酸树脂', 'IA8000': '丙烯酸树脂', 'DBE': 'DBE溶剂',
    'DMP': '混合二元酸酯', 'DPM': '二丙二醇甲醚', 'MIBK': '甲基异丁基酮', '10%135': 'ACA-EAA1 10%稀释液',
    '50173M': '聚酯树脂50173-M-40', '50170M': '聚酯树脂50170-M-52', '40%818': '环氧树脂40%固含',
    'RJ561': '聚酯树脂50561-R-60', '气硅': '气相二氧化硅', '6#炭黑-阿克苏': '炭黑(阿克苏6#)',
    '209-基料': '209基料', '209-白浆': '209白浆', '100#溶剂油': '100号溶剂油(补加)',
    'IR191': '9型环氧树脂', 'IR170': '环氧树脂50%固含', 'RF956': '酚醛固化剂PR8219-65',
    '7096': '助剂7096', 'BYK-306': '流平剂BYK-306', '3%气硅混合料': '气硅混合料3%',
}
SMILES = {
    '正丁醇': 'CCCCO', '二甲苯': 'Cc1ccccc1C', '补加混合液': 'CCCCOCCO', 'TM004': 'CCCCOCCO',
    'TZ240': 'CCCCOC(=O)C', 'TT444': 'CCC(=O)C', 'TT066': 'O=C1CCCCC1', 'TZ161': 'CC(C)OC(=O)C',
    'TZ425': 'COC(=O)CCCC(=O)OC', 'TM982': 'CC(C)OC', 'TM024': 'CCCCOCCOCCO', 'TZ221': 'CCCCOCCO',
    '10%磷酸': 'OP(=O)(O)O', 'MIBK': 'CC(C)CC(=O)C', 'DPM': 'CC(OCC(C)=O)C', 'DMP': 'COC(=O)CCCC(=O)OC',
}
FG_KEYS = ['fg_epoxy', 'fg_oh', 'fg_cooh', 'fg_ester', 'fg_amine', 'fg_amide', 'fg_arom', 'fg_ether']


def series_of(name):
    """配方系列（配方族）：去掉名称尾部的序号段，如 6G-13B-2A→6G-13B、808-4-A→808-4、10#-1→10#。"""
    parts = txt(name).split('-')
    while len(parts) > 1 and re.fullmatch(r'\d*[A-Z#]?\d*', parts[-1]) and parts[-1] != '':
        if re.fullmatch(r'\d+([A-Z]#?|#)?', parts[-1]) or re.fullmatch(r'[A-Z]\d*', parts[-1]):
            parts.pop()
            continue
        break
    return '-'.join(parts) if len(parts) > 1 else txt(name)


# ---------------------------------------------------------------- 汇总记录
def quant_wide(key, raw):
    """宽表性能原始写法 → (数值, 备注)。"""
    v = num(raw)
    if v is not None:
        return v, ''
    s = txt(raw)
    m = re.match(r'^(\d+(?:\.\d+)?)([+＞>])$', s)
    if m:
        note = {'+': f'原始记录{s}（超量程/截尾，按{m.group(1)}计）'}.get(m.group(2), f'原始记录{s}')
        return float(m.group(1)), note
    if key == '水煮':
        m = re.match(r'^(\d+)(?:\.\d+)?$', s)
        if m:
            return float(m.group(1)), ''
    return None, f'原始记录{s}（无法定级）'


def merge_wide(recs806, recs814):
    """8.6 与 8.14 为同批配方的两次汇总：共有配方以 8.14 为准，8.14 缺项回落 8.6；
    T弯 取两份中更细的写法；仅 8.6 记录的（C7 系列）保留。"""
    by814 = {}
    for r in recs814:
        by814.setdefault(r['样本ID'], []).append(r)
    out, used_pair = [], set()

    for r in recs806:
        later = by814.get(r['样本ID'])
        if not later:
            out.append(r)
            continue
        prim = [x for x in later if x['线棒'] == r['线棒']]
        if not prim:                       # 8.14 未保留该线棒号的复测记录 → 保留 8.6 原始记录
            out.append(r)
            continue
        used_pair.add((r['样本ID'], r['线棒']))
        p = prim[0]
        merged = dict(p)
        merged['组分'] = OrderedDict(p['组分']) if p['组分'] else OrderedDict(r['组分'])
        perf = OrderedDict()
        for k in ('T弯', 'MEK', '水煮'):
            a, b = r['性能'].get(k), p['性能'].get(k)
            if b is None:
                if a is not None:
                    perf[k] = a
            else:
                if a is not None and k == 'T弯' and len(txt(a).split('.')[-1]) > len(txt(b).split('.')[-1]):
                    perf[k] = a
                else:
                    perf[k] = b
        merged['性能'] = perf
        merged['批次'] = '8.14'
        merged['日期'] = '2025-08-14'
        merged['配料批次'] = '8.6'
        out.append(merged)
    for r in recs814:
        if (r['样本ID'], r['线棒']) not in used_pair:
            out.append(r)
    return out


def wide_records(files):
    """三份配料测试汇总 → 每配方一条样本 + 各线棒号的性能记录。"""
    per_file = {}
    for batch, kw, date, bt, bm in WIDE_SOURCES:
        per_file[batch] = parse_wide(files[kw], batch, date, (bt, bm))
    recs = per_file['7.26'] + merge_wide(per_file['8.6'], per_file['8.14'])
    samples = OrderedDict()
    for r in recs:
        sid = r['样本ID']
        s = samples.get(sid)
        if s is None:
            s = samples[sid] = {'样本ID': sid, '系列': r['系列'], '体系': SYS_EPOXY, '组分': OrderedDict(r['组分']),
                                '工艺': {'烘烤温度': r['烘烤'][0], '烘烤时间': r['烘烤'][1], '基材': '镀铬铁',
                                         '批次': r.get('配料批次', r['批次']), '线棒号': r['线棒'], '备注': ''},
                                '性能': []}
        elif r['组分'] and r['组分'] != s['组分']:
            s.setdefault('组分冲突', []).append((r['批次'], r['线棒']))
        bt, bm = s['工艺']['烘烤温度'], s['工艺']['烘烤时间']
        cond = f'{bt}℃/{bm}min'
        if r['线棒'] != s['工艺']['线棒号']:
            cond += f'｜{r["线棒"]}线棒'
        for key, raw in r['性能'].items():
            v, note = quant_wide(key, raw)
            tgt = {'T弯': 'T弯', 'MEK': 'MEK擦拭', '水煮': '水煮等级'}[key]
            if v is None:
                s.setdefault('无法定级', []).append(f'{tgt}={show(raw)}')
                continue
            s['性能'].append({'目标': tgt, '测试值': v, '测试条件': cond, '测试日期': r['日期'],
                              '备注': note, '标签状态': '实测', '标签来源': '实验室'})
    return samples, per_file


def blend_provenance(path):
    """8.6「换算明细」：复配配方 → 来源配方与取样量（写入工艺条件备注）。"""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        if '换算明细' not in wb.sheetnames:
            return {}
    except Exception:
        return {}
    ws = wb['换算明细']
    hdr = [txt(c.value) for c in ws[1]]
    ci = {n: i for i, n in enumerate(hdr)}
    out = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        cid, src = txt(row[ci['复配ID']]), txt(row[ci['来源配方']])
        q = num(row[ci['复配取样量(g)']])
        if cid and src and q:
            out[cid].append(f'{src} 取{round(q, 2)}g')
    return {k: '复配自 ' + ' + '.join(v) for k, v in out.items()}


def carryover_series():
    """R8/R9/R10（200℃/10min 批次）：本次原始文件无出处，沿用数据集既有登记。"""
    if not os.path.exists(LEGACY_PKL):
        return OrderedDict()
    D = pickle.load(open(LEGACY_PKL, 'rb'))
    out = OrderedDict()
    for x in D['all_samples']:
        sid = x['样本ID']
        if not re.match(r'^R(8|9|10)-', sid):
            continue
        comp = OrderedDict()
        for k, v in x['组分'].items():
            c = canon(CODE_FIX.get(txt(k), txt(k)))
            comp[c] = round(comp.get(c, 0.0) + float(v), 4)
        s = {'样本ID': sid, '系列': x['系列'], '体系': SYS_EPOXY, '组分': comp,
             '工艺': {'烘烤温度': x.get('烘烤温度') or 200, '烘烤时间': x.get('烘烤时间') or 10,
                      '基材': '镀铬铁', '批次': '7.26', '线棒号': '14#', '备注': ''},
             '性能': []}
        cond = f"{s['工艺']['烘烤温度']}℃/{s['工艺']['烘烤时间']}min"
        for key, tgt in (('T弯', 'T弯'), ('MEK', 'MEK擦拭'), ('水煮', '水煮等级')):
            v = x.get(key)
            if v is None or (isinstance(v, float) and v != v):
                continue
            s['性能'].append({'目标': tgt, '测试值': float(v), '测试条件': cond, '测试日期': '2025-07-26',
                              '备注': '', '标签状态': '实测', '标签来源': '实验室'})
        out[sid] = s
    return out


def matrix_records(files):
    """矩阵表 → 样本记录（跨表同名配方合并/分列）。"""
    defs = {}
    allrec, audit = [], []
    for kw, system in MATRIX_SOURCES:
        recs, aud = parse_matrix_blocks(files[kw], system, defs)
        allrec += recs
        audit += aud
    samples = OrderedDict()
    for r in allrec:
        name = txt(r['样本ID']).lstrip('‘’\'"')
        comp = OrderedDict()
        for k, v in r['组分'].items():
            c = canon(CODE_FIX.get(k, k))
            comp[c] = round(comp.get(c, 0.0) + float(v), 6)
        sid = name
        got, unq = quantify(r)
        bt, bm = 205, 17
        cond = f'{bt}℃/{bm}min'
        base = '镀锡铁' if '镀锡' in r['工艺文本'] else ('镀铬铁' if ('镀铬' in r['工艺文本'] or '铬铁' in r['工艺文本']) else '')
        line = ''
        m = re.search(r'(?:刮棒|线棒|涂布棒)[^\d]{0,3}(\d{1,2}#)', r['工艺文本'])
        if m:
            line = m.group(1)
        elif re.search(r'-?(\d{1,2})#', r['工艺文本']) and '涂' in r['工艺文本']:
            line = re.search(r'(\d{1,2})#', r['工艺文本']).group(1) + '#'
        perf = []
        for key, (v, band, note) in got.items():
            tgt = {'T弯': 'T弯', 'MEK': 'MEK擦拭', '水煮': '水煮等级'}[key]
            perf.append({'目标': tgt, '测试值': v if band is None else band, '测试条件': cond,
                         '测试日期': r['日期'], '备注': f'原始记录 {note}', '标签状态': '实测',
                         '标签来源': '实验室', 'band': band})
        proc_note = r['工艺文本'][:180]
        if sid in samples:
            old = samples[sid]
            if all(abs(old['组分'].get(k, -1) - v) < 1e-4 for k, v in comp.items()) \
                    and len(old['组分']) == len(comp):
                for p in perf:
                    p['测试条件'] += f'｜{r["sheet"]}批次'
                    old['性能'].append(p)
                if old['工艺']['备注'] and proc_note not in old['工艺']['备注']:
                    old['工艺']['备注'] += ' / ' + proc_note
                old['工艺']['批次'] = old['工艺']['批次'] if old['工艺']['批次'] == r['sheet'] else '多批'
                continue
            sid = f'{name}@{r["sheet"]}'
        s = {'样本ID': sid, '系列': series_of(name),
             '体系': r['体系'], '组分': comp,
             '工艺': {'烘烤温度': bt, '烘烤时间': bm, '基材': base, '批次': r['sheet'],
                      '线棒号': line, '备注': proc_note},
             '性能': perf, '无法定级': unq}
        samples[sid] = s
    return samples, audit


# ---------------------------------------------------------------- 原料库与描述符
def mat_name(code):
    comp = COMPO.get(code)
    if comp and comp[0]:
        return str(comp[0]).split('（')[0].strip()
    return NAMES.get(code, code)


def build_materials(samples):
    used = defaultdict(set)
    for s in samples.values():
        for c in s['组分']:
            used[c].add(s['体系'])
    full = {}
    for c in sorted(used):
        d = dict(MAT[c]) if c in MAT else est_material(c)
        if '数据来源' not in d:
            d['数据来源'] = '类别典型值(工作台估算登记)'
        if c in SELF_EVIDENT:                    # 代码名已指明类别与到货状态
            d.update(SELF_EVIDENT[c])
            d['数据来源'] = 'handbook:名称自证'
        full[c] = d
    for k in list(full):
        for role in ('role', 'rtype'):
            full[k].setdefault(role, '助剂' if role == 'role' else '其他')
    changed, merge, pending = HF.apply(full)
    for c in merge:
        full.pop(c, None)
        for s in samples.values():
            if c in s['组分']:
                tgt = merge[c]
                if tgt not in full:
                    full[tgt] = dict(MAT[tgt]) if tgt in MAT else est_material(tgt)
                s['组分'][tgt] = round(s['组分'].pop(c) + s['组分'].get(tgt, 0.0), 4)
    tds_changed, tds_prov = tds_sds.apply(full)
    used2 = defaultdict(set)
    for s in samples.values():
        for c in s['组分']:
            used2[c].add(s['体系'])
    for c in sorted(set(used2) - set(full)):
        full[c] = est_material(c)
        full[c]['数据来源'] = '类别典型值(工作台估算登记)'
    return full, used2, pending


def systems_of(codes):
    ss = sorted(codes)
    if len(ss) == 1:
        return ss[0]
    if set(ss) == {SYS_EPOXY, SYS_PEIBI, SYS_JINHUANG} or len(ss) > 1:
        return '通用'
    return ss[0]


def descriptor_row(comp, mat):
    """配方级描述符：与模板内置公式同式（加权按质量分数、官能团密度按每 100g 配方摩尔数）。"""
    items = [(c, float(a)) for c, a in comp.items() if a and float(a) > 0 and c in mat]
    total = sum(a for _, a in items)
    if total <= 0:
        return None
    roles = ['树脂', '固化剂', '溶剂', '助剂', '颜料']
    rts = ['环氧', '酚醛', '聚酯', '乙烯基', '丙烯酸', '聚氨酯']
    rf = {r: 0.0 for r in roles}
    tf = {r: 0.0 for r in rts}
    for c, a in items:
        rf[mat[c]['role']] += a / total
        if mat[c]['rtype'] in tf:
            tf[mat[c]['rtype']] += a / total
    d = OrderedDict()
    d['总用量(g)'] = total
    d['组分数量'] = len(items)
    for r in roles:
        d[{'树脂': '树脂占比', '固化剂': '固化剂占比', '溶剂': '溶剂占比',
           '助剂': '助剂占比', '颜料': '颜料占比'}[r]] = rf[r]
    d['固化剂/树脂比'] = rf['固化剂'] / rf['树脂'] if rf['树脂'] > 0 else 0.0
    for r in rts:
        d[f'{r}树脂占比'] = tf[r]
    wkeys = CONT_DESC[:22]
    for k in wkeys:
        d[{'NV': '加权固含(%)', 'density': '加权密度', 'Mw': '加权分子量', 'EEW': '加权EEW',
           'AV': '加权酸值', 'OHV': '加权羟值', 'amine': '加权胺值', 'func': '加权官能度',
           'Tg': '加权Tg', 'bp': '加权沸点', 'fp': '加权闪点', 'dD': '加权δD', 'dP': '加权δP',
           'dH': '加权δH', 'pol': '加权极性', 'evap': '加权挥发速率', 'C': '加权C(%)',
           'H': '加权H(%)', 'O': '加权O(%)', 'N': '加权N(%)', 'S': '加权S(%)',
           'Cl': '加权Cl(%)'}[k]] = sum((num(mat[c][k]) or 0.0) * a / total for c, a in items)
    fgname = {'fg_epoxy': '环氧基密度(mol/100g)', 'fg_oh': '羟基密度', 'fg_cooh': '羧基密度',
              'fg_ester': '酯基密度', 'fg_amine': '胺基密度', 'fg_amide': '酰胺密度',
              'fg_arom': '芳香环密度', 'fg_ether': '醚键密度'}
    for k in FG_KEYS:
        d[fgname[k]] = sum((num(mat[c][k]) or 0.0) * a for c, a in items) / 100.0
    d['蜡含量(%)'] = sum((num(mat[c]['wax']) or 0.0) * a / total for c, a in items)
    d['颜料含量(%)'] = sum((num(mat[c]['pig']) or 0.0) * a / total for c, a in items)
    return d


# ---------------------------------------------------------------- 写出工作簿
HDR_FILL = None
DET_HEADERS = ['样本ID', '系列', '体系', '原料代码', '用量(g)', '角色', '树脂类型',
               '固含贡献', '密度贡献', '分子量贡献', 'EEW贡献', '酸值贡献', '羟值贡献', '胺值贡献',
               '官能度贡献', 'Tg贡献', '沸点贡献', '闪点贡献', 'δD贡献', 'δP贡献', 'δH贡献',
               '极性贡献', '挥发速率贡献', 'C贡献', 'H贡献', 'O贡献', 'N贡献', 'S贡献', 'Cl贡献',
               '环氧基贡献', '羟基贡献', '羧基贡献', '酯基贡献', '胺基贡献', '酰胺贡献',
               '芳香环贡献', '醚键贡献', '蜡贡献', '颜料贡献']
MAT_HEADERS = ['原料代码', '原料名称', '所属体系', '角色', '树脂类型', 'SMILES', '描述符状态',
               '固含NV(%)', '密度(g/cm³)', '分子量(g/mol)', '环氧当量EEW(g/eq)', '酸值AV(mgKOH/g)',
               '羟值OHV(mgKOH/g)', '胺值(mgKOH/g)', '官能度', 'Tg(℃)', '沸点(℃)', '闪点(℃)',
               'Hansen δD', 'Hansen δP', 'Hansen δH', '极性指数', '相对挥发速率', 'C(%)', 'H(%)',
               'O(%)', 'N(%)', 'S(%)', 'Cl(%)', '环氧基(mol/100g)', '羟基(mol/100g)', '羧基(mol/100g)',
               '酯基(mol/100g)', '胺基(mol/100g)', '酰胺(mol/100g)', '芳香环(mol/100g)', '醚键(mol/100g)',
               '蜡含量(%)', '颜料含量(%)', '数据来源', '备注']
def as_date(v):
    """'YYYY-MM-DD' → 日期单元格（避免日期以文本形式存储）。"""
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})$', txt(v))
    if not m:
        return v or None
    import datetime
    return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))


PERF_HEADERS = ['样本ID', '体系', '目标属性', '测试值', '单位', '标签状态', '标签来源', '不确定性',
                '测试条件', '测试日期', '备注']
PROC_HEADERS = ['样本ID', '体系', '烘烤温度(℃)', '烘烤时间(min)', '膜厚(g/m²)', '基材', '批次',
                '线棒号', '备注']
DICT_ROWS = [
    ('样本ID', '配方明细/性能结果/工艺条件/配方级描述符/建模输入', '文本', '-', '全局唯一，关联各表', 'D1-1'),
    ('系列', '配方明细/配方级描述符/建模输入', '文本', '-', '配方系列/家族，用于系列目标编码建模', 'D1'),
    ('体系', '体系配置/配方明细', '文本', '-', '化学体系类别，需在体系配置登记', '环氧酚醛'),
    ('固化机制', '体系配置', '文本', '-', '体系交联反应机制', '环氧-酚醛缩合'),
    ('目标属性', '体系配置/性能结果', '文本', '-', '性能名称，需在体系配置登记', 'T弯'),
    ('方向', '体系配置', '枚举', '-', '越低越好/越高越好，决定建模目标方向', '越低越好'),
    ('数据类型', '体系配置', '枚举', '-', '连续/计数/等级/分类，决定回归或分类建模', '连续'),
    ('原料代码', '原料主数据/配方明细', '文本', '-', '原料唯一编码，与原料主数据一致；未登记代码标红', 'IR190'),
    ('用量', '配方明细', '数值', 'g', '该组分在样本中的质量份（原始记录口径）', '66.0'),
    ('角色', '原料主数据/配方明细', '枚举', '-', '树脂/固化剂/溶剂/助剂/颜料', '树脂'),
    ('树脂类型', '原料主数据/配方明细', '枚举', '-', '环氧/酚醛/聚酯/乙烯基/丙烯酸/聚氨酯/氨基/其他', '环氧'),
    ('SMILES', '原料主数据', '文本', '-', '结构明确的原料给出结构式，支撑分子描述符计算', 'CCCCO'),
    ('描述符状态', '原料主数据', '枚举', '-', '已计算/专有估算', '已计算'),
    ('固含NV', '原料主数据', '数值', '%', '按到货状态的固体含量', '36'),
    ('环氧当量EEW', '原料主数据', '数值', 'g/eq', '含 1mol 环氧基的到货产品质量', '2640'),
    ('羟值OHV', '原料主数据', '数值', 'mgKOH/g', '中和 1g 样品羟基所需 KOH 毫克数', '30'),
    ('官能团密度', '原料主数据', '数值', 'mol/100g', '每 100g 到货产品的官能团摩尔数', '0.038'),
    ('数据来源', '原料主数据', '文本', '-', '描述符出处：TDS/SDS 档案实测、TDS 同族推断、送检组成、手册值、估算', 'TDS/SDS'),
    ('各「贡献」列', '配方明细', '数值', '-', '用量 × 原料描述符，供配方级聚合', '2.508'),
    ('总用量/组分数量', '配方级描述符', '数值', 'g/-', '样本组分数与质量份之和', '74.9'),
    ('树脂占比等', '配方级描述符', '数值', '-', '各角色质量分数；固化剂/树脂比 = 固化剂占比/树脂占比', '0.88'),
    ('加权固含等', '配方级描述符', '数值', '-', 'Σ(用量×描述符)/总用量', '40.2'),
    ('环氧基密度等', '配方级描述符', '数值', 'mol/100g', 'Σ(用量×官能团密度)/100', '0.035'),
    ('测试值', '性能结果', '数值', '-', '性能测试结果数值（原始写法见备注）', '17.415'),
    ('T弯/MEK/水煮实测', '建模输入', '数值', '-', '一行=一个样本，三个目标实测值并列；无实测记为空', '17.415'),
    ('T弯/MEK/水煮预测', '建模输入', '数值', '-', '工作台回写的预测值与对应不确定性（录入时为空）', '18.6'),
    ('标签状态', '性能结果/建模输入', '枚举', '-', '实测/伪标签/推荐测试/人工复核；无对应测试值即无标签', '实测'),
    ('标签来源', '性能结果', '文本', '-', '实验室/模型预测/主动学习/人工复核', '实验室'),
    ('不确定性', '性能结果/建模输入', '数值', '-', '模型预测的树间标准差（实测记录留空）', '0.42'),
    ('测试条件', '性能结果', '文本', '-', '烘烤制度；同配方多线棒/多批次时在此标注', '205℃/17min｜10#线棒'),
    ('测试日期', '性能结果', '日期', 'yyyy-mm-dd', '该条性能记录的测试日期，按日期格式存储', '2025-08-14'),
    ('烘烤温度/时间', '工艺条件', '数值', '℃/min', '固化工艺参数；原始记录未给出的按 NaN 处理', '205/17'),
    ('基材/膜厚/批次/线棒号', '工艺条件', '文本', '-', '试涂基材、膜厚、配料批次、涂布线棒', '镀铬铁/8.14/14#'),
    ('备注', '性能结果/工艺条件', '文本', '-', '性能原始写法（区间、截尾、腐蚀档位等）与工艺原文', '原始记录 >15mm腐蚀'),
]


PKL = os.path.join(HERE, '..', 'data', 'merged_data.pkl')


def write_payload(samples, mat, path):
    """同步 data/merged_data.pkl（合并版数据集与实验脚本共用的中间产物，沿用既有 schema）。"""
    all_samples, lab, unlab = [], [], []
    for s in samples.values():
        prim = {}
        for pr in s['性能']:
            if not isinstance(pr['测试值'], str):
                prim.setdefault(pr['目标'], pr['测试值'])
        row = {'样本ID': s['样本ID'], '体系': s['体系'], '系列': s['系列'], '组分': dict(s['组分']),
               '烘烤温度': s['工艺']['烘烤温度'], '烘烤时间': s['工艺']['烘烤时间'],
               'T弯': prim.get('T弯'), 'MEK': prim.get('MEK擦拭'), '水煮': prim.get('水煮等级'),
               '标签状态': '实测' if prim else '无标签', '来源': f"{s['体系']}·{s['工艺']['批次']}"}
        all_samples.append(row)
        (lab if prim else unlab).append(row)
    import pandas as pd
    D = {'full_mat': mat, 'new_mats': [c for c in mat if c not in MAT],
         'lab_samples': lab, 'unlab_samples': unlab, 'all_samples': all_samples,
         'desc_df': pd.DataFrame([{'样本ID': r['样本ID'], '体系': r['体系']} for r in all_samples])}
    pickle.dump(D, open(path, 'wb'))
    return len(all_samples), len(lab), len(unlab)


def write_workbook(samples, mat, used, out_path, stats):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import FormulaRule

    HDR_FILL = PatternFill('solid', fgColor='1F2937')
    HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    ZEBRA_1 = PatternFill('solid', fgColor='FFFFFF')
    ZEBRA_2 = PatternFill('solid', fgColor='F7F9FC')
    KPI_FILL = PatternFill('solid', fgColor='EAF2FF')
    GREEN_FILL = PatternFill('solid', fgColor='E3F5EA')
    RED_FILL = PatternFill('solid', fgColor='FDE8E8')
    THIN = Side(style='thin', color='D9DEE7')
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    BODY_FONT = Font(name='Arial', size=10)
    BOLD_FONT = Font(name='Arial', size=10, bold=True)
    TITLE_FONT = Font(name='Arial', size=14, bold=True, color='1F2937')

    def style_table(ws, n_cols, n_rows, kpi_cols=None, status_col=None, header_row=1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(header_row, c)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = BORDER
        for r in range(header_row + 1, header_row + n_rows + 1):
            fill = ZEBRA_1 if (r - header_row) % 2 == 0 else ZEBRA_2
            for c in range(1, n_cols + 1):
                cell = ws.cell(r, c)
                cell.fill = fill
                cell.font = BODY_FONT
                cell.border = BORDER
                cell.alignment = Alignment(vertical='center')
        if kpi_cols:
            for c in kpi_cols:
                for r in range(header_row + 1, header_row + n_rows + 1):
                    ws.cell(r, c).fill = KPI_FILL
                    ws.cell(r, c).font = BOLD_FONT
        if status_col:
            for r in range(header_row + 1, header_row + n_rows + 1):
                if ws.cell(r, status_col).value == '实测':
                    ws.cell(r, status_col).fill = GREEN_FILL

    wb = Workbook()
    sys_names = list(SYSTEM_CONFIG)
    all_names = sys_names + ['通用']
    tgt_names = sorted({t for _, _, ts in SYSTEM_CONFIG.values() for t, *_ in ts})

    # 使用说明
    ws = wb.active
    ws.title = '使用说明'
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 126
    lines = [
        '终极版涂料配方-性能数据集模板 v3',
        '',
        '一、模板定位',
        '统一承载不同化学体系（环氧酚醛、环氧配比方案、聚酯金黄）的配方、性能与工艺数据，',
        '以「原料描述符」替代「原料编码」作为建模特征，使模型具备跨体系、跨新组分的泛化能力。',
        '',
        '二、数据构成',
        f'配方样本 {stats["n_samples"]} 条（' + '，'.join(f'{k} {v} 条' for k, v in stats['by_system'].items()) + '）。',
        f'组分记录 {stats["n_det"]} 行；性能实测 {stats["n_perf"]} 条'
        f'（T弯 {stats["by_target"].get("T弯", 0)}、MEK擦拭 {stats["by_target"].get("MEK擦拭", 0)}、'
        f'水煮等级 {stats["by_target"].get("水煮等级", 0)}）；工艺条件 {stats["n_proc"]} 条；原料主数据 {stats["n_mat"]} 种。',
        f'有实测标签 {stats["n_lab"]} 条，无实测标签 {stats["n_unlab"]} 条。',
        '',
        '三、数据来源',
        '1. 7.26配料测试汇总：R01–R7 系列 106 个配方，烘烤 200℃/10min。',
        '2. 8.6配料测试汇总（含换算明细、测试原始数据）：C7 系列 8 个配方，烘烤 205℃/17min。',
        '3. 8.14配料测试汇总（含测试原始数据）：D1–D7/C4–C6 系列 175 个配方，与 8.6 同一批配方，以更晚的 8.14 为准（补齐 8.6 尚未出结果的水煮等级），T弯取两份记录中更细的写法。',
        '4. 3NX240913-6C--AI研发26.7.22配比方案：环氧配比方案体系配方，性能项取 T弯G 冲击-5%硫酸铜'
        '腐蚀判定；该表的水煮/杀菌项为定性记录（泛白、合格、-），未定级。',
        '5. 聚酯金黄-AI(1)：聚酯金黄体系配方，性能项取 T弯 / MEK / 121℃*60min 水煮。',
        '6. R8/R9/R10 系列 56 个配方（烘烤 200℃/10min）：沿用数据集既有登记，本次原始文件未含该批记录。',
        '',
        '四、工作表结构',
        '1. 体系配置：登记体系、固化机制、目标属性（单位/方向/数据类型）。',
        '2. 原料主数据：登记原料描述符（SMILES、固含/当量/官能团密度等）与描述符数据来源。',
        '3. 配方明细：长格式（每行 = 一个样本中的一个组分），含「系列」列用于系列编码建模。',
        '4. 性能结果：每条测试记录一行，含测试条件/测试日期，原始写法记在备注。',
        '5. 工艺条件：烘烤温度/时间/膜厚/基材/批次/线棒号。',
        '6. 配方级描述符：由配方明细 + 原料主数据聚合的样本级特征矩阵。',
        '7. 建模输入：宽表，一行 = 一个样本，三个目标的实测值与特征并列（预测值/不确定性由工作台回写）。',
        '8. 数据字典：全部字段的口径说明。',
        '',
        '五、录入与口径规则',
        '1. 原料代码必须与「原料主数据」一致且唯一；样本ID全局唯一；未登记的原料代码在配方明细中标红。',
        '2. 用量统一为质量份(g)，按原始记录口径；矩阵表的百分比/1000KG/500克等折算列不另立样本，其上的性能记录归入对应配方。',
        '3. 同一配方的 10#/14#/18# 线棒试涂：组成分一条样本，各线棒的测试值分别成行，'
        '在「测试条件」标注线棒号。',
        '4. 性能取值沿用实验室口径：字段带 MEK/T弯/水煮 即视为同一种测试；T弯G 冲击-5%硫酸铜腐蚀判定'
        '按档位定档（0腐蚀→0、10mm→10、15mm→15，开区间 <10mm 取 5~8、>15mm 取 16~20，'
        '开区间内按配方交联密度定档），点状/有改善等无法定级者留空，原始写法记入备注。',
        '5. MEK 记录值恰为 300 视为右截尾（真实值 ≥300）；水煮等级 1（最好）~4（最差），'
        '记作 4+ 及以上者并入 4，均在备注保留原始写法。',
        '6. 原始记录未给出的烘烤条件按空值处理（未记录≠零固化）。',
        '7. 体系名称必须在「体系配置」中登记；新增配方按「样本ID + 系列 + 原料代码 + 用量」逐行录入，'
        '性能与工艺以样本ID关联。',
        '',
        '六、重新生成',
        'python scripts/reingest_template.py            由 raw/ 下的 5 份原始文件重新生成本文件。',
        'python scripts/reingest_template.py --report   只打印解析审计（配方列识别、合计行校验、错位检测）。',
    ]
    for r, text in enumerate(lines, start=1):
        cell = ws.cell(r, 2, text)
        if r == 1:
            cell.font = TITLE_FONT
        elif text[:2] in ('一、', '二、', '三、', '四、', '五、', '六、'):
            cell.font = BOLD_FONT
        else:
            cell.font = BODY_FONT
        cell.alignment = Alignment(vertical='center', wrap_text=True)

    # 体系配置
    ws = wb.create_sheet('体系配置')
    sys_headers = ['体系名称', '固化机制', '典型树脂类型', '目标属性', '单位', '方向', '数据类型', '适用标准/说明']
    ws.append(sys_headers)
    rows = []
    for name, (mech, rtype, ts) in SYSTEM_CONFIG.items():
        for t in ts:
            rows.append([name, mech, rtype, *t])
    for row in rows:
        ws.append(row)
    style_table(ws, len(sys_headers), len(rows), kpi_cols=[1, 4, 5, 6, 7])
    for i, w in enumerate([14, 20, 30, 14, 8, 12, 10, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    # 原料主数据
    ws = wb.create_sheet('原料主数据')
    ws.append(MAT_HEADERS)
    codes = sorted(mat, key=lambda c: (systems_of(used.get(c, {'通用'})), c))
    for code in codes:
        d = mat[code]
        smi = SMILES.get(code, '')
        status = d.get('描述符状态') or ('已计算' if smi else '专有估算')
        row = [code, mat_name(code), systems_of(used.get(code, {'通用'})), d['role'], d['rtype'],
               smi or None, status]
        for k in CONT_DESC:
            v = num(d.get(k))
            row.append(round(v, 6) if v is not None else None)
        note = None
        if d.get('TDS档案'):
            note = f"TDS档案:{d['TDS档案']}" if isinstance(d['TDS档案'], str) else 'TDS/SDS 档案实测'
        elif d.get('prov'):
            note = '描述符出处见工作台 tds_sds/compo_rules'
        row += [d.get('数据来源', '类别典型值'), note]
        ws.append(row)
    n_mat = len(codes)
    style_table(ws, len(MAT_HEADERS), n_mat, kpi_cols=[8, 11, 12, 13, 14, 15, 30, 31, 32, 33, 34, 35, 36, 37])
    for i, w in enumerate([16, 24, 12, 8, 9, 18, 10] + [10] * 32 + [22, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    for dv, colref in ((DataValidation(type='list', formula1=f'"{",".join(all_names)}"', allow_blank=True), 'C'),
                       (DataValidation(type='list', formula1='"树脂,固化剂,溶剂,助剂,颜料"', allow_blank=True), 'D'),
                       (DataValidation(type='list', formula1='"环氧,酚醛,聚酯,乙烯基,丙烯酸,聚氨酯,氨基,其他"', allow_blank=True), 'E')):
        ws.add_data_validation(dv)
        dv.add(f'{colref}2:{colref}1000')

    # 配方明细
    ws = wb.create_sheet('配方明细')
    ws.append(DET_HEADERS)
    det_rows = 0
    for s in samples.values():
        for code, amt in s['组分'].items():
            d = mat.get(code)
            if d is None or amt <= 0:
                continue
            row = [s['样本ID'], s['系列'], s['体系'], code, round(float(amt), 4), d['role'], d['rtype']]
            for k in CONT_DESC:
                v = num(d.get(k)) or 0.0
                row.append(round(float(amt) * v, 6))
            ws.append(row)
            det_rows += 1
    style_table(ws, len(DET_HEADERS), det_rows, kpi_cols=[5])
    for i, w in enumerate([16, 12, 12, 16, 10, 9, 9] + [10] * 32, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.conditional_formatting.add(
        f'D2:D{det_rows + 1}',
        FormulaRule(formula=['ISNA(MATCH($D2,\'原料主数据\'!$A:$A,0))'], fill=RED_FILL, stopIfTrue=True))
    dv = DataValidation(type='list', formula1=f'"{",".join(sys_names)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f'C2:C{max(det_rows + 200, 1000)}')

    # 性能结果
    ws = wb.create_sheet('性能结果')
    ws.append(PERF_HEADERS)
    n_perf = 0
    for s in samples.values():
        for p in sorted(s['性能'], key=lambda x: (TARGETS.index(x['目标']), x['测试条件'])):
            v = p['测试值']
            if isinstance(v, str):
                continue
            ws.append([s['样本ID'], s['体系'], p['目标'], round(float(v), 4), UNIT[p['目标']],
                       p['标签状态'], p['标签来源'], None, p['测试条件'], as_date(p['测试日期']),
                       p.get('备注') or None])
            n_perf += 1
    style_table(ws, len(PERF_HEADERS), n_perf, kpi_cols=[4], status_col=6)
    for i, w in enumerate([16, 12, 12, 10, 8, 10, 10, 10, 26, 12, 46], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    for dv, colref in ((DataValidation(type='list', formula1=f'"{",".join(sys_names)}"', allow_blank=True), 'B'),
                       (DataValidation(type='list', formula1=f'"{",".join(tgt_names)}"', allow_blank=True), 'C'),
                       (DataValidation(type='list', formula1='"实测,伪标签,推荐测试,人工复核"', allow_blank=True), 'F')):
        ws.add_data_validation(dv)
        dv.add(f'{colref}2:{colref}{max(n_perf + 200, 1000)}')

    # 工艺条件
    ws = wb.create_sheet('工艺条件')
    ws.append(PROC_HEADERS)
    for s in samples.values():
        p = s['工艺']
        ws.append([s['样本ID'], s['体系'], p['烘烤温度'], p['烘烤时间'], p.get('膜厚') or None,
                   p.get('基材') or None, p['批次'], p.get('线棒号') or None, p.get('备注') or None])
    style_table(ws, len(PROC_HEADERS), len(samples))
    for i, w in enumerate([16, 12, 12, 12, 10, 10, 12, 9, 60], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    dv = DataValidation(type='list', formula1=f'"{",".join(sys_names)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f'B2:B{max(len(samples) + 200, 1000)}')

    # 配方级描述符
    ws = wb.create_sheet('配方级描述符')
    first = next(d for d in (descriptor_row(x['组分'], mat) for x in samples.values()) if d)
    fd_headers = ['样本ID', '系列', '体系'] + list(first.keys())
    ws.append(fd_headers)
    desc = {}
    for s in samples.values():
        d = descriptor_row(s['组分'], mat)
        desc[s['样本ID']] = d
        if d is None:
            continue
        ws.append([s['样本ID'], s['系列'], s['体系']] + [round(v, 6) if isinstance(v, float) else v
                                                         for v in d.values()])
    style_table(ws, len(fd_headers), len(samples), kpi_cols=[4, 6, 7, 18, 21, 22, 23])
    for i, w in enumerate([16, 12, 12] + [12] * (len(fd_headers) - 3), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    # 建模输入（宽表：一行 = 一个样本，三个目标的实测/预测/不确定性并列）
    ws = wb.create_sheet('建模输入')
    short = {'T弯': 'T弯', 'MEK擦拭': 'MEK', '水煮等级': '水煮'}
    mi_headers = (['样本ID', '系列', '体系', '标签状态']
                  + [f'{short[t]}实测' for t in TARGETS] + [f'{short[t]}预测' for t in TARGETS]
                  + [f'{short[t]}不确定性' for t in TARGETS] + fd_headers[3:])
    ws.append(mi_headers)
    n_mi = 0
    for s in samples.values():
        d = desc.get(s['样本ID'])
        if d is None:
            continue
        have = {}
        for pr in s['性能']:                       # 同目标多条记录（不同线棒/批次）取首条（14#线棒）
            if not isinstance(pr['测试值'], str):
                have.setdefault(pr['目标'], pr['测试值'])
        ws.append([s['样本ID'], s['系列'], s['体系'], '实测' if have else '无标签']
                  + [round(float(have[t]), 4) if t in have else None for t in TARGETS]
                  + [None] * (len(TARGETS) * 2)   # 预测值/不确定性由工作台回写
                  + [round(x, 6) if isinstance(x, float) else x for x in d.values()])
        n_mi += 1
    style_table(ws, len(mi_headers), n_mi, kpi_cols=[5, 6, 7], status_col=4)
    for i, w in enumerate([16, 12, 12, 10] + [11] * (len(mi_headers) - 4), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'E2'

    # 数据字典
    ws = wb.create_sheet('数据字典')
    ws.append(['字段名', '所属工作表', '类型', '单位', '口径说明', '示例'])
    for row in DICT_ROWS:
        ws.append(list(row))
    style_table(ws, 6, len(DICT_ROWS))
    for i, w in enumerate([28, 40, 8, 12, 64, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    # 统一显示格式：文本列按文本、计数列整数、数值列最多 6 位小数（不截断有效位）
    TEXT_COLS = {'原料主数据': [1, 2, 3, 4, 5, 6, 7, 40, 41],
                 '配方明细': [1, 2, 3, 4, 6, 7],
                 '性能结果': [1, 2, 3, 5, 6, 7, 9, 11],
                 '工艺条件': [1, 2, 6, 7, 8, 9],
                 '配方级描述符': [1, 2, 3],
                 '建模输入': [1, 2, 3, 4],
                 '体系配置': [1, 2, 3, 4, 5, 6, 7, 8],
                 '数据字典': [1, 2, 3, 4, 5, 6]}
    INT_COLS = {'配方级描述符': [5], '建模输入': [15]}
    for name, tcols in TEXT_COLS.items():
        sh = wb[name]
        for r in range(2, sh.max_row + 1):
            for c in range(1, sh.max_column + 1):
                cell = sh.cell(r, c)
                if cell.value is None:
                    continue
                if c in tcols:
                    cell.number_format = '@'
                elif c in INT_COLS.get(name, []):
                    cell.number_format = '0'
                elif name == '性能结果' and c == 10:
                    cell.number_format = 'yyyy-mm-dd'
                else:
                    cell.number_format = '0.######'
    wb.save(out_path)
    return {'配方明细': det_rows, '性能结果': n_perf, '工艺条件': len(samples),
            '配方级描述符': len(samples), '建模输入': n_mi, '原料主数据': n_mat}


# ---------------------------------------------------------------- 主流程
def main():
    report = '--report' in sys.argv
    files = {}
    for _, kw, _, _, _ in WIDE_SOURCES:
        files[kw] = find_source(kw)
    for kw, _ in MATRIX_SOURCES:
        files[kw] = find_source(kw)
    print('原始文件：')
    for k, v in files.items():
        print(f'  {k:20s} {os.path.basename(v)}')
    build_vocab([files[kw] for kw, _ in MATRIX_SOURCES])

    wide, per_file = wide_records(files)
    print(f'\n配料测试汇总：7.26={len(per_file["7.26"])} 行 / 8.6={len(per_file["8.6"])} 行 / '
          f'8.14={len(per_file["8.14"])} 行 → 合并后配方 {len(wide)} 条')
    carry = carryover_series()
    notes = blend_provenance(files['8.6配料测试汇总'])
    for sid, txt_ in notes.items():
        if sid in wide:
            wide[sid]['工艺']['备注'] = (txt_ + (' / ' if wide[sid]['工艺']['备注'] else '') + wide[sid]['工艺']['备注'])[:180]
    samples = OrderedDict()
    for sid, s in wide.items():
        samples[sid] = s
    for sid, s in carry.items():
        samples.setdefault(sid, s)
    print(f'R8/R9/R10 沿用既有登记：{len(carry)} 条')

    mrec, audit = matrix_records(files)
    for sid, s in mrec.items():
        samples[sid] = s
    cblocks = [b for a in audit for b in a['块'] if b['原料行']]
    cov = sum(int(str(b['合计可校验列']).split('/')[0]) for b in cblocks)
    cov_all = sum(int(str(b['合计可校验列']).split('/')[1]) for b in cblocks)
    print(f'矩阵表解析：配方 {len(mrec)} 条（配比方案+聚酯金黄）｜{cov}/{cov_all} 条配方列通过合计行校验')
    for a in audit:
        for b in a['块']:
            if b['错位取标']:
                d = '上' if b['错位取标'] > 0 else '下'
                print(f"  [错位] {a['sheet']} 名称行{b['名称行']}: 性能数值相对标签整块{d}移 {abs(b['错位取标'])} 行，已按标签语义取回")
            if b['合计差异']:
                print(f"  [合计不符] {a['sheet']} {b['合计差异']}")
        if a.get('仅性能无组成'):
            print(f"  [有性能无组成，未录入] {a['sheet']}: {a['仅性能无组成']}")

    for s in samples.values():                     # 用量按 4 位小数入账，贡献/聚合列与之自洽
        s['组分'] = OrderedDict((c, round(v, 4)) for c, v in sorted(s['组分'].items()))
    mat, used, pending = build_materials(samples)
    bands = [s for s in samples.values() if any(isinstance(p['测试值'], str) for p in s['性能'])]
    if bands:
        tmp = [{'样本ID': s['样本ID'], '组分': s['组分'],
                'T弯': next(p['测试值'] for p in s['性能'] if isinstance(p['测试值'], str))} for s in bands]
        _assign_bands(tmp, mat)
        for s, t in zip(bands, tmp):
            for p in s['性能']:
                if isinstance(p['测试值'], str):
                    p['备注'] = (p['备注'] + f'；档位{p["测试值"]}按配方交联密度定档').strip('；')
                    p['测试值'] = t['T弯']
    n_perf = sum(1 for s in samples.values() for p in s['性能'] if not isinstance(p['测试值'], str))
    stats = {'n_samples': len(samples),
             'n_mat': len(mat),
             'n_lab': sum(1 for s in samples.values() if s['性能']),
             'n_unlab': sum(1 for s in samples.values() if not s['性能']),
             'by_system': {sys_: sum(1 for s in samples.values() if s['体系'] == sys_)
                           for sys_ in SYSTEM_CONFIG},
             'n_det': sum(len(s['组分']) for s in samples.values()),
             'n_perf': n_perf,
             'n_proc': len(samples),
             'by_target': {t: sum(1 for s in samples.values() for p in s['性能']
                                  if p['目标'] == t and not isinstance(p['测试值'], str))
                           for t in TARGETS}}
    print(f'\n汇总：样本 {stats["n_samples"]} 条 {stats["by_system"]}｜组分 {stats["n_det"]} 行｜'
          f'性能 {stats["n_perf"]} 条 {stats["by_target"]}｜原料 {len(mat)} 种')
    unl = [s['样本ID'] for s in samples.values() if not s['性能']]
    print(f'无实测标签样本：{len(unl)} 条')
    if report:
        return
    written = write_workbook(samples, mat, used, OUT, stats)
    print('已写出', OUT)
    print('  各表行数:', written)
    ns, nl, nu = write_payload(samples, mat, PKL)
    print(f'已写出 {os.path.relpath(PKL, ROOT)}（样本 {ns}｜实测 {nl}｜无标签 {nu}）')


if __name__ == '__main__':
    main()
